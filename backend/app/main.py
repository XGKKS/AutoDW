import asyncio
import base64
import csv
import hashlib
import io
import json
import re
import os
import logging
import traceback
import uuid
import threading
from functools import partial
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from fastapi import FastAPI, UploadFile, File, HTTPException, Body, Form, Request
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import requests
import openpyxl
from docx import Document
from concurrent.futures import ThreadPoolExecutor, as_completed

executor = ThreadPoolExecutor(max_workers=10)

cache_lock = threading.Lock()
roots_lock = threading.Lock()
governance_progress_lock = threading.Lock()
governance_progress_store: Dict[str, Dict[str, Any]] = {}

import sys
import os
# 添加父目录到路径，支持绝对导入
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.models import (
    GenerateDDLRequest,
    GenerateDDLResponse,
    TextGenerateTaskRequest,
    TestConnectionRequest,
    TestConnectionResponse,
    WordRootsResponse,
    WordRootItem,
    GovernanceRequest,
    DbConnectionRequest,
    ExecuteDDLRequest
)
from app.field_type_resolver import get_excel_field_type, resolve_field_type
from app.native_db_executor import execute_native_ddl, test_native_connection

from app.validators.ddl_validator import DDLValidator
from app.config.db_examples import get_db_example
from app.root_governance import (
    load_standard_roots,
    normalize_standard_root_record,
    save_standard_roots,
    load_historical_roots,
    save_historical_roots,
    load_effective_historical_roots,
    save_effective_historical_roots,
    add_to_historical_roots,
    format_standard_roots_for_prompt,
    migrate_from_legacy,
    build_governance_prompt,
    build_governance_merge_prompt,
    chunk_roots_for_governance,
    dedupe_governed_roots,
    GOVERNANCE_CHUNK_SIZE,
    GOVERNANCE_MERGE_CHUNK_SIZE,
    prepare_historical_roots_for_governance,
    filter_historical_roots_against_standard,
    apply_governance_result,
    enrich_standard_roots_with_historical_types,
    BUSINESS_DOMAINS,
)

from app.root_policy import (
    DEFAULT_ABBR_MAX_LEN,
    get_root_constraints,
    get_root_reuse_principle,
    infer_theme_prefix,
    render_theme_prefix_guide,
    resolve_abbr_max_len,
)

app = FastAPI(title="数仓建表 AI 助手")


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    logger.error(f"请求验证失败: {exc.errors()}")
    return JSONResponse(
        status_code=422,
        content={
            "code": 1,
            "message": f"数据验证失败: {str(exc.errors())}",
            "details": exc.errors()
        },
    )


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

LOG_DIR = os.path.join(BASE_DIR, "logs")
LOG_FILE = os.path.join(LOG_DIR, f"app_{datetime.now().strftime('%Y%m%d')}.log")

os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

WORD_ROOTS_FILE = os.path.join(BASE_DIR, "word_roots.json")
STANDARDS_FILE = os.path.join(BASE_DIR, "dev_standards.json")
STANDARDS_DIR = os.path.join(BASE_DIR, "standards")
HISTORY_FILE = os.path.join(BASE_DIR, "ddl_history.json")
DB_CONNECTIONS_FILE = os.path.join(BASE_DIR, "db_connections.json")
DB_SECRET_KEY_FILE = os.path.join(BASE_DIR, "db_secret.key")
CUSTOM_PROMPT_FILE = os.path.join(BASE_DIR, "custom_prompt.txt")
BUILTIN_DIR = os.path.join(BASE_DIR, "builtin")
BUILTIN_PROMPT_FILE = os.path.join(BUILTIN_DIR, "default_prompt.txt")
BUILTIN_STANDARDS_FILE = os.path.join(BUILTIN_DIR, "default_standards.md")

progress_store = {}
task_results = {}
task_validation_flags = {}
system_prompt_cache = None
batch_cache = {}
task_cancel_flags = {}  # 任务终止标志

def update_progress(task_id, current, total, table_name=None, stage=None, matched_count=None,
                    unmatched_count=None, total_fields=None, enable_validation=None,
                    field_progress=None, source_label=None):
    """
    更新任务进度
    
    Args:
        task_id: 任务ID
        current: 当前步骤编号（1-8）
        total: 总步骤数（7或8）
        table_name: 当前处理的表名
        stage: 当前阶段的描述
        matched_count: 已匹配的历史词根数量
        unmatched_count: 未匹配的字段数量
        total_fields: 总字段数量
        enable_validation: 是否启用最终校验
    """
    if stage is None and isinstance(table_name, str):
        stage_prefixes = ("✅", "🔍", "🔧", "⚠", "❌")
        stage_markers = (
            "字段级处理完成",
            "统一校验",
            "统一修正",
            "修正失败",
            "校验失败",
            "完成建表",
            "组装DDL",
            "jieba分词",
            "历史词根匹配",
            "生成字段名",
            "生成字段英文名",
        )
        if table_name.startswith(stage_prefixes) or any(marker in table_name for marker in stage_markers):
            stage = table_name
            table_name = None

    if enable_validation is None:
        enable_validation = task_validation_flags.get(task_id, False)

    existing_data = progress_store.get(task_id, {})
    step1_title = source_label or existing_data.get('source_label') or "解析Excel"

    milestones = [
        {
            "step": 1,
            "title": step1_title,
            "icon": "[1/8]",
            "status": "completed" if current >= 1 else "pending"
        },
        {
            "step": 2,
            "title": "字段分组",
            "icon": "[2/8]",
            "status": "completed" if current >= 2 else "pending"
        },
        {
            "step": 3,
            "title": "jieba分词",
            "icon": "[3/8]",
            "status": "active" if current == 3 else ("completed" if current > 3 else "pending"),
            "sub_progress": None
        },
        {
            "step": 4,
            "title": "历史词根匹配",
            "icon": "[4/8]",
            "status": "active" if current == 4 else ("completed" if current > 4 else "pending"),
            "sub_progress": None
        },
        {
            "step": 5,
            "title": "生成字段名",
            "icon": "[5/8]",
            "status": "active" if current == 5 else ("completed" if current > 5 else "pending"),
            "sub_progress": None
        },
        {
            "step": 6,
            "title": "组装DDL",
            "icon": "[6/8]",
            "status": "active" if current == 6 else ("completed" if current > 6 else "pending"),
            "sub_progress": None
        },
        {
            "step": 7,
            "title": "最终校验",
            "icon": "[7/8]",
            "status": "active" if current == 7 else ("completed" if current > 7 else "pending"),
            "optional": not enable_validation,  # 标记为可选步骤
            "sub_progress": None
        },
        {
            "step": 8,
            "title": "完成建表",
            "icon": "[8/8]",
            "status": "completed" if current >= 8 else "pending"
        }
    ]
    
    # 处理子进度
    if stage:
        if "分词：" in stage or "jieba分词" in stage:
            for m in milestones:
                if m["step"] == 3:
                    m["sub_progress"] = stage.split("[")[-1].replace("]", "") if "[" in stage else stage
        elif "历史词根匹配" in stage or "词根匹配" in stage:
            for m in milestones:
                if m["step"] == 4:
                    m["sub_progress"] = stage.split("[")[-1].replace("]", "") if "[" in stage else stage
        elif "生成字段英文名" in stage or "生成字段名" in stage:
            for m in milestones:
                if m["step"] == 5:
                    m["sub_progress"] = stage.split("[")[-1].replace("]", "") if "[" in stage else stage
        elif "组装DDL" in stage:
            for m in milestones:
                if m["step"] == 6:
                    m["sub_progress"] = stage.split("[")[-1].replace("]", "")
        elif "校验" in stage or "修正" in stage:
            for m in milestones:
                if m["step"] == 7:
                    m["sub_progress"] = stage
    
    progress_data = {
        'current': current,
        'total': total,
        'table_name': table_name,
        'stage': stage,
        'milestones': milestones,
        'overall_progress': int((current / total) * 100)
    }
    if source_label:
        progress_data['source_label'] = source_label
    elif 'source_label' in existing_data:
        progress_data['source_label'] = existing_data['source_label']
    
    # 从现有数据中保留统计信息，除非明确提供了新值
    if matched_count is not None:
        progress_data['matched_count'] = matched_count
        logger.info(f"【进度更新】任务 {task_id} 保存匹配统计: matched_count={matched_count}")
    elif 'matched_count' in existing_data:
        progress_data['matched_count'] = existing_data['matched_count']
        
    if unmatched_count is not None:
        progress_data['unmatched_count'] = unmatched_count
        logger.info(f"【进度更新】任务 {task_id} 保存匹配统计: unmatched_count={unmatched_count}")
    elif 'unmatched_count' in existing_data:
        progress_data['unmatched_count'] = existing_data['unmatched_count']
        
    if total_fields is not None:
        progress_data['total_fields'] = total_fields
        logger.info(f"【进度更新】任务 {task_id} 保存匹配统计: total_fields={total_fields}")
    elif 'total_fields' in existing_data:
        progress_data['total_fields'] = existing_data['total_fields']

    if field_progress is not None:
        progress_data['field_progress'] = field_progress
        logger.info(f"【进度更新】任务 {task_id} 字段生成进度: {field_progress}")
    elif 'field_progress' in existing_data:
        progress_data['field_progress'] = existing_data['field_progress']
    
    progress_store[task_id] = progress_data

def seed_from_builtin(builtin_path, target_path, content_type="text"):
    if not os.path.exists(builtin_path):
        logger.warning(f"内置默认文件不存在: {builtin_path}")
        return None
    try:
        with open(builtin_path, 'r', encoding='utf-8') as f:
            content = f.read()
        if content_type == "standards":
            data = {
                "version": "1.0",
                "last_modified": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "content": content
            }
            with open(target_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        else:
            with open(target_path, 'w', encoding='utf-8') as f:
                f.write(content)
        logger.info(f"已从内置默认文件初始化: {target_path}")
        return content
    except Exception as e:
        logger.error(f"从内置默认文件初始化失败: {builtin_path} -> {target_path}, 错误: {e}")
        return None


DEFAULT_STANDARDS_CONTENT = """## 数仓开发规范（内置默认规范）

### 一、词根原则

#### 规则1：一词一根原则（核心）
同一个中文业务概念，在全局范围内**只能对应一个英文词根**，严禁出现同义词根混用。

**示例**：
- 「安全」→ `safety`（全局唯一），禁用 `safe`、`security`（除非指机密性）
- 「修改」→ `modify`，禁用 `update`、`edit`（除非指界面编辑）
- 「创建」→ `create`，禁用 `build`、`make`、`generate`
- 「部门」→ `dept`（全局唯一），禁用 `department`、`dep`

#### 规则2：词根规范性与唯一性
- **严禁** `_attr`、`biz_attr`、`fld_`、`tbl_` 等无意义占位符
- **严禁**中文字段名，所有字段必须使用英文词根组合
- **同表唯一**：每张表内的字段英文名必须绝对唯一，禁止重复
- **全局唯一**：同一词根在整个数据仓库中含义保持一致

#### 规则3：词根形式服从前台模式
- 词根库同时维护 `full_root` 和 `abbr_root`，但实际建表只能使用当前请求选择的模式。
- 当前为“优先全称”时，只能使用 `full_root`，不得混用 `abbr_root`。
- 当前为“优先缩写”时，只能使用 `abbr_root`，且每个缩写词根长度不得超过当前配置的上限。
- 未命中词根库的新概念，也必须按当前模式生成命名用词根。

#### 规则4：词根库维护
- `full_root` 使用完整英文词根，`abbr_root` 使用不超过当前配置上限的缩写词根。
- 不在规范中指定业务词根映射，实际映射以用户词根库为准。

### 二、标准字段原则

#### 规则1：标准字段命名规则
**普通字段命名**：`词根1_词根2_..._词根N_基础词根`
- 最后一个词根为基础词根
- 词根数量控制在2-4个之间，标准字段总长度禁止超过60

### 三、表命名规范

#### 规则1：表名分层前缀
- `ods_` - 原始数据层
- `dim_` - 维度层
- `dwd_` - 明细事实层
- `dws_` - 汇总事实层
- `ads_` - 应用层

#### 规则2：表名命名格式
`分层前缀_业务域_表名`
**示例**：`dwd_order_detail`、`dim_product_info`

### 四、字段类型规范

| 基础词根 | 适用场景 | 推荐类型 |
|---------|---------|---------|
| `id` | ID/唯一标识 | VARCHAR(64) |
| `name` | 名称 | VARCHAR(128) |
| `code` | 编码 | VARCHAR(64) |
| `num` | 数量 | INT/BIGINT |
| `amt` | 金额 | DECIMAL(18,2) |
| `date` | 日期 | DATE |
| `time` | 时间 | DATETIME |
| `flag` | 标志位 | TINYINT(1) |
| `desc` | 描述 | VARCHAR(512) |

### 五、约束规范
- 外键约束根据业务需求定义
- 字段非空约束根据业务规则定义

### 六、注释规范
- 所有表必须添加表注释 COMMENT
- 所有字段必须添加字段注释 COMMENT
- 注释使用中文，简洁明了

### 七、索引规范
- 根据查询需求合理创建索引
- 避免过度索引影响写入性能
- 复合索引遵循最左前缀原则"""

USER_PROMPT_TEMPLATE = """你是一位数据仓库专家。请根据以下参考信息生成 DDL，不要有思考和任何解释过程：

【建表需求】
{description}

{industry_context_block}

【词根参考】
{word_roots_content}

【数据库类型】
{db_type}
{db_example}

【开发规范】
{standards_content}

{root_constraints}

要求：
1. 严格按照建表需求创建表，确保所有字段都被正确创建。
2. 字段命名必须符合【词根命名规则】。
3. 严格遵循开发规范（表名，字段名等）。
4. 只输出 CREATE TABLE 和 COMMENT 语句，不要解释和思考过程。
5. 严格按照【数据库类型】中的格式生成DDL，包括表名格式和注释语法。
6. 表名分层前缀与 schema 映射：ods -> ods, dim -> dim, dwd -> dwd, dws -> dws, ads -> ads, input -> input
7. {root_reuse_principle}
8. 【新词根扩展】如果已保存的词根中检索不到要建表的字段，请按当前词根模式生成命名用词根，并在 SQL 输出结束后，单独一行按以下格式补全词根库信息：【新词根】词根全称(例如：create_time):词根缩写(例如：crt_tm):中文名称:推荐字段类型"""

TABLE_NAME_PROMPT = """
【任务】根据以下信息生成符合规范的英文表名，不要有思考和任何解释过程

【中文表名】{chinese_table_name}

{industry_context_block}

【开发规范】
{standards_content}

【数据库类型】{db_type}

【表名格式要求】{table_format}

【输出格式】
请只输出英文表名，不要包含任何解释或额外内容。

【输出示例】
{table_name_example}
"""

NEW_ROOTS_OUTPUT_INSTRUCTION = """

【新词根返回格式】
如发现新词根，必须在 SQL 输出结束后单独输出，每行一个，格式固定为：
【新词根】词根全称:词根缩写:中文名称:推荐字段类型
不要输出“词根全称”“词根缩写”“中文名称”“推荐字段类型”等字段标签。
示例：【新词根】create_time:crt_tm:创建时间:DATETIME
"""


def append_new_roots_instruction(prompt: str) -> str:
    if '【新词根返回格式】' in prompt:
        return prompt
    return f"{prompt}{NEW_ROOTS_OUTPUT_INSTRUCTION}"


def append_root_mode_constraints(prompt: str, root_constraints: str, root_reuse_principle: str) -> str:
    """Ensure old custom prompts cannot bypass the selected root mode."""
    if root_constraints in prompt and root_reuse_principle in prompt:
        return prompt
    return f"""{prompt}

【当前词根模式强制约束】
{root_constraints}
{root_reuse_principle}
"""


def normalize_industry_context(industry_context: Optional[str]) -> str:
    return str(industry_context or "").strip()


def build_industry_context_block(industry_context: Optional[str]) -> str:
    normalized = normalize_industry_context(industry_context)
    if not normalized:
        return ""
    return f"【行业背景】\n{normalized}"


def merge_standards_with_industry(standards_content: str, industry_context: Optional[str]) -> str:
    normalized = normalize_industry_context(industry_context)
    standards_text = str(standards_content or "").strip()
    if not normalized:
        return standards_text
    industry_block = f"## 行业背景\n{normalized}"
    if industry_block in standards_text:
        return standards_text
    if not standards_text:
        return industry_block
    return f"{standards_text}\n\n{industry_block}"


def get_system_prompt() -> str:
    global system_prompt_cache
    if system_prompt_cache is not None:
        return system_prompt_cache
    
    standards = load_standards()
    content = standards.get('content', '')
    
    if not content:
        system_prompt_cache = "你是一位数据仓库专家，请根据用户的建表需求生成符合规范的DDL语句，不要有思考和任何解释过程。"
    else:
        system_prompt_cache = f"""你是一位数据仓库专家，必须严格遵循以下开发规范：

{content}

请根据用户的建表需求，使用规范中的词根和命名规则生成正确的DDL语句，不要有思考和任何解释过程。"""
    
    logger.info(f"System Prompt缓存已初始化，长度: {len(system_prompt_cache)}")
    return system_prompt_cache

def refresh_system_prompt_cache():
    global system_prompt_cache
    system_prompt_cache = None
    logger.info("System Prompt缓存已刷新")


def load_legacy_default_standard() -> dict:
    legacy_default_path = os.path.join(STANDARDS_DIR, "default.json")
    if not os.path.exists(legacy_default_path):
        return {}

    try:
        with open(legacy_default_path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
        content = (data.get('content') or '').strip()
        if not content:
            return {}
        return {
            "content": content,
            "updated_at": data.get('last_modified', ''),
            "name": data.get('name', '') or "默认规范"
        }
    except Exception as e:
        logger.error(f"读取历史默认规范失败: {e}")
        return {}


def load_all_standards() -> list:
    logger.info(f"加载所有规范文件")
    standards = []
    
    if not os.path.exists(STANDARDS_DIR):
        os.makedirs(STANDARDS_DIR, exist_ok=True)
        logger.info(f"创建规范目录: {STANDARDS_DIR}")
    
    # 先加载自定义规范，检查是否有激活的
    custom_standards = []
    has_active_custom = False
    
    try:
        for filename in os.listdir(STANDARDS_DIR):
            if filename.endswith('.json'):
                filepath = os.path.join(STANDARDS_DIR, filename)
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    data = json.load(f)
                    standard_id = filename.replace('.json', '')
                    if standard_id == "default":
                        continue
                    is_active = data.get('is_active', False)
                    if is_active:
                        has_active_custom = True
                    custom_standards.append({
                        "id": standard_id,
                        "name": data.get('name', '未命名规范'),
                        "content": data.get('content', ''),
                        "updated_at": data.get('last_modified', ''),
                        "is_active": is_active
                    })
    except Exception as e:
        logger.error(f"加载规范列表失败: {e}")
    
    # 加载默认规范，设置正确的激活状态
    legacy_default = load_legacy_default_standard()
    if os.path.exists(STANDARDS_FILE):
        try:
            with open(STANDARDS_FILE, 'r', encoding='utf-8-sig') as f:
                data = json.load(f)
                content = data.get('content', '').strip()
                if not content:
                    if legacy_default:
                        standards.append({
                            "id": "default",
                            "name": "默认规范",
                            "content": legacy_default["content"],
                            "updated_at": legacy_default["updated_at"],
                            "is_active": not has_active_custom
                        })
                    else:
                        standards.append({
                            "id": "default",
                            "name": "默认规范（内置）",
                            "content": DEFAULT_STANDARDS_CONTENT,
                            "updated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            "is_active": not has_active_custom  # 如果没有自定义规范激活，则默认规范激活
                        })
                else:
                    standards.append({
                        "id": "default",
                        "name": "默认规范",
                        "content": content,
                        "updated_at": data.get('last_modified', ''),
                        "is_active": not has_active_custom
                    })
        except Exception as e:
            logger.error(f"加载旧版规范文件失败: {e}")
            if legacy_default:
                standards.append({
                    "id": "default",
                    "name": "默认规范",
                    "content": legacy_default["content"],
                    "updated_at": legacy_default["updated_at"],
                    "is_active": not has_active_custom
                })
            else:
                standards.append({
                    "id": "default",
                    "name": "默认规范（内置）",
                    "content": DEFAULT_STANDARDS_CONTENT,
                    "updated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "is_active": not has_active_custom
                })
    else:
        if legacy_default:
            standards.append({
                "id": "default",
                "name": "默认规范",
                "content": legacy_default["content"],
                "updated_at": legacy_default["updated_at"],
                "is_active": not has_active_custom
            })
        else:
            standards.append({
                "id": "default",
                "name": "默认规范（内置）",
                "content": DEFAULT_STANDARDS_CONTENT,
                "updated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "is_active": not has_active_custom
            })
    
    # 添加自定义规范
    standards.extend(custom_standards)
    
    standards.sort(key=lambda x: x['updated_at'] or '', reverse=True)
    logger.info(f"共加载 {len(standards)} 个规范")
    return standards


def save_standard_by_id(standard_id: str, content: str, name: str = None, is_active: bool = False):
    logger.info(f"保存规范: {standard_id}")

    if standard_id == "default":
        try:
            data = {
                "version": "1.0",
                "last_modified": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "content": content
            }
            with open(STANDARDS_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info("默认规范保存成功")
            refresh_system_prompt_cache()
            return True
        except Exception as e:
            logger.error(f"保存默认规范失败: {e}")
            return False
    
    if not os.path.exists(STANDARDS_DIR):
        os.makedirs(STANDARDS_DIR, exist_ok=True)
    
    filepath = os.path.join(STANDARDS_DIR, f"{standard_id}.json")
    try:
        data = {
            "name": name or "未命名规范",
            "content": content,
            "last_modified": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "is_active": is_active
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"规范 {standard_id} 保存成功")
        refresh_system_prompt_cache()
        return True
    except Exception as e:
        logger.error(f"保存规范失败: {e}")
        return False


def delete_standard_by_id(standard_id: str) -> bool:
    logger.info(f"删除规范: {standard_id}")
    
    if standard_id == "default":
        if os.path.exists(STANDARDS_FILE):
            try:
                with open(STANDARDS_FILE, 'r', encoding='utf-8-sig') as f:
                    data = json.load(f)
                data['content'] = ''
                data['last_modified'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                with open(STANDARDS_FILE, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                logger.info("默认规范内容已清空")
            except Exception as e:
                logger.error(f"清空默认规范失败: {e}")
                return False
        refresh_system_prompt_cache()
        return True
    
    filepath = os.path.join(STANDARDS_DIR, f"{standard_id}.json")
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            logger.info(f"规范文件已删除: {filepath}")
            return True
        except Exception as e:
            logger.error(f"删除规范文件失败: {e}")
            return False
    
    return False


def load_standards() -> dict:
    if os.path.exists(STANDARDS_DIR):
        try:
            for filename in os.listdir(STANDARDS_DIR):
                if not filename.endswith('.json'):
                    continue
                if filename == 'default.json':
                    continue
                filepath = os.path.join(STANDARDS_DIR, filename)
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    active_data = json.load(f)
                content = active_data.get('content', '').strip()
                if active_data.get('is_active', False) and content:
                    logger.info(f"使用已启用规范: {active_data.get('name', filename)}")
                    return {
                        "version": "1.0",
                        "last_modified": active_data.get('last_modified', datetime.now().strftime('%Y-%m-%d')),
                        "content": content
                    }
        except Exception as e:
            logger.error(f"加载已启用规范失败: {e}")

    logger.info(f"加载规范文件: {STANDARDS_FILE}")
    if os.path.exists(STANDARDS_FILE):
        try:
            with open(STANDARDS_FILE, 'r', encoding='utf-8-sig') as f:
                data = json.load(f)
                content = data.get('content', '').strip()
                if not content:
                    legacy_default = load_legacy_default_standard()
                    if legacy_default:
                        logger.info("主规范为空，使用历史默认规范内容")
                        return {
                            "version": "1.0",
                            "last_modified": legacy_default.get('updated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                            "content": legacy_default["content"]
                        }
                    logger.info("规范文件内容为空，尝试从内置默认文件初始化")
                    builtin_content = seed_from_builtin(BUILTIN_STANDARDS_FILE, STANDARDS_FILE, "standards")
                    if builtin_content:
                        with open(STANDARDS_FILE, 'r', encoding='utf-8-sig') as f:
                            data = json.load(f)
                    else:
                        logger.info("内置默认文件不可用，使用硬编码兜底规范")
                        data = {"version": "1.0", "last_modified": datetime.now().strftime('%Y-%m-%d'), "content": DEFAULT_STANDARDS_CONTENT}
                logger.info(f"成功加载规范文件，版本: {data.get('version', 'unknown')}")
                return data
        except Exception as e:
            logger.error(f"加载规范文件失败: {e}")
    logger.info("规范文件不存在，尝试从内置默认文件初始化")
    legacy_default = load_legacy_default_standard()
    if legacy_default:
        logger.info("主规范文件不存在，使用历史默认规范内容")
        return {
            "version": "1.0",
            "last_modified": legacy_default.get('updated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            "content": legacy_default["content"]
        }
    builtin_content = seed_from_builtin(BUILTIN_STANDARDS_FILE, STANDARDS_FILE, "standards")
    if builtin_content:
        return {"version": "1.0", "last_modified": datetime.now().strftime('%Y-%m-%d'), "content": builtin_content}
    logger.info("内置默认文件不可用，使用硬编码兜底规范")
    return {"version": "1.0", "last_modified": datetime.now().strftime('%Y-%m-%d'), "content": DEFAULT_STANDARDS_CONTENT}


def save_standards(content: str):
    logger.info(f"保存规范文件: {STANDARDS_FILE}")
    try:
        data = {
            "version": "1.0",
            "last_modified": datetime.now().strftime('%Y-%m-%d'),
            "content": content
        }
        with open(STANDARDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info("规范文件保存成功")
        refresh_system_prompt_cache()
    except Exception as e:
        logger.error(f"保存规范文件失败: {e}")
        raise


def delete_standards(standard_id: str):
    logger.info(f"删除规范: {standard_id}")
    return delete_standard_by_id(standard_id)


def save_ddl_history(ddl_data: dict):
    import uuid
    history = []
    try:
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                history = json.load(f)
    except:
        history = []
    
    record = {
        "id": str(uuid.uuid4())[:8],
        "timestamp": datetime.now().isoformat(),
        "type": ddl_data.get('type', 'single'),
        "tables_count": ddl_data.get('tables_count', 1),
        "success_count": ddl_data.get('success_count', 1),
        "ddl": ddl_data.get('ddl', ''),
        "description": ddl_data.get('description', ''),
        "db_type": ddl_data.get('db_type', ''),
        "root_match_priority": ddl_data.get('root_match_priority', '')
    }
    
    history.insert(0, record)
    history = history[:20]
    
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def get_ddl_history(limit: int = 20):
    try:
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                history = json.load(f)
                return history[:limit]
    except:
        pass
    return []


def delete_ddl_history(record_id: str):
    try:
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                history = json.load(f)
            
            history = [h for h in history if h.get('id') != record_id]
            
            with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
            return True
    except:
        pass
    return False


def _get_db_cipher():
    try:
        from cryptography.fernet import Fernet
    except ImportError as exc:
        raise RuntimeError("缺少密码加密依赖 cryptography") from exc

    if os.path.exists(DB_SECRET_KEY_FILE):
        with open(DB_SECRET_KEY_FILE, 'rb') as f:
            key = f.read().strip()
    else:
        key = Fernet.generate_key()
        with open(DB_SECRET_KEY_FILE, 'wb') as f:
            f.write(key)
    return Fernet(key)


def encrypt_db_password(password: str) -> str:
    if not password:
        return ""
    return _get_db_cipher().encrypt(password.encode('utf-8')).decode('utf-8')


def decrypt_db_password(encrypted_password: str) -> str:
    if not encrypted_password:
        return ""
    return _get_db_cipher().decrypt(encrypted_password.encode('utf-8')).decode('utf-8')


def load_db_connections(include_password: bool = False) -> List[Dict[str, Any]]:
    try:
        if os.path.exists(DB_CONNECTIONS_FILE):
            with open(DB_CONNECTIONS_FILE, 'r', encoding='utf-8') as f:
                connections = json.load(f)
        else:
            connections = []
    except Exception as e:
        logger.error(f"加载数据库连接配置失败: {e}")
        connections = []

    result = []
    for conn in connections:
        item = dict(conn)
        encrypted_password = item.pop("password_encrypted", "")
        item.pop("jdbc_url", None)
        item.pop("driver_class", None)
        item.pop("driver_path", None)
        item["has_password"] = bool(encrypted_password)
        if include_password:
            item["password"] = decrypt_db_password(encrypted_password)
        result.append(item)
    return result


def save_db_connections(connections: List[Dict[str, Any]]):
    safe_connections = []
    for conn in connections:
        item = dict(conn)
        item.pop("password", None)
        item.pop("has_password", None)
        safe_connections.append(item)
    with open(DB_CONNECTIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(safe_connections, f, ensure_ascii=False, indent=2)


def get_db_connection(connection_id: str, include_password: bool = True) -> Optional[Dict[str, Any]]:
    connections = load_db_connections(include_password=include_password)
    return next((conn for conn in connections if conn.get("id") == connection_id), None)


def upsert_db_connection(data: DbConnectionRequest, connection_id: Optional[str] = None) -> Dict[str, Any]:
    stored_connections = []
    if os.path.exists(DB_CONNECTIONS_FILE):
        try:
            with open(DB_CONNECTIONS_FILE, 'r', encoding='utf-8') as f:
                stored_connections = json.load(f)
        except Exception:
            stored_connections = []

    existing = None
    if connection_id:
        existing = next((conn for conn in stored_connections if conn.get("id") == connection_id), None)
        if not existing:
            raise HTTPException(status_code=404, detail="数据库连接不存在")

    password_encrypted = existing.get("password_encrypted", "") if existing else ""
    if data.password:
        password_encrypted = encrypt_db_password(data.password)

    record = {
        "id": connection_id or str(uuid.uuid4())[:8],
        "name": data.name,
        "db_type": data.db_type,
        "host": data.host,
        "port": data.port,
        "database": data.database,
        "username": data.username,
        "password_encrypted": password_encrypted,
        "updated_at": datetime.now().isoformat(),
    }

    if existing:
        index = stored_connections.index(existing)
        stored_connections[index] = record
    else:
        record["created_at"] = datetime.now().isoformat()
        stored_connections.insert(0, record)

    save_db_connections(stored_connections)
    public_record = dict(record)
    public_record.pop("password_encrypted", None)
    public_record["has_password"] = bool(password_encrypted)
    return public_record


def update_ddl_history_execution(record_id: str, execution_info: Dict[str, Any]) -> bool:
    try:
        if not os.path.exists(HISTORY_FILE):
            return False
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            history = json.load(f)
        for record in history:
            if record.get("id") == record_id:
                record["execute_status"] = execution_info.get("status", "")
                record["execute_time"] = execution_info.get("time", "")
                record["connection_name"] = execution_info.get("connection_name", "")
                record["execute_message"] = execution_info.get("message", "")
                record["executed_count"] = execution_info.get("executed_count", 0)
                with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
                    json.dump(history, f, ensure_ascii=False, indent=2)
                return True
    except Exception as e:
        logger.error(f"更新建表执行历史失败: {e}")
    return False

MAX_ROOTS_PER_REQUEST = 50
COMMON_ROOTS_COUNT = 20

requests_session = requests.Session()


def set_governance_progress(task_id: Optional[str], **updates):
    if not task_id:
        return
    with governance_progress_lock:
        state = governance_progress_store.setdefault(task_id, {})
        state.update(updates)
        state["updated_at"] = datetime.now().isoformat()


def get_governance_progress(task_id: str) -> Dict[str, Any]:
    with governance_progress_lock:
        return dict(governance_progress_store.get(task_id, {}))
requests_session.headers.update({'Content-Type': 'application/json'})
requests_session.timeout = 300
requests_session.keep_alive = True
requests_session.proxies = {"http": None, "https": None}
requests_session.trust_env = False


def log_governance_event(
    event: str,
    *,
    task_id: Optional[str] = None,
    phase: Optional[str] = None,
    chunk_index: Optional[int] = None,
    total_chunks: Optional[int] = None,
    merge_round: Optional[int] = None,
    prompt_kind: Optional[str] = None,
    prompt_length: Optional[int] = None,
    prompt_hash: Optional[str] = None,
    item_count: Optional[int] = None,
    result_count: Optional[int] = None,
    elapsed_ms: Optional[int] = None,
    retry_count: Optional[int] = None,
    extra: Optional[Dict[str, Any]] = None,
):
    details = [f"event={event}"]
    if task_id:
        details.append(f"task_id={task_id}")
    if phase:
        details.append(f"phase={phase}")
    if chunk_index is not None and total_chunks is not None:
        details.append(f"chunk={chunk_index}/{total_chunks}")
    elif chunk_index is not None:
        details.append(f"chunk={chunk_index}")
    if merge_round is not None:
        details.append(f"merge_round={merge_round}")
    if prompt_kind:
        details.append(f"prompt_kind={prompt_kind}")
    if prompt_length is not None:
        details.append(f"prompt_chars={prompt_length}")
    if prompt_hash:
        details.append(f"prompt_hash={prompt_hash}")
    if item_count is not None:
        details.append(f"item_count={item_count}")
    if result_count is not None:
        details.append(f"result_count={result_count}")
    if elapsed_ms is not None:
        details.append(f"elapsed_ms={elapsed_ms}")
    if retry_count is not None:
        details.append(f"retry_count={retry_count}")
    if extra:
        for key, value in extra.items():
            details.append(f"{key}={value}")
    logger.info("【词根治理】%s", ", ".join(details))


def load_word_roots() -> list:
    """Load roots used by DDL generation; never fall back to legacy word_roots.json."""
    try:
        roots = load_standard_roots()
        logger.info(f"加载标准词根文件，共 {len(roots)} 条")
        return roots
    except Exception as e:
        logger.error(f"加载标准词根失败: {e}")
        return []


def load_legacy_word_roots_for_governance() -> list:
    """Load historical roots for governance from the canonical governance source."""
    roots = load_effective_historical_roots()
    logger.info(f"加载治理历史词根，共 {len(roots)} 条")
    return roots
def save_word_roots(roots: list):
    """save to standard_roots"""
    save_standard_roots(roots)
    logger.info(f"saved {len(roots)} standard roots")
def merge_and_save_roots(new_roots: list):
    logger.info(f"合并新词根: {len(new_roots)} 条")
    existing = load_word_roots()

    for root in new_roots:
        chinese_name = root.get('chinese_name')
        if not chinese_name:
            continue

        existing_idx = next((i for i, r in enumerate(existing) if r.get('chinese_name') == chinese_name), None)
        if existing_idx is not None:
            existing[existing_idx] = root
            logger.info(f"覆盖词根: {chinese_name} -> {root.get('full_root')}")
        else:
            existing.append(root)
            logger.info(f"新增词根: {chinese_name} -> {root.get('full_root')}")

    save_word_roots(existing)


def classify_roots_by_domain(roots: list) -> Dict[str, list]:
    domain_map = {}
    for root in roots:
        domain = root.get('business_domain', '基础通用')
        if domain not in domain_map:
            domain_map[domain] = []
        domain_map[domain].append(root)
    logger.info(f"词根按业务域分类: {list(domain_map.keys())}")
    return domain_map


def identify_target_domains(description: str, domain_map: Dict[str, list]) -> List[str]:
    description_lower = description.lower()
    matched_domains = []

    for domain in domain_map.keys():
        if domain != '基础通用' and domain in description_lower:
            matched_domains.append(domain)
            logger.info(f"匹配到业务域: {domain}")

    if not matched_domains:
        logger.info("未识别到特定业务域，将使用所有词根")

    return matched_domains


def extract_keywords_from_description(description: str) -> List[str]:
    chinese_pattern = re.compile(r'[\u4e00-\u9fa5]+')
    chinese_words = chinese_pattern.findall(description)

    stopwords = {'的', '了', '和', '是', '在', '有', '与', '或', '等', '及', '等', '包含', '包括', '以及', '包括'}
    keywords = [w for w in chinese_words if len(w) >= 2 and w not in stopwords]

    expanded_keywords = set(keywords)
    for keyword in keywords:
        for i in range(2, len(keyword)):
            prefix = keyword[:i]
            suffix = keyword[i:]
            if len(prefix) >= 2:
                expanded_keywords.add(prefix)
            if len(suffix) >= 2:
                expanded_keywords.add(suffix)

    expanded_list = sorted(list(expanded_keywords), key=lambda x: -len(x))
    logger.info(f"从描述中提取关键词: {expanded_list}")
    return expanded_list


def filter_roots_by_keywords(keywords: List[str], roots: list, max_count: int = 50) -> Tuple[List[dict], List[dict]]:
    matched = []
    unmatched = []

    for root in roots:
        chinese_name = root.get('chinese_name', '')
        full_root = root.get('full_root', '').lower()
        abbr_root = root.get('abbr_root', '').lower()

        is_matched = False
        for keyword in keywords:
            if keyword in chinese_name or keyword in full_root or keyword in abbr_root:
                is_matched = True
                break

        if is_matched:
            matched.append(root)
        else:
            unmatched.append(root)

    logger.info(f"关键词匹配: {len(matched)} 条，剩余: {len(unmatched)} 条")

    if len(matched) > max_count:
        matched = matched[:max_count]

    return matched, unmatched


def get_common_roots(roots: list, count: int = 20) -> List[dict]:
    return roots[:count] if len(roots) > count else roots


def format_word_roots_for_prompt(roots: list, priority: str = 'full') -> str:
    """delegate to root_governance"""
    return format_standard_roots_for_prompt(roots, priority)
def parse_word_roots_text(text: str) -> list:
    logger.info("解析文本格式词根")
    roots = []
    for line in text.strip().split('\n'):
        line = line.strip()
        if not line or line.startswith('#') or line.startswith('//'):
            continue

        parts = line.replace('：', ':').split(':')
        if len(parts) >= 1:
            values = [part.strip() for part in parts]
            business_domain = ''
            domain_code = ''
            chinese_name = ''
            full_root = ''
            abbr_root = ''
            recommended_type = ''

            if len(values) >= 6:
                business_domain, domain_code, chinese_name, full_root, abbr_root, recommended_type = values[:6]
            elif len(values) >= 4:
                chinese_name, full_root, abbr_root, recommended_type = values[:4]
            elif len(values) >= 2:
                chinese_name, full_root = values[:2]
                abbr_root = values[2] if len(values) > 2 else ''
                recommended_type = values[3] if len(values) > 3 else ''

            if not chinese_name:
                continue

            roots.append({
                'business_domain': business_domain,
                'domain_code': domain_code,
                'chinese_name': chinese_name,
                'full_root': full_root,
                'abbr_root': abbr_root,
                'recommended_type': recommended_type
            })
    logger.info(f"解析到 {len(roots)} 条词根")
    return roots


def parse_excel_content(file_content: bytes) -> list:
    logger.info("解析Excel文件词根")
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    roots = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num == 1:
                continue
            if row and len(row) >= 1:
                business_domain = str(row[0]).strip() if len(row) > 0 and row[0] else ''
                domain_code = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                chinese_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                full_root = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                abbr_root = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                recommended_type = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                if not chinese_name:
                    continue
                roots.append({
                    'business_domain': business_domain,
                    'domain_code': domain_code,
                    'chinese_name': chinese_name,
                    'full_root': full_root,
                    'abbr_root': abbr_root,
                    'recommended_type': recommended_type
                })
    logger.info(f"Excel解析到 {len(roots)} 条词根")
    return roots


def parse_excel_to_text(file_content: bytes) -> str:
    logger.info("解析Excel文件为文本")
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    result = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_content = []
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
            if row_text.strip():
                sheet_content.append(row_text)
        if sheet_content:
            result.append(f"【{sheet}】")
            result.extend(sheet_content)
    logger.info(f"Excel解析到 {len(result)} 行文本")
    return "\n".join(result)


def parse_docx_content(file_content: bytes) -> str:
    doc = Document(io.BytesIO(file_content))
    result = []
    for para in doc.paragraphs:
        if para.text.strip():
            result.append(para.text)
    return "\n".join(result)


def parse_csv_content(file_content: bytes) -> str:
    try:
        text = file_content.decode("utf-8")
    except UnicodeDecodeError:
        text = file_content.decode("gbk")
    return text


def parse_csv_roots(file_content: bytes) -> list:
    text = parse_csv_content(file_content)
    reader = csv.reader(io.StringIO(text))
    roots = []
    for row_num, row in enumerate(reader, 1):
        if row_num == 1:
            continue
        values = [str(cell).strip() if cell is not None else '' for cell in row]
        if not any(values):
            continue
        business_domain = values[0] if len(values) > 0 else ''
        domain_code = values[1] if len(values) > 1 else ''
        chinese_name = values[2] if len(values) > 2 else ''
        full_root = values[3] if len(values) > 3 else ''
        abbr_root = values[4] if len(values) > 4 else ''
        recommended_type = values[5] if len(values) > 5 else ''
        if not chinese_name:
            continue
        roots.append({
            'business_domain': business_domain,
            'domain_code': domain_code,
            'chinese_name': chinese_name,
            'full_root': full_root,
            'abbr_root': abbr_root,
            'recommended_type': recommended_type,
        })
    return roots


def filter_and_prepare_roots(description: str, priority: str = 'full') -> Tuple[list, str]:
    all_roots = load_word_roots()

    if not all_roots:
        logger.info("历史词根库为空，返回空列表")
        return [], "无"

    domain_map = classify_roots_by_domain(all_roots)

    target_domains = identify_target_domains(description, domain_map)

    selected_roots = []
    matched_chinese_names = set()

    if target_domains:
        for domain in target_domains:
            if domain in domain_map:
                for root in domain_map[domain]:
                    if root.get('chinese_name') not in matched_chinese_names:
                        selected_roots.append(root)
                        matched_chinese_names.add(root.get('chinese_name'))
        logger.info(f"按业务域筛选: {len(selected_roots)} 条")

    keywords = extract_keywords_from_description(description)
    if keywords:
        matched_by_keyword, unmatched = filter_roots_by_keywords(keywords, all_roots, MAX_ROOTS_PER_REQUEST)

        for root in matched_by_keyword:
            if root.get('chinese_name') not in matched_chinese_names:
                selected_roots.append(root)
                matched_chinese_names.add(root.get('chinese_name'))

        logger.info(f"关键词筛选后: {len(selected_roots)} 条")

    if len(selected_roots) < MAX_ROOTS_PER_REQUEST:
        for root in all_roots:
            if len(selected_roots) >= MAX_ROOTS_PER_REQUEST:
                break
            if root.get('chinese_name') not in matched_chinese_names:
                selected_roots.append(root)
                matched_chinese_names.add(root.get('chinese_name'))

        logger.info(f"补充词根后: {len(selected_roots)} 条")

    formatted = format_word_roots_for_prompt(selected_roots, priority)
    return selected_roots, formatted


def _extract_new_roots_from_response_legacy(content: str, existing_roots: list) -> list:
    new_roots = []
    if '【新词根】' not in content:
        return new_roots

    logger.info("检测到【新词根】标记")
    
    existing_chinese = {r.get('chinese_name', '') for r in existing_roots}
    
    lines = content.split('\n')
    in_new_roots_section = False
    new_roots_lines = []
    
    for line in lines:
        if '【新词根】' in line:
            in_new_roots_section = True
            line = line.replace('【新词根】', '').strip()
            if line:
                new_roots_lines.append(line)
        elif in_new_roots_section:
            if line.strip().startswith('【') and '【新词根】' not in line:
                break
            elif line.strip():
                new_roots_lines.append(line.strip())
    
    logger.info(f"找到 {len(new_roots_lines)} 行新词根内容")
    
    for line in new_roots_lines:
        if not line or line.startswith('--'):
            continue
        
        line = line.strip()
        line = line.replace('：', ':').replace('，', ',').replace('（', '(').replace('）', ')')
        
        parts = re.split(r'[:\-–—=,，]', line, maxsplit=3)
        parts = [p.strip() for p in parts if p.strip()]
        
        if len(parts) >= 4:
            full_root = parts[0]
            abbr_root = parts[1]
            chinese_name = parts[2]
            recommended_type = parts[3]
            
            if chinese_name and chinese_name not in existing_chinese:
                new_roots.append({
                    'business_domain': '基础通用',
                    'chinese_name': chinese_name,
                    'full_root': full_root,
                    'abbr_root': abbr_root,
                    'recommended_type': recommended_type
                })
                existing_chinese.add(chinese_name)
                logger.info(f"提取新词根: {chinese_name} -> {full_root} ({abbr_root})")

    logger.info(f"共提取 {len(new_roots)} 个新词根")
    return new_roots


def extract_new_roots_from_response(content: str, existing_roots: list) -> list:
    """Extract roots emitted as: 【新词根】full_root:abbr_root:chinese_name:recommended_type."""
    marker = '【新词根】'
    if marker not in content:
        return []

    logger.info("检测到【新词根】标记")
    existing_chinese = {r.get('chinese_name', '') for r in existing_roots}
    label_aliases = {
        '词根全称', '全称', 'full_root',
        '词根缩写', '缩写', 'abbr_root',
        '中文名称', '中文名', '中文含义', 'chinese_name',
        '推荐字段类型', '推荐类型', '字段类型', 'recommended_type'
    }

    new_roots = []
    collecting = False
    for raw_line in content.splitlines():
        line = raw_line.strip()
        if marker in line:
            collecting = True
            line = line.split(marker, 1)[1].strip()
        elif collecting:
            if line.startswith('【') and marker not in line:
                break
        else:
            continue

        if not line:
            continue

        line = re.sub(r'^(?:--|#|//)\s*', '', line).strip()
        line = re.sub(r'^\s*(?:[-*]|\d+[.)、])\s*', '', line).strip()
        line = line.replace('：', ':').replace('，', ',').replace('（', '(').replace('）', ')')
        parts = re.split(r'[:\-–—,，\s]+', line, maxsplit=7)
        parts = [p.strip() for p in parts if p.strip()]
        parts = [p for p in parts if p not in label_aliases]

        if len(parts) < 4:
            continue

        full_root, abbr_root, chinese_name = parts[:3]
        recommended_type = ":".join(parts[3:])
        if chinese_name and chinese_name not in existing_chinese:
            new_roots.append({
                'business_domain': '基础通用',
                'chinese_name': chinese_name,
                'full_root': full_root,
                'abbr_root': abbr_root,
                'recommended_type': recommended_type
            })
            existing_chinese.add(chinese_name)
            logger.info(f"提取新词根: {chinese_name} -> {full_root} ({abbr_root})")

    logger.info(f"共提取 {len(new_roots)} 个新词根")
    return new_roots


def remove_new_roots_markers(content: str) -> str:
    """从DDL内容中移除【新词根】标记行，只保留CREATE TABLE语句"""
    lines = content.split('\n')
    clean_lines = []
    skip_mode = False
    
    for line in lines:
        if '【新词根】' in line:
            skip_mode = True
            continue
        if skip_mode:
            if line.strip() and not line.startswith('--') and not line.strip().startswith('【'):
                skip_mode = False
            else:
                continue
        if line.strip():
            clean_lines.append(line)
    
    return '\n'.join(clean_lines)


def remove_primary_key_clauses(content: str) -> str:
    """Remove generated primary-key declarations while preserving field definitions."""
    if not content:
        return content

    lines = []
    for line in content.splitlines():
        if re.search(r'^\s*,?\s*PRIMARY\s+KEY\s*\(', line, re.IGNORECASE):
            continue

        cleaned = re.sub(
            r'\s+PRIMARY\s+KEY\b',
            '',
            line,
            flags=re.IGNORECASE,
        )
        lines.append(cleaned)

    return '\n'.join(lines)


def extract_llm_content(result: dict) -> str:
    choices = result.get("choices") or []
    if not choices:
        return ""
    choice = choices[0] or {}
    message = choice.get("message") or {}
    delta = choice.get("delta") or {}
    content = message.get("content")
    if not content:
        content = delta.get("content")
    if isinstance(content, list):
        content = "\n".join(
            item.get("text", "") if isinstance(item, dict) else str(item)
            for item in content
        )
    if content and str(content).strip():
        return str(content).strip()
    fallback_text = str(choices[0].get("text") or "").strip()
    if fallback_text:
        return fallback_text
    reasoning = str(message.get("reasoning_content") or "").strip()
    if reasoning:
        logger.warning("LLM returned reasoning_content but no final content; treating response as empty")
    return ""


def extract_json_array_from_text(text: str) -> list:
    if not text:
        return []

    cleaned = re.sub(r"```json\s*", "", text, flags=re.IGNORECASE).replace("```", "").strip()
    try:
        parsed = json.loads(cleaned)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        pass

    start = cleaned.find('[')
    end = cleaned.rfind(']')
    if start == -1 or end == -1 or end <= start:
        return []

    snippet = cleaned[start:end + 1]
    try:
        parsed = json.loads(snippet)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        logger.warning("治理结果 JSON 解析失败")
        return []


def extract_json_object_from_text(text: str) -> dict:
    if not text:
        return {}

    cleaned = re.sub(r"```json\s*", "", text, flags=re.IGNORECASE).replace("```", "").strip()
    if not cleaned:
        return {}

    try:
        parsed = json.loads(cleaned)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        pass

    start = cleaned.find('{')
    end = cleaned.rfind('}')
    if start == -1 or end == -1 or end <= start:
        return {}

    snippet = cleaned[start:end + 1]
    try:
        parsed = json.loads(snippet)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        logger.warning("结构化表定义 JSON 解析失败")
        return {}


def extract_governance_objects_from_text(text: str) -> list:
    if not text:
        return []

    cleaned = re.sub(r"```json\s*", "", text, flags=re.IGNORECASE).replace("```", "").strip()
    if not cleaned:
        return []

    def _extract_from_parsed(parsed_value) -> list:
        if isinstance(parsed_value, dict):
            if "choices" in parsed_value:
                nested_text = extract_llm_content(parsed_value)
                if nested_text:
                    return extract_governance_objects_from_text(nested_text)
            if (
                parsed_value.get("domain_code")
                or parsed_value.get("business_domain")
                or parsed_value.get("chinese_name")
                or parsed_value.get("full_root")
                or parsed_value.get("abbr_root")
            ):
                return [parsed_value]
            return []
        if isinstance(parsed_value, list):
            extracted = []
            for item in parsed_value:
                if isinstance(item, dict):
                    extracted.extend(_extract_from_parsed(item))
            return extracted
        return []

    try:
        parsed = json.loads(cleaned)
        extracted = _extract_from_parsed(parsed)
        if extracted:
            return extracted
    except Exception:
        pass

    objects = []
    decoder = json.JSONDecoder()
    index = 0
    while index < len(cleaned):
        while index < len(cleaned) and cleaned[index] not in '{[':
            index += 1
        if index >= len(cleaned):
            break
        try:
            parsed, end = decoder.raw_decode(cleaned[index:])
        except Exception:
            index += 1
            continue
        objects.extend(_extract_from_parsed(parsed))
        index += max(end, 1)

    return objects


def extract_governance_objects_from_response(response) -> list:
    response_text = ""

    if hasattr(response, "iter_lines"):
        try:
            response.encoding = "utf-8"
        except Exception:
            pass
        streamed_chunks = []
        try:
            for raw_line in response.iter_lines(decode_unicode=True):
                if raw_line is None:
                    continue
                line = str(raw_line).strip()
                if not line:
                    continue
                if line.startswith("data:"):
                    line = line[5:].strip()
                if line == "[DONE]":
                    break
                if line.startswith("{") and line.endswith("}"):
                    try:
                        chunk_obj = json.loads(line)
                        content = extract_llm_content(chunk_obj)
                        if content:
                            streamed_chunks.append(content)
                        continue
                    except Exception:
                        pass
                streamed_chunks.append(line)
        except Exception as e:
            logger.warning("治理流式读取失败，回退到完整响应: %s", e)
        response_text = "\n".join(piece for piece in streamed_chunks if piece)

    if not response_text:
        try:
            response.encoding = "utf-8"
        except Exception:
            pass
        if hasattr(response, "json"):
            try:
                payload = response.json()
                response_text = extract_llm_content(payload)
            except Exception:
                response_text = ""
        if not response_text:
            try:
                response_text = response.text or ""
            except Exception:
                response_text = ""

    return extract_governance_objects_from_text(response_text)


def governance_error_response(detail: str, status_code: int = 500) -> JSONResponse:
    return JSONResponse(
        status_code=status_code,
        content={
            "code": 1,
            "message": detail,
            "detail": detail,
        },
    )


def normalize_governed_root(root: dict) -> dict:
    domain_code = str(root.get('domain_code') or 'pub').strip().lower()
    if domain_code not in BUSINESS_DOMAINS:
        domain_code = 'pub'

    business_domain = str(root.get('business_domain') or BUSINESS_DOMAINS.get(domain_code, '通用')).strip()
    chinese_name = str(root.get('chinese_name') or '').strip()
    full_root = str(root.get('full_root') or '').strip()
    abbr_root = str(root.get('abbr_root') or '').strip()
    recommended_type = str(root.get('recommended_type') or '').strip()
    governance_status = str(root.get('governance_status') or 'governed').strip()
    merged_from = root.get('merged_from') if isinstance(root.get('merged_from'), list) else []

    return {
        'domain_code': domain_code,
        'business_domain': business_domain,
        'chinese_name': chinese_name,
        'full_root': full_root,
        'abbr_root': abbr_root,
        'recommended_type': recommended_type,
        'governance_status': governance_status,
        'merged_from': merged_from,
    }


def run_governance_llm_round(
    *,
    phase_label: str,
    api_url: str,
    api_key: str,
    model: str,
    temperature: float,
    prompt: str,
    timeout_seconds: int,
    prompt_kind: str,
    task_id: Optional[str] = None,
    chunk_index: Optional[int] = None,
    total_chunks: Optional[int] = None,
    merge_round: Optional[int] = None,
    item_count: Optional[int] = None,
) -> dict:
    llm_url = build_llm_url(api_url)
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": get_system_prompt()},
            {"role": "user", "content": prompt},
        ],
        "max_tokens": 4096,
        "temperature": temperature,
    }
    payload = add_deepseek_thinking_body(api_url, payload)
    prompt_length = len(prompt)
    prompt_lines = prompt.count("\n") + 1 if prompt else 0
    prompt_hash = hashlib.sha256(prompt.encode("utf-8")).hexdigest()[:16] if prompt else ""
    payload_length = len(json.dumps(payload, ensure_ascii=False))
    logger.info(
        "%s请求准备: llm_url=%s, model=%s, temperature=%s, timeout=%ss, prompt_kind=%s, prompt_chars=%s, prompt_lines=%s, prompt_hash=%s, payload_chars=%s",
        phase_label,
        llm_url,
        model,
        temperature,
        timeout_seconds,
        prompt_kind,
        prompt_length,
        prompt_lines,
        prompt_hash,
        payload_length,
    )
    log_governance_event(
        "llm_request_prepared",
        task_id=task_id,
        phase=phase_label,
        chunk_index=chunk_index,
        total_chunks=total_chunks,
        merge_round=merge_round,
        prompt_kind=prompt_kind,
        prompt_length=prompt_length,
        prompt_hash=prompt_hash,
        item_count=item_count,
        extra={
            "model": model,
            "timeout_s": timeout_seconds,
            "payload_chars": payload_length,
            "temperature": temperature,
        },
    )
    started_at = datetime.now()
    try:
        response = requests.post(
            llm_url,
            headers=headers,
            json=payload,
            timeout=timeout_seconds,
            stream=True,
        )
        llm_elapsed_ms = int((datetime.now() - started_at).total_seconds() * 1000)
        objects = extract_governance_objects_from_response(response)
        response_text_length = 0
        logger.info(
            "%s上游响应: status=%s, elapsed_ms=%s, response_items=%s",
            phase_label,
            response.status_code,
            llm_elapsed_ms,
            len(objects),
        )
        log_governance_event(
            "llm_response_received",
            task_id=task_id,
            phase=phase_label,
            chunk_index=chunk_index,
            total_chunks=total_chunks,
            merge_round=merge_round,
            prompt_kind=prompt_kind,
            prompt_length=prompt_length,
            prompt_hash=prompt_hash,
            item_count=item_count,
            result_count=len(objects),
            elapsed_ms=llm_elapsed_ms,
            extra={"status_code": response.status_code},
        )
        if response.status_code != 200:
            detail = ""
            try:
                detail = response.text[:500] if response.text else ""
            except Exception:
                detail = ""
            logger.error("%s失败: HTTP %s %s", phase_label, response.status_code, detail)
            log_governance_event(
                "llm_http_error",
                task_id=task_id,
                phase=phase_label,
                chunk_index=chunk_index,
                total_chunks=total_chunks,
                merge_round=merge_round,
                prompt_kind=prompt_kind,
                prompt_length=prompt_length,
                prompt_hash=prompt_hash,
                item_count=item_count,
                elapsed_ms=llm_elapsed_ms,
                extra={"status_code": response.status_code, "detail_preview": detail[:120]},
            )
            return {
                "ok": False,
                "message": f"词根治理失败: HTTP {response.status_code}",
                "stage": phase_label,
                "prompt_kind": prompt_kind,
                "prompt_length": prompt_length,
                "llm_elapsed_ms": llm_elapsed_ms,
                "llm_url": llm_url,
            }
        normalized = [
            normalize_governed_root(item)
            for item in objects
            if isinstance(item, dict) and (item.get("full_root") or item.get("abbr_root") or item.get("chinese_name"))
        ]
        if not normalized:
            logger.error("%s返回为空或无法解析为 JSON 数组", phase_label)
            log_governance_event(
                "llm_empty_or_unparsed",
                task_id=task_id,
                phase=phase_label,
                chunk_index=chunk_index,
                total_chunks=total_chunks,
                merge_round=merge_round,
                prompt_kind=prompt_kind,
                prompt_length=prompt_length,
                prompt_hash=prompt_hash,
                item_count=item_count,
                elapsed_ms=llm_elapsed_ms,
            )
            return {
                "ok": False,
                "message": "词根治理失败: 模型返回为空或结果无法解析",
                "stage": phase_label,
                "prompt_kind": prompt_kind,
                "prompt_length": prompt_length,
                "llm_elapsed_ms": llm_elapsed_ms,
                "llm_url": llm_url,
            }

        log_governance_event(
            "llm_round_succeeded",
            task_id=task_id,
            phase=phase_label,
            chunk_index=chunk_index,
            total_chunks=total_chunks,
            merge_round=merge_round,
            prompt_kind=prompt_kind,
            prompt_length=prompt_length,
            prompt_hash=prompt_hash,
            item_count=item_count,
            result_count=len(normalized),
            elapsed_ms=llm_elapsed_ms,
        )
        return {
            "ok": True,
            "data": normalized,
            "stage": phase_label,
            "prompt_kind": prompt_kind,
            "prompt_length": prompt_length,
            "prompt_lines": prompt_lines,
            "prompt_hash": prompt_hash,
            "llm_elapsed_ms": llm_elapsed_ms,
            "llm_url": llm_url,
        }
    except requests.exceptions.ReadTimeout as e:
        logger.error(
            "%s超时: %s | model=%s | api_url=%s | llm_url=%s | prompt_chars=%s | prompt_hash=%s",
            phase_label,
            e,
            model,
            api_url,
            llm_url,
            prompt_length,
            prompt_hash,
        )
        log_governance_event(
            "llm_timeout",
            task_id=task_id,
            phase=phase_label,
            chunk_index=chunk_index,
            total_chunks=total_chunks,
            merge_round=merge_round,
            prompt_kind=prompt_kind,
            prompt_length=prompt_length,
            prompt_hash=prompt_hash,
            item_count=item_count,
            extra={"error": str(e)},
        )
        return {
            "ok": False,
            "message": "词根治理失败: 上游模型响应超时，请稍后重试或减少词根规模",
            "stage": phase_label,
            "prompt_kind": prompt_kind,
            "prompt_length": prompt_length,
            "llm_url": llm_url,
        }
    except requests.exceptions.RequestException as e:
        logger.error(
            "%s请求异常: %s | model=%s | api_url=%s | llm_url=%s | prompt_chars=%s | prompt_hash=%s",
            phase_label,
            e,
            model,
            api_url,
            llm_url,
            prompt_length,
            prompt_hash,
        )
        log_governance_event(
            "llm_request_exception",
            task_id=task_id,
            phase=phase_label,
            chunk_index=chunk_index,
            total_chunks=total_chunks,
            merge_round=merge_round,
            prompt_kind=prompt_kind,
            prompt_length=prompt_length,
            prompt_hash=prompt_hash,
            item_count=item_count,
            extra={"error": str(e)},
        )
        return {
            "ok": False,
            "message": f"词根治理失败: {str(e)}",
            "stage": phase_label,
            "prompt_kind": prompt_kind,
            "prompt_length": prompt_length,
            "llm_url": llm_url,
        }
    except ValueError as e:
        logger.error(
            "%s结果解析失败: %s | model=%s | api_url=%s | llm_url=%s | prompt_hash=%s",
            phase_label,
            e,
            model,
            api_url,
            llm_url,
            prompt_hash,
        )
        log_governance_event(
            "llm_parse_error",
            task_id=task_id,
            phase=phase_label,
            chunk_index=chunk_index,
            total_chunks=total_chunks,
            merge_round=merge_round,
            prompt_kind=prompt_kind,
            prompt_hash=prompt_hash,
            item_count=item_count,
            extra={"error": str(e)},
        )
        return {
            "ok": False,
            "message": "词根治理失败: 模型返回内容无法解析",
            "stage": phase_label,
            "prompt_kind": prompt_kind,
            "prompt_length": prompt_length,
            "llm_url": llm_url,
        }
    except Exception as e:
        logger.exception(
            "%s未知异常 | model=%s | api_url=%s | llm_url=%s | prompt_chars=%s | prompt_hash=%s",
            phase_label,
            model,
            api_url,
            llm_url,
            prompt_length,
            prompt_hash,
        )
        log_governance_event(
            "llm_unknown_exception",
            task_id=task_id,
            phase=phase_label,
            chunk_index=chunk_index,
            total_chunks=total_chunks,
            merge_round=merge_round,
            prompt_kind=prompt_kind,
            prompt_length=prompt_length,
            prompt_hash=prompt_hash,
            item_count=item_count,
            extra={"error": str(e)},
        )
        return {
            "ok": False,
            "message": f"词根治理失败: {str(e)}",
            "stage": phase_label,
            "prompt_kind": prompt_kind,
            "prompt_length": prompt_length,
            "llm_url": llm_url,
        }


def run_governance_chunk(index: int, total_chunks: int, req: GovernanceRequest, prompt: str) -> Tuple[int, dict]:
    result = run_governance_llm_round(
        phase_label=f"词根治理分批 {index}/{total_chunks}",
        api_url=req.llm_config.api_url,
        api_key=req.llm_config.api_key,
        model=req.llm_config.model,
        temperature=req.llm_config.temperature or 0.3,
        prompt=prompt,
        timeout_seconds=600,
        prompt_kind="chunk",
        task_id=req.task_id,
        chunk_index=index,
        total_chunks=total_chunks,
    )
    return index, result

def build_llm_url(api_url: str) -> str:
    """构建LLM请求URL，避免重复拼接/chat/completions"""
    api_url = api_url.rstrip('/')
    if api_url.endswith('/chat/completions'):
        return api_url
    return f"{api_url}/chat/completions"




def is_deepseek_native_api(api_url: str) -> bool:
    normalized = (api_url or '').strip().rstrip('/').lower()
    if normalized.endswith('/chat/completions'):
        normalized = normalized[:-len('/chat/completions')].rstrip('/')
    return normalized == 'https://api.deepseek.com'


def add_deepseek_thinking_body(api_url: str, payload: dict) -> dict:
    """OpenAI SDK extra_body maps to top-level JSON fields for raw HTTP calls."""
    if is_deepseek_native_api(api_url):
        payload['thinking'] = {'type': 'disabled'}
    return payload


def extract_create_table_content(response_content: str) -> str:
    if not response_content:
        return ""
    match = re.search(r'CREATE\s+TABLE', response_content, re.IGNORECASE)
    if not match:
        return response_content.strip()
    sql = response_content[match.start():].strip()
    fence = sql.find('```')
    if fence > 0:
        sql = sql[:fence].strip()
    return sql

def clean_ddl_response(response_content: str, table_name: str = "") -> Tuple[bool, str]:
    """清理LLM响应，提取有效的DDL内容
    
    Args:
        response_content: LLM的原始响应内容
        table_name: 表名（用于日志）
    
    Returns:
        Tuple[bool, str]: (是否有效, 清理后的DDL内容)
    """
    if not response_content:
        logger.warning(f"表 {table_name} 的LLM响应为空")
        return False, f"-- 生成失败: LLM响应为空"
    
    create_table_pattern = r'\bCREATE\s+TABLE\b'
    
    if not re.search(create_table_pattern, response_content, re.IGNORECASE):
        logger.warning(f"表 {table_name} 的LLM响应中未找到CREATE TABLE语句")
        
        lines = response_content.split('\n')
        clean_lines = []
        
        descriptive_patterns = [
            r'^已就绪',
            r'^根据',
            r'^我将',
            r'^以下是',
            r'^提示：',
            r'^说明：',
            r'^注意：',
            r'^【.*】.*请',
            r'^已收到',
            r'^好的',
            r'^明白',
            r'^收到后将',
            r'^请检查'
        ]
        
        for line in lines:
            stripped = line.strip()
            
            should_skip = False
            for pattern in descriptive_patterns:
                if re.match(pattern, stripped):
                    should_skip = True
                    break
            
            if not should_skip and stripped and not stripped.startswith('--'):
                clean_lines.append(line)
        
        filtered_content = '\n'.join(clean_lines).strip()
        
        if filtered_content:
            logger.info(f"表 {table_name} 使用过滤后的响应内容，长度: {len(filtered_content)}")
            return True, filtered_content
        
        warning_patterns = [
            r'请提供',
            r'需要',
            r'请补充',
            r'缺少',
            r'请输入'
        ]
        
        for pattern in warning_patterns:
            if pattern in response_content:
                return False, f"-- LLM需要更多信息：{response_content.strip()}"
        
        return False, f"-- 生成失败: LLM未生成有效的DDL语句（原始响应: {response_content[:100]}...)"
    
    lines = response_content.split('\n')
    clean_lines = []
    in_create_table = False
    
    descriptive_patterns = [
        r'^已就绪',
        r'^请提供',
        r'^根据',
        r'^收到后',
        r'^我将',
        r'^以下是',
        r'^提示：',
        r'^说明：',
        r'^注意：',
        r'^【.*】.*请',
        r'^已收到',
        r'^好的',
        r'^明白',
        r'^收到后将',
        r'^请检查'
    ]
    
    def should_skip_line(line: str) -> bool:
        """判断是否应该跳过该行"""
        stripped = line.strip()
        
        for pattern in descriptive_patterns:
            if re.match(pattern, stripped):
                return True
        
        if stripped.startswith('--'):
            return True
        
        return False
    
    for line in lines:
        stripped = line.strip()
        
        if re.search(create_table_pattern, stripped, re.IGNORECASE):
            in_create_table = True
        
        if in_create_table:
            if should_skip_line(line):
                continue
            else:
                clean_lines.append(line)
        else:
            if should_skip_line(line):
                continue
            else:
                if stripped and not stripped.startswith('#'):
                    clean_lines.append(line)
    
    clean_content = '\n'.join(clean_lines).strip()
    
    if not clean_content:
        return False, f"-- 生成失败: DDL内容为空"
    
    if not re.search(create_table_pattern, clean_content, re.IGNORECASE):
        return False, f"-- 生成失败: 未找到有效的CREATE TABLE语句"
    
    return True, remove_primary_key_clauses(clean_content)


@app.post("/api/test-connection", response_model=TestConnectionResponse)
async def test_connection(req: TestConnectionRequest):
    logger.info(f"测试连接请求 - API URL: {req.api_url}, Model: {req.model}")
    start_time = datetime.now()
    try:
        headers = {
            "Authorization": f"Bearer {req.api_key}",
            "Content-Type": "application/json"
        }
        data = {
            "model": req.model,
            "messages": [{"role": "user", "content": "Hi"}],
            "max_tokens": 10
        }

        data = add_deepseek_thinking_body(req.api_url, data)
        response = requests_session.post(
            build_llm_url(req.api_url),
            headers=headers,
            json=data,
            timeout=30
        )

        elapsed = (datetime.now() - start_time).total_seconds()
        logger.info(f"请求完成，耗时: {elapsed:.2f}秒，状态码: {response.status_code}")

        if response.status_code == 200:
            return TestConnectionResponse(code=0, message="连接成功")
        else:
            return TestConnectionResponse(code=response.status_code, message=f"连接失败: {response.text}")
    except Exception as e:
        elapsed = (datetime.now() - start_time).total_seconds()
        logger.error(f"连接测试异常（耗时 {elapsed:.2f}秒）: {e}")
        return TestConnectionResponse(code=500, message=f"连接异常: {str(e)}")


@app.post("/api/generate-ddl", response_model=GenerateDDLResponse)
async def generate_ddl(req: GenerateDDLRequest):
    total_start_time = datetime.now()
    abbr_max_len = resolve_abbr_max_len(req.llm_config.abbr_max_len)
    
    logger.info("="*50)
    logger.info("开始生成DDL请求")
    logger.info(f"数据库类型: {req.db_type}")
    logger.info(f"建表需求: {req.description[:100]}...")

    step1_start = datetime.now()
    if req.word_roots_input and req.word_roots_input.content:
        logger.info("检测到用户上传了新词根，正在合并...")
        parsed_roots = []
        if req.word_roots_input.type == "text":
            parsed_roots = parse_word_roots_text(req.word_roots_input.content)
        elif req.word_roots_input.type == "file_base64":
            try:
                file_data = base64.b64decode(req.word_roots_input.content)
                parsed_roots = parse_excel_content(file_data)
            except Exception as e:
                logger.error(f"词根文件解析失败: {e}")

        if parsed_roots:
            merge_and_save_roots(parsed_roots)
            logger.info(f"已合并 {len(parsed_roots)} 条新词根")

    step1_elapsed = (datetime.now() - step1_start).total_seconds()

    step2_start = datetime.now()
    filtered_roots, word_roots_content = filter_and_prepare_roots(req.description, req.root_match_priority)
    logger.info(f"词根筛选完成: {len(filtered_roots)} 条")
    step2_elapsed = (datetime.now() - step2_start).total_seconds()

    db_type_name = {'mysql': 'MySQL', 'postgresql': 'PostgreSQL', 'oracle': 'Oracle'}.get(req.db_type, req.db_type)

    step3_start = datetime.now()
    
    standards = load_standards()
    industry_context = normalize_industry_context(req.llm_config.industry_context)
    standards_content = merge_standards_with_industry(standards.get('content', ''), industry_context)
    
    root_match_name = {'full': '全称', 'abbr': '缩写'}.get(req.root_match_priority, req.root_match_priority)
    
    db_example_config = get_db_example(req.db_type, req.root_match_priority)
    db_example = f"表名格式: {db_example_config['table_format']}\n\n示例DDL:\n{db_example_config['example']}"
    industry_context_block = build_industry_context_block(industry_context)
    
    root_constraints = get_root_constraints(req.root_match_priority, abbr_max_len)
    root_reuse_principle = get_root_reuse_principle(req.root_match_priority, abbr_max_len)
    theme_prefix = infer_theme_prefix(req.description, req.db_type, standards_content=standards_content)
    if theme_prefix:
        theme_prefix_block = (
            f"\n\n【主题域缩写】{theme_prefix}"
            f"\n【主题域缩写规则】\n{render_theme_prefix_guide(standards_content)}"
            f"\n【表名要求】表名必须使用 分层前缀_业务域_业务名 的结构，例如 dwd_{theme_prefix}_order_detail。"
        )
    else:
        theme_prefix_block = (
            "\n\n【表名要求】表名必须使用 分层前缀_业务域_业务名 的结构。"
            "\n未明确命中业务域时，不要默认使用 pub；只能根据表名语义选择一个业务域。"
        )
    
    if req.custom_prompt:
        logger.info("使用自定义提示词")
        try:
            prompt = req.custom_prompt.format(
                word_roots_content=word_roots_content,
                description=req.description,
                industry_context=industry_context,
                industry_context_block=industry_context_block,
                db_type=db_type_name,
                root_match_priority=root_match_name,
                standards_content=standards_content,
                db_example=db_example,
                root_constraints=root_constraints,
                root_reuse_principle=root_reuse_principle
            )
        except KeyError:
            logger.warning("自定义提示词中缺少必要的占位符，将附加建表需求")
            prompt = f"{req.custom_prompt}\n\n【建表需求】\n{req.description}\n\n{industry_context_block}\n\n【数据库类型】\n{db_type_name}\n{db_example}\n\n【词根参考】\n{word_roots_content}"
        
        # 检查提示词是否包含建表需求信息，如果没有则添加
        if req.description and req.description not in prompt:
            logger.info("自定义提示词中未包含建表需求，自动附加")
            prompt = f"{req.custom_prompt}\n\n【建表需求】\n{req.description}\n\n{industry_context_block}\n\n【数据库类型】\n{db_type_name}\n{db_example}\n\n【词根参考】\n{word_roots_content}"
        prompt = append_root_mode_constraints(prompt, root_constraints, root_reuse_principle)
    else:
        logger.info("使用默认提示词")
        prompt = USER_PROMPT_TEMPLATE.format(
            word_roots_content=word_roots_content,
            description=req.description,
            industry_context_block=industry_context_block,
            db_type=db_type_name,
            root_match_priority=root_match_name,
            standards_content=standards_content,
            db_example=db_example,
            root_constraints=root_constraints,
            root_reuse_principle=root_reuse_principle
        )

    prompt = f"{prompt}{theme_prefix_block}"
    prompt = append_new_roots_instruction(prompt)
    logger.info(f"提示词长度: {len(prompt)} 字符")
    step3_elapsed = (datetime.now() - step3_start).total_seconds()
    logger.info(f"发送请求到 LLM: {req.llm_config.api_url}")
    logger.info(f"数据库类型: {db_type_name}")
    logger.info(f"实际提示词内容:\n{prompt[:500]}...")

    llm_start_time = datetime.now()

    headers = {
        "Authorization": f"Bearer {req.llm_config.api_key}",
        "Content-Type": "application/json"
    }
    
    system_prompt = get_system_prompt()
    logger.info(f"System Prompt长度: {len(system_prompt)} 字符")
    
    data = {
        "model": req.llm_config.model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": 2048,
        "temperature": req.llm_config.temperature
    }
    
    data = add_deepseek_thinking_body(req.llm_config.api_url, data)
    max_retries = 3
    retry_delay = 5
    response_content = None
    
    for attempt in range(max_retries):
        try:
            logger.info(f"开始发送LLM请求（第 {attempt + 1} 次）...")
            
            response = requests_session.post(
                build_llm_url(req.llm_config.api_url),
                headers=headers,
                json=data,
                timeout=300
            )
            
            if response.status_code == 200:
                response.encoding = 'utf-8'
                result = response.json()
                response_content = extract_llm_content(result)
                logger.info(f"LLM响应成功，内容长度: {len(response_content)} 字符")
                break
            elif response.status_code in [429, 500, 502, 503, 504]:
                logger.warning(f"LLM请求失败，状态码: {response.status_code}，准备重试...")
                if attempt < max_retries - 1:
                    import time
                    time.sleep(retry_delay * (attempt + 1))
            else:
                logger.error(f"LLM请求失败，状态码: {response.status_code}")
                break
        except Exception as e:
            logger.warning(f"LLM请求异常: {e}，准备重试...")
            if attempt < max_retries - 1:
                import time
                time.sleep(retry_delay * (attempt + 1))
    
    llm_elapsed = (datetime.now() - llm_start_time).total_seconds()
    step4_start = datetime.now()
    
    content = ""
    new_roots = []
    violations = []
    
    try:
        if response_content is not None:
            content = re.sub(r'```sql\s*', '', response_content)
            content = re.sub(r'\s*```', '\n', content)
            
            is_valid, clean_content = clean_ddl_response(content, "单独建表")
            
            if not is_valid:
                logger.warning(f"单独建表: {clean_content}")
                return GenerateDDLResponse(
                    code=1,
                    data=clean_content,
                    extracted_roots=[],
                    violations=[],
                    message=clean_content
                )
            
            content = clean_content
            
            all_roots = load_word_roots()
            new_roots = extract_new_roots_from_response(content, all_roots)
            content = remove_new_roots_markers(content)
            
            logger.info(f"检测到 {len(new_roots)} 个新词根，将返回给前端由用户确认保存")
            
            violations = []
            
            if req.enable_validation:
                validator = DDLValidator(
                    word_roots=load_word_roots(),
                    standards=load_standards(),
                    root_match_priority=req.root_match_priority,
                    abbr_max_len=abbr_max_len,
                )
                violations = validator.validate(content, req.db_type)
                
                if violations:
                    error_count = sum(1 for v in violations if v['level'] == 'error')
                    logger.info(f"检测到 {len(violations)} 个规范违规（{error_count} 个错误，{len(violations)-error_count} 个警告）")
                    
                    logger.info("【违规详情】")
                    for i, v in enumerate(violations, 1):
                        logger.info(f"  [{i}] [{v['level']}] {v['rule']}: {v['message']}")
                    
                    if error_count > 0:
                        max_correction_attempts = 3
                        correction_attempt = 0
                        correction_history = []
                        
                        while correction_attempt < max_correction_attempts and error_count > 0:
                            correction_attempt += 1
                            logger.info(f"开始第{correction_attempt}次纠正...")
                            
                            fix_prompt = validator.generate_fix_prompt(content, violations, correction_attempt)
                            llm_correction_start_time = datetime.now()
                            
                            data = {
                                "model": req.llm_config.model,
                                "messages": [
                                    {"role": "system", "content": system_prompt},
                                    {"role": "user", "content": fix_prompt}
                                ],
                                "max_tokens": 2048,
                                "temperature": req.llm_config.temperature
                            }
                            
                            data = add_deepseek_thinking_body(req.llm_config.api_url, data)
                            response = requests_session.post(
                                build_llm_url(req.llm_config.api_url),
                                headers=headers,
                                json=data,
                                timeout=300
                            )
                            
                            if response.status_code == 200:
                                response.encoding = 'utf-8'
                                result = response.json()
                                content = extract_llm_content(result)
                                content = re.sub(r'```sql\s*', '', content)
                                content = re.sub(r'\s*```', '\n', content)
                                
                                violations = validator.validate(content, req.db_type)
                                error_count = sum(1 for v in violations if v['level'] == 'error')
                                warning_count = len(violations) - error_count
                                
                                llm_correction_elapsed = (datetime.now() - llm_correction_start_time).total_seconds()
                                correction_history.append({
                                    'attempt': correction_attempt,
                                    'elapsed': llm_correction_elapsed,
                                    'error_count': error_count,
                                    'warning_count': warning_count
                                })
                                
                                if error_count > 0:
                                    logger.info(f"第{correction_attempt}次纠正后仍有 {error_count} 个错误")
                                else:
                                    logger.info(f"第{correction_attempt}次纠正成功，所有错误已修复")
                                    break
                            else:
                                logger.error(f"第{correction_attempt}次纠正请求失败: {response.status_code}")
                                break
                        
                        if correction_history:
                            logger.info("【修正历史】")
                            for record in correction_history:
                                logger.info(
                                    f"  第{record['attempt']}轮: "
                                    f"耗时{record['elapsed']:.2f}秒, "
                                    f"错误{record['error_count']}个, "
                                    f"警告{record['warning_count']}个"
                                )
                        
                        if error_count > 0 and correction_attempt >= max_correction_attempts:
                            logger.warning(f"经过{max_correction_attempts}轮纠正，仍存在 {error_count} 个错误")
            else:
                logger.info("跳过三层校验")
        
        step4_elapsed = (datetime.now() - step4_start).total_seconds()
        total_elapsed = (datetime.now() - total_start_time).total_seconds()
        
        logger.info("="*50)
        logger.info("【性能分析】各阶段耗时:")
        logger.info(f"  本地处理-词根合并: {step1_elapsed:.3f}秒")
        logger.info(f"  本地处理-词根筛选: {step2_elapsed:.3f}秒")
        logger.info(f"  本地处理-Prompt组装: {step3_elapsed:.3f}秒")
        logger.info(f"  LLM模型推理耗时: {llm_elapsed:.2f}秒")
        logger.info(f"  本地处理-响应解析: {step4_elapsed:.3f}秒")
        logger.info(f"  总耗时: {total_elapsed:.2f}秒")
        logger.info("="*50)
        
        if response_content is None:
            return GenerateDDLResponse(code=1, data="", extracted_roots=[], violations=[], message="LLM request failed or returned no content")

        if not content.strip():
            return GenerateDDLResponse(code=1, data="", extracted_roots=[], violations=[], message="LLM returned no final content")

        response_data = {
            "code": 0 if not violations or sum(1 for v in violations if v['level'] == 'error') == 0 else 1,
            "data": content,
            "extracted_roots": new_roots,
            "violations": violations
        }
        
        return GenerateDDLResponse(
            code=response_data["code"],
            data=content,
            extracted_roots=new_roots,
            violations=violations if violations else [],
            message="合规生成完成" if not violations else f"生成完成但存在 {len(violations)} 个违规项"
        )
    except Exception as e:
        llm_elapsed = (datetime.now() - llm_start_time).total_seconds()
        total_elapsed = (datetime.now() - total_start_time).total_seconds()
        logger.error(f"请求异常（耗时 {total_elapsed:.2f}秒，LLM耗时 {llm_elapsed:.2f}秒）: {e}")
        logger.error(f"堆栈信息: {traceback.format_exc()}")
        return GenerateDDLResponse(code=500, data="", message=f"请求异常: {str(e)}")


@app.post("/api/parse-file")
async def parse_file(file: UploadFile = File(...), parse_type: str = Form("auto")):
    logger.info(f"解析文件请求: {file.filename}, 解析类型: {parse_type}")
    content = await file.read()
    filename = file.filename or ""

    try:
        if filename.endswith(".xlsx"):
            if parse_type == "roots":
                parsed = parse_excel_content(content)
            else:
                parsed = parse_excel_to_text(content)
            return {"code": 0, "data": parsed}
        elif filename.endswith(".docx"):
            parsed = parse_docx_content(content)
        elif filename.endswith(".csv"):
            parsed = parse_csv_roots(content) if parse_type == "roots" else parse_csv_content(content)
        elif filename.endswith(".txt"):
            text = content.decode("utf-8")
            parsed = parse_word_roots_text(text) if parse_type == "roots" else text
        else:
            raise HTTPException(status_code=400, detail="不支持的文件类型")

        return {"code": 0, "data": parsed}
    except Exception as e:
        logger.error(f"文件解析失败: {e}")
        raise HTTPException(status_code=500, detail=f"文件解析失败: {str(e)}")


@app.get("/api/download-template")
async def download_template():
    logger.info("下载词根模板请求")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "词根模板"

    headers = ["业务域", "域编码", "中文名称", "全称词根", "缩写词根", "推荐类型"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    example_rows = [
        ["客户", "cust", "客户", "customer", "cust", "VARCHAR(64)"],
        ["订单", "ord", "订单", "order", "ord", "INT"],
        ["产品", "prod", "产品", "product", "prod", "VARCHAR(128)"]
    ]
    for row_num, row_data in enumerate(example_rows, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_data

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    return StreamingResponse(
        virtual_workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=词根模板.xlsx"}
    )


@app.get("/api/download-batch-template")
async def download_batch_template():
    logger.info("下载批量建表模板请求")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批量建表"

    headers = ["表名", "表分层", "主题域", "字段名", "推荐字段类型"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)

    example_rows = [
        ["订单明细表", "dwd", "order", "订单ID", "VARCHAR(64)"],
        ["订单明细表", "dwd", "", "用户ID", "VARCHAR(64)"],
        ["订单明细表", "dwd", "", "订单金额", "DECIMAL(12,2)"],
        ["订单明细表", "dwd", "", "下单时间", "DATETIME"],
        ["产品维度表", "dim", "prod", "产品ID", "VARCHAR(64)"],
        ["产品维度表", "dim", "", "产品名称", "VARCHAR(128)"],
        ["产品维度表", "dim", "", "产品分类", "VARCHAR(32)"],
        ["销售汇总表", "dws", "mkt", "统计日期", "DATE"],
        ["销售汇总表", "dws", "", "销售总额", "DECIMAL(15,2)"],
    ]
    for row_num, row_data in enumerate(example_rows, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_data

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    return StreamingResponse(
        virtual_workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''%E6%89%B9%E9%87%8F%E5%BB%BA%E8%A1%A8%E6%A8%A1%E6%9D%BF.xlsx"}
    )


@app.get("/api/word-roots", response_model=WordRootsResponse)
async def get_word_roots():
    logger.info("获取词根列表请求")
    roots = load_word_roots()
    return WordRootsResponse(code=0, data=roots)


@app.post("/api/word-roots")
async def save_word_roots_api(roots: List[Dict[str, Any]] = Body(...)):
    logger.info(f"保存词根请求: {len(roots)} 条")
    save_word_roots(roots)
    return {"code": 0, "message": "词根保存成功"}


@app.delete("/api/word-roots")
async def clear_word_roots():
    logger.info("清除所有标准词根请求")
    save_word_roots([])
    return {"code": 0, "message": "标准词根已清除"}



@app.get("/api/standard-roots")
async def get_standard_roots():
    roots = load_standard_roots()
    historical_roots = load_effective_historical_roots()
    enriched_roots = enrich_standard_roots_with_historical_types(roots, historical_roots)
    if enriched_roots != roots:
        save_standard_roots(enriched_roots)
        roots = load_standard_roots()
    else:
        roots = enriched_roots
    return {"code": 0, "data": roots}


@app.put("/api/standard-roots/{root_id}")
async def update_standard_root(root_id: str, payload: Dict[str, Any] = Body(...)):
    roots = load_standard_roots()
    index = next((i for i, root in enumerate(roots) if str(root.get("root_id") or "") == root_id), -1)
    if index == -1:
        return {"code": 1, "message": "标准词根不存在"}

    existing = roots[index]
    updated = normalize_standard_root_record({**existing, **payload, "root_id": root_id}, existing)
    roots[index] = updated
    save_standard_roots(roots)
    return {"code": 0, "message": "标准词根更新成功", "data": updated}


@app.delete("/api/standard-roots/{root_id}")
async def delete_standard_root(root_id: str):
    roots = load_standard_roots()
    filtered = [root for root in roots if str(root.get("root_id") or "") != root_id]
    if len(filtered) == len(roots):
        return {"code": 1, "message": "标准词根不存在"}
    save_standard_roots(filtered)
    return {"code": 0, "message": "标准词根删除成功", "data": filtered}

@app.get("/api/historical-roots")
async def get_historical_roots():
    roots = load_effective_historical_roots()
    return {"code": 0, "data": roots}

@app.post("/api/historical-roots")
async def save_historical_roots_api(roots: List[Dict[str, Any]] = Body(...)):
    save_effective_historical_roots(roots)
    return {"code": 0, "message": "historical roots saved"}

@app.get("/api/business-domains")
async def get_business_domains():
    return {"code": 0, "data": BUSINESS_DOMAINS}

@app.post("/api/governance/preview")
async def governance_preview():
    try:
        historical = load_effective_historical_roots()
        return {"code": 0, "data": {"count": len(historical), "roots": historical}}
    except Exception as e:
        logger.exception("词根治理预览失败")
        return governance_error_response(f"词根治理失败: {str(e)}")


@app.get("/api/governance/progress/{task_id}")
async def governance_progress(task_id: str):
    progress = get_governance_progress(task_id)
    if not progress:
        return {"code": 1, "message": "治理任务不存在", "data": None}
    return {"code": 0, "data": progress}


@app.post("/api/governance/run")
async def governance_run(req: GovernanceRequest):
    task_id = req.task_id or str(uuid.uuid4())
    industry_context = normalize_industry_context(req.llm_config.industry_context)
    if req.task_id != task_id:
        req.task_id = task_id
    try:
        log_governance_event(
            "task_started",
            task_id=task_id,
            phase="governance_run",
            extra={
                "requested_workers": req.max_workers,
                "model": req.llm_config.model,
                "industry_context_chars": len(industry_context),
            },
        )
        set_governance_progress(
            task_id,
            status="running",
            stage="loading_history",
            message="正在加载历史词根",
            raw_root_count=0,
            prepared_root_count=0,
            filtered_root_count=0,
            excluded_standard_count=0,
            standard_root_count=0,
            chunk_count=0,
            requested_workers=req.max_workers,
            actual_workers=0,
            completed_chunks=0,
            current_chunk=0,
            active_chunks=[],
            merge_started=False,
            final_candidate_count=0,
        )
        historical = load_effective_historical_roots()
        if not historical:
            log_governance_event(
                "task_no_historical_roots",
                task_id=task_id,
                phase="loading_history",
            )
            set_governance_progress(
                task_id,
                status="completed",
                stage="completed",
                message="暂无可治理的历史词根",
                raw_root_count=0,
                prepared_root_count=0,
                filtered_root_count=0,
                excluded_standard_count=0,
                standard_root_count=len(load_standard_roots()),
                chunk_count=0,
                actual_workers=0,
            )
            return {"code": 0, "message": "暂无可治理的历史词根", "data": []}

        started_at = datetime.now()
        prepared_historical = prepare_historical_roots_for_governance(historical)
        standard_roots = load_standard_roots()
        filter_result = filter_historical_roots_against_standard(prepared_historical, standard_roots)
        filtered_historical = filter_result["kept"]
        excluded_standard_count = filter_result["excluded_count"]
        log_governance_event(
            "roots_prepared",
            task_id=task_id,
            phase="preprocess",
            item_count=len(historical),
            result_count=len(filtered_historical),
            extra={
                "prepared_root_count": len(prepared_historical),
                "excluded_standard_count": excluded_standard_count,
                "standard_root_count": filter_result["standard_count"],
            },
        )
        logger.info(
            "词根治理输入: 原始 %s 条, 去重后 %s 条, 标准已存在剔除 %s 条, 最终待治理 %s 条, 标准词根 %s 条",
            len(historical),
            len(prepared_historical),
            excluded_standard_count,
            len(filtered_historical),
            filter_result["standard_count"],
        )
        set_governance_progress(
            task_id,
            stage="filtering_standard",
            message=f"已按标准词根过滤，剔除 {excluded_standard_count} 条冲突词根，剩余 {len(filtered_historical)} 条待治理",
            raw_root_count=len(historical),
            prepared_root_count=len(prepared_historical),
            filtered_root_count=len(filtered_historical),
            excluded_standard_count=excluded_standard_count,
            standard_root_count=filter_result["standard_count"],
        )

        if not filtered_historical:
            log_governance_event(
                "task_skipped_all_standardized",
                task_id=task_id,
                phase="preprocess",
                item_count=len(historical),
                result_count=0,
                extra={
                    "prepared_root_count": len(prepared_historical),
                    "excluded_standard_count": excluded_standard_count,
                },
            )
            set_governance_progress(
                task_id,
                status="completed",
                stage="completed",
                message="历史词根在去重后均已存在于标准词根，无需再次治理",
                chunk_count=0,
                actual_workers=0,
                completed_chunks=0,
                current_chunk=0,
                active_chunks=[],
                merge_started=False,
                final_candidate_count=0,
                elapsed_ms=int((datetime.now() - started_at).total_seconds() * 1000),
            )
            return {
                "code": 0,
                "message": "历史词根已被标准词根覆盖，无需再次治理",
                "data": [],
                "meta": {
                    "task_id": task_id,
                    "raw_root_count": len(historical),
                    "prepared_root_count": len(prepared_historical),
                    "filtered_root_count": 0,
                    "excluded_standard_count": excluded_standard_count,
                    "standard_root_count": filter_result["standard_count"],
                    "chunk_size": GOVERNANCE_CHUNK_SIZE,
                    "chunk_count": 0,
                    "requested_workers": max(1, req.max_workers),
                    "actual_workers": 0,
                    "chunk_candidate_count": 0,
                    "final_candidate_count": 0,
                    "prompt_length": 0,
                    "elapsed_ms": int((datetime.now() - started_at).total_seconds() * 1000),
                    "llm_elapsed_ms": 0,
                }
            }

        chunks = chunk_roots_for_governance(filtered_historical, GOVERNANCE_CHUNK_SIZE)
        requested_workers = max(1, req.max_workers)
        actual_workers = min(requested_workers, max(1, len(chunks)))
        chunk_item_counts = [len(chunk) for chunk in chunks]
        log_governance_event(
            "chunks_built",
            task_id=task_id,
            phase="chunking",
            item_count=len(filtered_historical),
            result_count=len(chunks),
            extra={
                "chunk_size": GOVERNANCE_CHUNK_SIZE,
                "requested_workers": requested_workers,
                "actual_workers": actual_workers,
                "chunk_item_counts": chunk_item_counts,
            },
        )
        set_governance_progress(
            task_id,
            stage="chunking",
            message="已完成切块，准备发起分批治理",
            raw_root_count=len(historical),
            prepared_root_count=len(prepared_historical),
            filtered_root_count=len(filtered_historical),
            excluded_standard_count=excluded_standard_count,
            standard_root_count=filter_result["standard_count"],
            chunk_count=len(chunks),
            chunk_size=GOVERNANCE_CHUNK_SIZE,
            requested_workers=requested_workers,
            actual_workers=actual_workers,
        )
        logger.info(
            "词根治理切块: chunk_size=%s, chunk_count=%s, requested_workers=%s, actual_workers=%s",
            GOVERNANCE_CHUNK_SIZE,
            len(chunks),
            requested_workers,
            actual_workers,
        )

        if len(chunks) == 1:
            single_prompt = build_governance_prompt(chunks[0], industry_context)
            single_result = run_governance_llm_round(
                phase_label="词根治理单批",
                api_url=req.llm_config.api_url,
                api_key=req.llm_config.api_key,
                model=req.llm_config.model,
                temperature=req.llm_config.temperature or 0.3,
                prompt=single_prompt,
                timeout_seconds=600,
                prompt_kind="chunk",
                task_id=task_id,
                chunk_index=1,
                total_chunks=1,
                item_count=len(chunks[0]),
            )
            if not single_result.get("ok"):
                set_governance_progress(
                    task_id,
                    status="failed",
                    stage="chunk_failed",
                    message=single_result.get("message", "词根治理失败"),
                )
                return governance_error_response(single_result.get("message", "词根治理失败"))

            normalized = dedupe_governed_roots(single_result.get("data", []))
            total_elapsed_ms = int((datetime.now() - started_at).total_seconds() * 1000)
            log_governance_event(
                "task_completed_single_chunk",
                task_id=task_id,
                phase="completed",
                chunk_index=1,
                total_chunks=1,
                item_count=len(chunks[0]),
                result_count=len(normalized),
                elapsed_ms=total_elapsed_ms,
            )
            set_governance_progress(
                task_id,
                status="completed",
                stage="completed",
                message=f"治理完成，生成 {len(normalized)} 条候选标准词根",
                completed_chunks=1,
                current_chunk=1,
                active_chunks=[],
                merge_started=False,
                final_candidate_count=len(normalized),
                elapsed_ms=total_elapsed_ms,
            )
            return {
                "code": 0,
                "message": f"治理完成，生成 {len(normalized)} 条标准词根候选",
                "data": normalized,
                "meta": {
                    "task_id": task_id,
                    "raw_root_count": len(historical),
                    "prepared_root_count": len(prepared_historical),
                    "filtered_root_count": len(filtered_historical),
                    "excluded_standard_count": excluded_standard_count,
                    "standard_root_count": filter_result["standard_count"],
                    "chunk_size": GOVERNANCE_CHUNK_SIZE,
                    "chunk_count": 1,
                    "requested_workers": requested_workers,
                    "actual_workers": 1,
                    "chunk_candidate_count": len(normalized),
                    "final_candidate_count": len(normalized),
                    "prompt_length": single_result.get("prompt_length", len(single_prompt)),
                    "elapsed_ms": total_elapsed_ms,
                    "llm_elapsed_ms": single_result.get("llm_elapsed_ms", 0),
                }
            }

        loop = asyncio.get_running_loop()
        with ThreadPoolExecutor(max_workers=actual_workers) as governance_executor:
            chunk_tasks = []
            for index, chunk in enumerate(chunks, start=1):
                chunk_prompt = build_governance_prompt(chunk, industry_context)
                log_governance_event(
                    "chunk_scheduled",
                    task_id=task_id,
                    phase="chunk_running",
                    chunk_index=index,
                    total_chunks=len(chunks),
                    prompt_kind="chunk",
                    prompt_length=len(chunk_prompt),
                    item_count=len(chunk),
                )
                future = loop.run_in_executor(
                    governance_executor,
                    partial(run_governance_chunk, index, len(chunks), req, chunk_prompt),
                )
                chunk_tasks.append((index, future))

            set_governance_progress(
                task_id,
                stage="chunk_running",
                message=f"正在并发治理，共 {len(chunks)} 批，当前并发 {actual_workers} 个线程",
                active_chunks=[index for index, _ in chunk_tasks[:actual_workers]],
            )

            chunk_results = []
            completed_chunks = 0
            active_chunks = [index for index, _ in chunk_tasks[:actual_workers]]
            for completed_future in asyncio.as_completed([future for _, future in chunk_tasks]):
                index, result = await completed_future
                completed_chunks += 1
                chunk_results.append((index, result))
                log_governance_event(
                    "chunk_finished",
                    task_id=task_id,
                    phase="chunk_running",
                    chunk_index=index,
                    total_chunks=len(chunks),
                    prompt_kind=result.get("prompt_kind"),
                    prompt_length=result.get("prompt_length"),
                    result_count=len(result.get("data", [])) if result.get("ok") else 0,
                    elapsed_ms=result.get("llm_elapsed_ms"),
                    extra={"ok": result.get("ok", False), "stage": result.get("stage", "")},
                )
                if index in active_chunks:
                    active_chunks.remove(index)
                next_chunk_number = completed_chunks + actual_workers
                if next_chunk_number <= len(chunks):
                    active_chunks.append(next_chunk_number)
                set_governance_progress(
                    task_id,
                    stage="chunk_running",
                    message=f"正在处理第 {completed_chunks}/{len(chunks)} 批，当前活跃批次: {', '.join(map(str, active_chunks)) if active_chunks else '无'}",
                    completed_chunks=completed_chunks,
                    current_chunk=completed_chunks,
                    active_chunks=active_chunks,
                )

            for _, result in chunk_results:
                if not result.get("ok"):
                    set_governance_progress(
                        task_id,
                        status="failed",
                        stage="chunk_failed",
                        message=result.get("message", "词根治理失败"),
                    )
                    return governance_error_response(result.get("message", "词根治理失败"))

            chunk_candidates = []
            chunk_prompt_lengths = []
            chunk_llm_elapsed_ms = 0
            for _, result in chunk_results:
                chunk_candidates.extend(result.get("data", []))
                chunk_prompt_lengths.append(result.get("prompt_length", 0))
                chunk_llm_elapsed_ms += result.get("llm_elapsed_ms", 0)

            merged_candidates = dedupe_governed_roots(chunk_candidates)
            log_governance_event(
                "chunks_merged_locally",
                task_id=task_id,
                phase="chunk_merged",
                item_count=len(chunk_candidates),
                result_count=len(merged_candidates),
                elapsed_ms=chunk_llm_elapsed_ms,
                extra={"chunk_prompt_lengths": chunk_prompt_lengths},
            )
            logger.info(
                "词根治理分批完成: 原始候选=%s, 去重后候选=%s, 分批累计耗时=%sms",
                len(chunk_candidates),
                len(merged_candidates),
                chunk_llm_elapsed_ms,
            )
            if not merged_candidates:
                set_governance_progress(
                    task_id,
                    status="failed",
                    stage="chunk_empty",
                    message="词根治理失败: 分批结果为空",
                )
                return governance_error_response("词根治理失败: 分批结果为空")

            current_candidates = merged_candidates
            merge_round = 0
            merge_llm_elapsed_ms = 0
            while len(current_candidates) > GOVERNANCE_MERGE_CHUNK_SIZE:
                merge_round += 1
                merge_batches = chunk_roots_for_governance(current_candidates, GOVERNANCE_MERGE_CHUNK_SIZE)
                if len(merge_batches) <= 1:
                    break
                log_governance_event(
                    "merge_round_started",
                    task_id=task_id,
                    phase="merge_running",
                    merge_round=merge_round,
                    item_count=len(current_candidates),
                    result_count=len(merge_batches),
                    extra={
                        "merge_chunk_size": GOVERNANCE_MERGE_CHUNK_SIZE,
                        "batch_item_counts": [len(batch) for batch in merge_batches],
                    },
                )

                set_governance_progress(
                    task_id,
                    stage="merge_running",
                    message=f"分批治理完成，正在第 {merge_round} 轮压缩 {len(current_candidates)} 条候选词根",
                    merge_started=True,
                    active_chunks=[],
                )

                merge_futures = []
                for batch_index, batch in enumerate(merge_batches, start=1):
                    batch_prompt = build_governance_merge_prompt(batch, industry_context)
                    log_governance_event(
                        "merge_batch_scheduled",
                        task_id=task_id,
                        phase="merge_running",
                        chunk_index=batch_index,
                        total_chunks=len(merge_batches),
                        merge_round=merge_round,
                        prompt_kind="merge",
                        prompt_length=len(batch_prompt),
                        item_count=len(batch),
                    )
                    future = loop.run_in_executor(
                        governance_executor,
                        partial(
                            run_governance_llm_round,
                            phase_label=f"词根治理全局合并-第{merge_round}轮 {batch_index}/{len(merge_batches)}",
                            api_url=req.llm_config.api_url,
                            api_key=req.llm_config.api_key,
                            model=req.llm_config.model,
                            temperature=req.llm_config.temperature or 0.3,
                            prompt=batch_prompt,
                            timeout_seconds=600,
                            prompt_kind="merge",
                            task_id=task_id,
                            chunk_index=batch_index,
                            total_chunks=len(merge_batches),
                            merge_round=merge_round,
                            item_count=len(batch),
                        ),
                    )
                    merge_futures.append(future)

                batch_results = await asyncio.gather(*merge_futures)
                batch_candidates = []
                for result in batch_results:
                    if not result.get("ok"):
                        set_governance_progress(
                            task_id,
                            status="failed",
                            stage="merge_failed",
                            message=result.get("message", "词根治理失败"),
                        )
                        return governance_error_response(result.get("message", "词根治理失败"))
                    batch_candidates.extend(result.get("data", []))
                    merge_llm_elapsed_ms += result.get("llm_elapsed_ms", 0)

                next_candidates = dedupe_governed_roots(batch_candidates)
                log_governance_event(
                    "merge_round_finished",
                    task_id=task_id,
                    phase="merge_running",
                    merge_round=merge_round,
                    item_count=len(batch_candidates),
                    result_count=len(next_candidates),
                    elapsed_ms=merge_llm_elapsed_ms,
                )
                if len(next_candidates) >= len(current_candidates):
                    logger.warning(
                        "词根治理合并未进一步收敛: round=%s, before=%s, after=%s",
                        merge_round,
                        len(current_candidates),
                        len(next_candidates),
                    )
                    break
                current_candidates = next_candidates

            merge_prompt = build_governance_merge_prompt(current_candidates, industry_context)
            set_governance_progress(
                task_id,
                stage="merge_running",
                message=f"分批治理完成，正在全局合并 {len(current_candidates)} 条候选词根",
                merge_started=True,
                active_chunks=[],
            )
            merge_result = await loop.run_in_executor(
                governance_executor,
                partial(
                    run_governance_llm_round,
                    phase_label="词根治理全局合并",
                    api_url=req.llm_config.api_url,
                    api_key=req.llm_config.api_key,
                    model=req.llm_config.model,
                    temperature=req.llm_config.temperature or 0.3,
                    prompt=merge_prompt,
                    timeout_seconds=600,
                    prompt_kind="merge",
                    task_id=task_id,
                    merge_round=merge_round + 1 if merge_round else 0,
                    item_count=len(current_candidates),
                ),
            )
            if not merge_result.get("ok"):
                set_governance_progress(
                    task_id,
                    status="failed",
                    stage="merge_failed",
                    message=merge_result.get("message", "词根治理失败"),
                )
                return governance_error_response(merge_result.get("message", "词根治理失败"))

            final_roots = dedupe_governed_roots(merge_result.get("data", []))
            total_elapsed_ms = int((datetime.now() - started_at).total_seconds() * 1000)
            log_governance_event(
                "task_completed",
                task_id=task_id,
                phase="completed",
                item_count=len(filtered_historical),
                result_count=len(final_roots),
                elapsed_ms=total_elapsed_ms,
                extra={
                    "chunk_count": len(chunks),
                    "merge_rounds": merge_round,
                    "chunk_candidate_count": len(merged_candidates),
                },
            )
            set_governance_progress(
                task_id,
                status="completed",
                stage="completed",
                message=f"治理完成，生成 {len(final_roots)} 条候选标准词根",
                completed_chunks=len(chunks),
                current_chunk=len(chunks),
                active_chunks=[],
                final_candidate_count=len(final_roots),
                elapsed_ms=total_elapsed_ms,
                llm_elapsed_ms=chunk_llm_elapsed_ms + merge_llm_elapsed_ms + merge_result.get("llm_elapsed_ms", 0),
            )
            return {
                "code": 0,
                "message": f"治理完成，生成 {len(final_roots)} 条标准词根候选",
                "data": final_roots,
                "meta": {
                    "task_id": task_id,
                    "raw_root_count": len(historical),
                    "prepared_root_count": len(prepared_historical),
                    "filtered_root_count": len(filtered_historical),
                    "excluded_standard_count": excluded_standard_count,
                    "standard_root_count": filter_result["standard_count"],
                    "chunk_size": GOVERNANCE_CHUNK_SIZE,
                    "chunk_count": len(chunks),
                    "requested_workers": requested_workers,
                    "actual_workers": actual_workers,
                    "chunk_candidate_count": len(merged_candidates),
                    "final_candidate_count": len(final_roots),
                    "prompt_length": merge_result.get("prompt_length", len(merge_prompt)),
                    "elapsed_ms": total_elapsed_ms,
                    "llm_elapsed_ms": chunk_llm_elapsed_ms + merge_llm_elapsed_ms + merge_result.get("llm_elapsed_ms", 0),
                }
            }
    except Exception as e:
        logger.exception("词根治理未知异常")
        log_governance_event(
            "task_failed",
            task_id=task_id,
            phase="failed",
            extra={"error": str(e)},
        )
        set_governance_progress(
            task_id,
            status="failed",
            stage="failed",
            message=f"词根治理失败: {str(e)}",
        )
        return governance_error_response(f"词根治理失败: {str(e)}")

@app.post("/api/governance/apply")
async def governance_apply(governed_roots: list = Body(...)):
    normalized = [normalize_governed_root(item) for item in governed_roots if isinstance(item, dict)]
    result = apply_governance_result(normalized)
    return {"code": 0, "message": f"governance applied: {len(result)} standard roots", "data": result}

@app.delete("/api/historical-roots")
async def clear_historical_roots():
    try:
        save_effective_historical_roots([])
        return {"code": 0, "message": "historical roots cleared"}
    except Exception as e:
        return {"code": 1, "message": str(e)}

@app.post("/api/save-new-roots")
async def save_new_roots_to_historical(new_roots: list = Body(...)):
    for r in new_roots:
        add_to_historical_roots(r)
    return {"code": 0, "message": f"{len(new_roots)} roots saved to historical"}
@app.get("/api/history-fields", response_model=WordRootsResponse)
async def get_history_fields():
    logger.info("获取历史字段请求")
    roots = load_word_roots()
    fields = [{"field_name": r.get('full_root', ''), "chinese_name": r.get('chinese_name', '')} for r in roots]
    return WordRootsResponse(code=0, data=roots)


@app.get("/api/logs")
async def get_logs(lines: int = 100):
    logger.info(f"获取日志请求: 最近 {lines} 行")
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                recent_lines = all_lines[-lines:] if len(all_lines) > lines else all_lines
                return {"code": 0, "data": "".join(recent_lines), "total_lines": len(all_lines)}
        return {"code": 0, "data": "", "total_lines": 0, "message": "日志文件不存在"}
    except Exception as e:
        logger.error(f"读取日志失败: {e}")
        return {"code": 1, "message": f"读取日志失败: {str(e)}"}


@app.get("/api/logs/list")
async def list_log_files():
    logger.info("获取日志文件列表请求")
    try:
        if os.path.exists(LOG_DIR):
            files = [f for f in os.listdir(LOG_DIR) if f.endswith('.log')]
            files.sort(reverse=True)
            return {"code": 0, "data": files}
        return {"code": 0, "data": []}
    except Exception as e:
        logger.error(f"获取日志文件列表失败: {e}")
        return {"code": 1, "message": str(e)}


@app.get("/api/logs/file/{filename}")
async def get_log_file(filename: str):
    logger.info(f"读取日志文件: {filename}")
    file_path = os.path.join(LOG_DIR, filename)
    if not os.path.exists(file_path):
        return {"code": 1, "message": "文件不存在"}
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return {"code": 0, "data": content}
    except Exception as e:
        logger.error(f"读取日志文件失败: {e}")
        return {"code": 1, "message": str(e)}


@app.get("/api/standards")
async def get_standards():
    logger.info("获取开发规范列表请求")
    standards = load_all_standards()
    return {"code": 0, "data": standards}


@app.post("/api/standards")
async def create_standard(data: dict = Body(...)):
    logger.info(f"创建新规范请求")
    content = data.get('content', '')
    name = data.get('name', '未命名规范')
    standard_id = str(uuid.uuid4())[:8]
    success = save_standard_by_id(standard_id, content, name)
    if success:
        return {"code": 0, "message": "规范创建成功", "data": {"id": standard_id}}
    return {"code": 1, "message": "规范创建失败"}


@app.put("/api/standards/{standard_id}")
async def update_standard_by_id(standard_id: str, data: dict = Body(...)):
    logger.info(f"更新规范请求: {standard_id}")
    content = data.get('content', '')
    name = data.get('name', None)
    success = save_standard_by_id(standard_id, content, name)
    if success:
        return {"code": 0, "message": "规范更新成功"}
    return {"code": 1, "message": "规范更新失败"}


@app.delete("/api/standards/{standard_id}")
async def delete_standard_api(standard_id: str):
    logger.info(f"删除规范请求: {standard_id}")
    success = delete_standard_by_id(standard_id)
    if success:
        return {"code": 0, "message": "删除成功"}
    return {"code": 1, "message": "删除失败"}


@app.put("/api/standards/{standard_id}/activate")
async def activate_standard(standard_id: str):
    logger.info(f"启用规范请求: {standard_id}")
    
    # 处理默认规范的启用
    if standard_id == "default":
        try:
            # 清空默认规范文件内容，使其使用内置默认
            if os.path.exists(STANDARDS_FILE):
                with open(STANDARDS_FILE, 'r', encoding='utf-8-sig') as f:
                    data = json.load(f)
            else:
                data = {}
            
            data['content'] = ''
            data['last_modified'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            with open(STANDARDS_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            # 禁用所有其他规范
            if os.path.exists(STANDARDS_DIR):
                for filename in os.listdir(STANDARDS_DIR):
                    if filename.endswith('.json'):
                        filepath = os.path.join(STANDARDS_DIR, filename)
                        with open(filepath, 'r', encoding='utf-8-sig') as f:
                            file_data = json.load(f)
                        file_data['is_active'] = False
                        with open(filepath, 'w', encoding='utf-8') as f:
                            json.dump(file_data, f, ensure_ascii=False, indent=2)
            
            refresh_system_prompt_cache()
            logger.info(f"默认规范已启用")
            return {"code": 0, "message": "已启用默认规范"}
        except Exception as e:
            logger.error(f"启用默认规范失败: {e}")
            return {"code": 1, "message": f"启用默认规范失败: {str(e)}"}
    
    # 处理自定义规范的启用
    if not os.path.exists(STANDARDS_DIR):
        return {"code": 1, "message": "规范目录不存在"}
    
    try:
        for filename in os.listdir(STANDARDS_DIR):
            if filename.endswith('.json'):
                filepath = os.path.join(STANDARDS_DIR, filename)
                with open(filepath, 'r', encoding='utf-8-sig') as f:
                    data = json.load(f)
                
                file_id = filename.replace('.json', '')
                data['is_active'] = (file_id == standard_id)
                
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
        
        refresh_system_prompt_cache()
        logger.info(f"规范 {standard_id} 已启用")
        return {"code": 0, "message": "已启用规范"}
    except Exception as e:
        logger.error(f"启用规范失败: {e}")
        return {"code": 1, "message": f"启用规范失败: {str(e)}"}


@app.put("/api/standards/{standard_id}/deactivate")
async def deactivate_standard(standard_id: str):
    logger.info(f"禁用规范请求: {standard_id}")
    
    if not os.path.exists(STANDARDS_DIR):
        return {"code": 1, "message": "规范目录不存在"}
    
    try:
        filepath = os.path.join(STANDARDS_DIR, f"{standard_id}.json")
        if not os.path.exists(filepath):
            return {"code": 1, "message": "规范不存在"}
        
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        data['is_active'] = False
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        refresh_system_prompt_cache()
        logger.info(f"规范 {standard_id} 已禁用")
        return {"code": 0, "message": "已禁用规范"}
    except Exception as e:
        logger.error(f"设置默认规范失败: {e}")
        return {"code": 1, "message": "设置失败"}


@app.post("/api/standards/reset")
async def reset_standards():
    logger.info("重置开发规范为内置默认")
    builtin_content = seed_from_builtin(BUILTIN_STANDARDS_FILE, STANDARDS_FILE, "standards")
    if builtin_content:
        refresh_system_prompt_cache()
        logger.info("开发规范已重置为内置默认")
        return {"code": 0, "data": builtin_content, "message": "已重置为内置默认规范"}
    return {"code": 1, "message": "内置默认规范文件不可用"}


@app.post("/api/standards/upload")
async def upload_standards(file: UploadFile = File(...)):
    logger.info(f"上传规范文件请求: {file.filename}")
    content = await file.read()
    filename = file.filename or ""
    
    try:
        if filename.endswith(".md") or filename.endswith(".txt"):
            try:
                text_content = content.decode("utf-8")
            except UnicodeDecodeError:
                text_content = content.decode("gbk")
        elif filename.endswith(".docx"):
            text_content = parse_docx_content(content)
        elif filename.endswith(".xlsx"):
            text_content = parse_excel_to_text(content)
        elif filename.endswith(".csv"):
            text_content = parse_csv_content(content)
        else:
            raise HTTPException(status_code=400, detail="不支持的文件类型，支持 .md, .txt, .docx, .xlsx, .csv")
        
        save_standards(text_content)
        logger.info(f"规范文件上传成功，内容长度: {len(text_content)}")
        return {"code": 0, "message": "规范文件上传成功"}
    except Exception as e:
        logger.error(f"上传规范文件失败: {e}")
        raise HTTPException(status_code=500, detail=f"上传规范文件失败: {str(e)}")


def parse_batch_table_excel(file_content: bytes) -> dict:
    logger.info("解析批量建表Excel文件")
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    tables = {}
    subject_domain_values = {}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = None
        row_num = 0
        
        for row in ws.iter_rows(values_only=True):
            row_num += 1
            
            if row_num == 1:
                headers = [str(cell).strip() if cell else '' for cell in row]
                continue
            
            if not row or all(not cell for cell in row):
                continue
            
            row_data = {headers[i]: str(row[i]).strip() if row[i] else '' for i in range(len(headers))}
            
            table_name = row_data.get('表名', '')
            table_layer = row_data.get('表分层', '').lower()
            subject_domain = row_data.get('主题域', '').strip().lower()
            field_name = row_data.get('字段名', '')
            field_type = get_excel_field_type(row_data) or resolve_field_type(field_name)
            
            if not table_name or not field_name:
                continue
            
            if table_name not in tables:
                tables[table_name] = {
                    'layer': table_layer,
                    'fields': [],
                    'user_specified_subject_domain': ''
                }
                subject_domain_values[table_name] = set()
            
            if subject_domain:
                subject_domain_values.setdefault(table_name, set()).add(subject_domain)
            
            tables[table_name]['fields'].append({
                'name': field_name,
                'type': field_type
            })
    
    for table_name, domains in subject_domain_values.items():
        if len(domains) > 1:
            conflict_values = ", ".join(sorted(domains))
            raise ValueError(f"表[{table_name}]存在多个冲突的主题域值: {conflict_values}，请统一后重试")
        if len(domains) == 1:
            tables[table_name]['user_specified_subject_domain'] = next(iter(domains))
    
    logger.info(f"Excel解析完成，共 {len(tables)} 张表")
    return tables


def generate_table_semantic_description(table_name: str, table_info: dict) -> str:
    """为批量建表生成包含语义信息的描述"""
    layer = table_info.get('layer', '')
    fields = table_info.get('fields', [])
    user_subject_domain = table_info.get('user_specified_subject_domain', '')
    
    field_names = [f['name'] for f in fields]
    field_types = [f['type'] for f in fields]
    
    semantic_parts = [f"表名：{table_name}"]
    semantic_parts.append(f"数据分层：{layer}")
    if user_subject_domain:
        semantic_parts.append(f"主题域缩写：{user_subject_domain}")
    semantic_parts.append(f"表用途：根据表名「{table_name}」及相关字段推断业务用途")
    
    if field_names:
        semantic_parts.append(f"字段列表（共{len(field_names)}个）：")
        for i, (name, ftype) in enumerate(zip(field_names, field_types), 1):
            semantic_parts.append(f"  {i}. {name}（{ftype}）")
    
    field_keywords = "、".join(field_names[:10])
    if len(field_names) > 10:
        field_keywords += f"等共{len(field_names)}个字段"
    semantic_parts.append(f"核心字段：{field_keywords}")
    
    return "\n".join(semantic_parts)


def normalize_single_table_schema(parsed_schema: dict) -> dict:
    table_name = str(parsed_schema.get("table_name", "") or "").strip()
    if not table_name:
        raise ValueError("未提取出表名")

    fields = parsed_schema.get("fields")
    if not isinstance(fields, list) or not fields:
        raise ValueError("未提取出有效字段")

    normalized_fields = []
    for field in fields:
        if not isinstance(field, dict):
            continue
        field_name = str(field.get("name", "") or "").strip()
        if not field_name:
            continue
        normalized_fields.append({
            "name": field_name,
            "type": str(field.get("type", "") or "").strip()
        })

    if not normalized_fields:
        raise ValueError("未提取出有效字段")

    return {
        table_name: {
            "layer": str(parsed_schema.get("layer", "") or "").strip().lower(),
            "fields": normalized_fields,
            "user_specified_subject_domain": str(
                parsed_schema.get("user_specified_subject_domain", "") or ""
            ).strip().lower()
        }
    }


def parse_single_text_to_table_schema(description: str, llm_config, db_type: str) -> dict:
    prompt = f"""请从下面的中文建表需求中提取单张表的结构化定义，并严格只输出 JSON 对象，不要输出解释。

要求：
1. 只能提取 1 张表；如果文本中明显描述了多张表，也仍只允许返回 1 张表，否则视为失败。
2. JSON 格式必须为：
{{
  "table_name": "中文表名",
  "layer": "ods|dim|dwd|dws|ads|input|",
  "user_specified_subject_domain": "主题域缩写或空字符串",
  "fields": [
    {{"name": "中文字段名", "type": "字段类型或空字符串"}}
  ]
}}
3. 如果字段类型无法明确判断，type 返回空字符串。
4. 不要生成英文表名或英文字段名。

数据库类型：{db_type}

建表需求：
{description}
"""

    headers = {
        "Authorization": f"Bearer {llm_config.api_key}",
        "Content-Type": "application/json"
    }
    data = {
        "model": llm_config.model,
        "messages": [
            {"role": "system", "content": "你是结构化信息抽取助手，只返回合法 JSON 对象。"},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": 2048,
        "temperature": 0
    }
    data = add_deepseek_thinking_body(llm_config.api_url, data)
    response = requests_session.post(
        build_llm_url(llm_config.api_url),
        headers=headers,
        json=data,
        timeout=300
    )
    if response.status_code != 200:
        raise ValueError(f"输入解析请求失败: {response.text}")

    response.encoding = 'utf-8'
    result = response.json()
    parsed_schema = extract_json_object_from_text(extract_llm_content(result))
    if not parsed_schema:
        raise ValueError("输入解析未返回合法 JSON")
    if "tables" in parsed_schema:
        tables = parsed_schema.get("tables")
        if isinstance(tables, list) and len(tables) != 1:
            raise ValueError("输入解析返回了多张表")
        if isinstance(tables, list) and len(tables) == 1 and isinstance(tables[0], dict):
            parsed_schema = tables[0]
    return normalize_single_table_schema(parsed_schema)

def generate_single_table_ddl(
    table_name: str,
    table_info: dict,
    word_roots: list,
    db_type: str,
    root_priority: str,
    custom_prompt: str = "",
    semantic_description: str = "",
    industry_context: str = "",
    abbr_max_len: int = DEFAULT_ABBR_MAX_LEN,
) -> str:
    fields_text = []
    for field in table_info['fields']:
        fields_text.append(f"- {field['name']} ({field['type']})")
    
    fields_description = "\n".join(fields_text)
    layer = table_info['layer']
    
    standards = load_standards()
    industry_context = normalize_industry_context(industry_context)
    standards_content = merge_standards_with_industry(standards.get('content', ''), industry_context)
    
    root_match_name = {'full': '全称', 'abbr': '缩写'}.get(root_priority, root_priority)
    
    user_subject_domain = (table_info.get('user_specified_subject_domain') or '').strip().lower()
    subject_domain_line = f"\n主题域缩写：{user_subject_domain}" if user_subject_domain else ""
    description = semantic_description if semantic_description else f"表名：{table_name}\n数据分层：{layer}{subject_domain_line}\n字段信息：\n{fields_description}"
    
    db_type_lower = {'mysql': 'mysql', 'postgresql': 'postgresql', 'oracle': 'oracle'}.get(db_type.lower(), 'mysql')
    db_example_config = get_db_example(db_type_lower, root_priority)
    db_example = f"表名格式: {db_example_config['table_format']}\n\n示例DDL:\n{db_example_config['example']}"
    industry_context_block = build_industry_context_block(industry_context)
    
    root_constraints = get_root_constraints(root_priority, abbr_max_len)
    root_reuse_principle = get_root_reuse_principle(root_priority, abbr_max_len)
    theme_prefix = user_subject_domain or infer_theme_prefix(description, table_name, layer, standards_content=standards_content)
    layer_prefix = f"{layer}_" if layer else ""
    if theme_prefix:
        theme_prefix_block = (
            f"\n\n【主题域缩写】{theme_prefix}"
            f"\n【主题域缩写规则】\n{render_theme_prefix_guide(standards_content)}"
            f"\n【表名要求】表名必须使用 {layer_prefix}{theme_prefix}_业务名 的结构。"
        )
    else:
        theme_prefix_block = (
            f"\n\n【表名要求】表名必须使用 {layer_prefix}业务域_业务名 的结构。"
            "\n未明确命中业务域时，不要默认使用 pub；只能根据表名语义选择一个业务域。"
        )
    
    if custom_prompt:
        try:
            prompt = custom_prompt.format(
                table_name=table_name,
                layer=layer,
                fields_description=fields_description,
                description=description,
                industry_context=industry_context,
                industry_context_block=industry_context_block,
                db_type=db_type,
                word_roots_content=format_word_roots_for_prompt(word_roots, root_priority),
                standards_content=standards_content,
                root_match_priority=root_match_name,
                db_example=db_example,
                root_constraints=root_constraints,
                root_reuse_principle=root_reuse_principle
            )
        except KeyError:
            logger.warning("自定义提示词中缺少必要的占位符，将附加建表需求")
            prompt = f"{custom_prompt}\n\n【建表需求】\n{description}\n\n{industry_context_block}\n\n【数据库类型】\n{db_type}\n{db_example}\n\n【词根参考】\n{format_word_roots_for_prompt(word_roots, root_priority)}"
        
        # 检查提示词是否包含建表需求信息，如果没有则添加
        if description and description not in prompt:
            logger.info("自定义提示词中未包含建表需求，自动附加")
            prompt = f"{custom_prompt}\n\n【建表需求】\n{description}\n\n{industry_context_block}\n\n【数据库类型】\n{db_type}\n{db_example}\n\n【词根参考】\n{format_word_roots_for_prompt(word_roots, root_priority)}"
        prompt = append_root_mode_constraints(prompt, root_constraints, root_reuse_principle)
    else:
        prompt = f"""你是一位数据仓库专家。请根据以下参考信息生成 DDL，不要任何解释和任何思考过程：

【建表需求】
{description}

{industry_context_block}

【词根参考】
{format_word_roots_for_prompt(word_roots, root_priority)}

【数据库类型】
{db_type}
{db_example}

【开发规范】
{standards_content}

{root_constraints}

要求：
1. 严格按照建表需求创建表，确保所有字段都被正确创建。
2. 字段命名必须符合【词根命名规则】。
3. 严格遵循开发规范（表名，字段名等）。
4. 只输出 CREATE TABLE 和 COMMENT 语句，不要解释和思考过程。
5. 严格按照【数据库类型】中的格式生成DDL，包括表名格式和注释语法。
6. {root_reuse_principle}
7. 【新词根扩展】如果已保存的词根中检索不到要建表的字段，请按当前词根模式生成命名用词根，并在 SQL 输出结束后，单独一行按以下格式补全词根库信息：【新词根】词根全称(例如：create_time):词根缩写(例如：crt_tm):中文名称:推荐字段类型"""
    
    return append_new_roots_instruction(prompt)


def process_single_table_for_batch(
    table_name: str,
    table_info: dict,
    all_roots: list,
    db_type_name: str,
    root_match_priority: str,
    api_key: str,
    api_url: str,
    model: str,
    custom_prompt: str = "",
    semantic_description: str = "",
    industry_context: str = "",
    temperature: float = 0.3,
    abbr_max_len: int = DEFAULT_ABBR_MAX_LEN,
) -> Tuple[str, str, list]:
    """
    处理单张表的DDL生成（线程安全）
    
    返回: (table_name, ddl_content, new_roots_list)
    """
    try:
        prompt = generate_single_table_ddl(
            table_name,
            table_info,
            all_roots,
            db_type_name,
            root_match_priority,
            custom_prompt,
            semantic_description,
            industry_context,
            abbr_max_len,
        )
        
        system_prompt = get_system_prompt()
        
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ]
        
        data = {
            "model": model,
            "messages": messages,
            "max_tokens": 2048,
            "temperature": temperature
        }
        
        data = add_deepseek_thinking_body(api_url, data)
        response_content = None
        max_retries = 3
        retry_delay = 5
        
        for attempt in range(max_retries):
            try:
                response = requests_session.post(
                    build_llm_url(api_url),
                    headers=headers,
                    json=data,
                    timeout=300
                )
                
                if response.status_code == 200:
                    response.encoding = 'utf-8'
                    result = response.json()
                    response_content = extract_llm_content(result)
                    break
                elif response.status_code in [429, 500, 502, 503, 504]:
                    logger.warning(f"处理表 {table_name} 第 {attempt + 1} 次失败，状态码: {response.status_code}，将重试")
                    if attempt < max_retries - 1:
                        import time
                        time.sleep(retry_delay * (attempt + 1))
                else:
                    break
            except Exception as e:
                logger.warning(f"处理表 {table_name} 第 {attempt + 1} 次异常: {e}")
                if attempt < max_retries - 1:
                    import time
                    time.sleep(retry_delay * (attempt + 1))
        
        if response_content is not None:
            original_content = response_content
            
            response_content = re.sub(r'```sql\s*', '', response_content)
            response_content = re.sub(r'```\s*', '\n', response_content)
            response_content = re.sub(r'```sql', '', response_content)
            
            response_content = re.sub(r'^```\w*\s*', '', response_content, flags=re.MULTILINE)
            response_content = re.sub(r'\s*```$', '', response_content, flags=re.MULTILINE)
            
            response_content = response_content.strip()
            
            is_valid, clean_content = clean_ddl_response(response_content, table_name)
            
            if not is_valid:
                logger.warning(f"表 {table_name} 的LLM响应无效: {clean_content}")
                return table_name, clean_content, []
            
            response_content = clean_content
            
            create_table_count = len(re.findall(r'\bCREATE\s+TABLE\b', response_content, re.IGNORECASE))
            if create_table_count > 1:
                logger.warning(f"表 {table_name} 的LLM响应包含 {create_table_count} 个CREATE TABLE语句，可能存在问题")
            
            logger.info(f"表 {table_name} 的DDL已生成，长度: {len(response_content)}")
            
            new_roots = extract_new_roots_from_response(response_content, all_roots)
            clean_ddl = remove_new_roots_markers(response_content)
            return table_name, clean_ddl, new_roots
        
        elif response is not None and response.status_code != 200:
            return table_name, f"-- 生成失败: {response.text}", []
        else:
            return table_name, f"-- 生成失败: 多次尝试后仍失败", []
    
    except Exception as e:
        logger.error(f"处理表 {table_name} 时出错: {e}")
        return table_name, f"-- 生成失败: {str(e)}", []


def run_structured_ddl_task(task_id: str, tables: dict, api_key: str, api_url: str, model: str, db_type: str,
                            root_match_priority: str, custom_prompt: str = "", enable_validation: bool = True,
                            max_workers: int = 5, temperature: float = 0.3, cut_mode: str = "accurate",
                            abbr_max_len: int = DEFAULT_ABBR_MAX_LEN, industry_context: str = "",
                            history_type: str = "batch", history_description: str = "", input_stage_label: str = "解析Excel文件..."):
    start_time = datetime.now()
    abbr_max_len = resolve_abbr_max_len(abbr_max_len)
    standards = load_standards()
    industry_context = normalize_industry_context(industry_context)
    standards_content = merge_standards_with_industry(standards.get('content', ''), industry_context)

    table_names = list((tables or {}).keys())
    if not table_names:
        raise ValueError("未解析到有效的表数据")

    logger.info(f"待处理表: {table_names}")
    logger.info(f"结构化建表任务 {task_id} 开始执行，模式: {history_type}，最大并发: {max_workers}")

    combined_description = "批量建表，涵盖以下表和字段：" + "；".join(
        [f"{t_name}({','.join([f['name'] for f in t_info.get('fields', [])])})" for t_name, t_info in tables.items()]
    )
    filtered_roots, _ = filter_and_prepare_roots(combined_description, root_match_priority)
    logger.info(f"结构化建表词根筛选完成: 精选 {len(filtered_roots)} 条词根")

    from app.processors.field_processor import FieldProcessor

    all_history_roots_for_matching = load_word_roots()
    field_processor = FieldProcessor(
        api_key, api_url, model, temperature,
        task_id=task_id, progress_callback=update_progress,
        max_workers=max_workers,
        custom_prompt=custom_prompt,
        word_roots=all_history_roots_for_matching,
        root_match_priority=root_match_priority,
        cut_mode=cut_mode,
        standards=standards_content,
        industry_context=industry_context,
        abbr_max_len=abbr_max_len,
        input_stage_label=input_stage_label
    )

    results = {}
    errors = []
    field_stats = None
    all_new_roots = []

    try:
        tables_data, field_mapping, field_stats, root_translations = field_processor.build_field_mapping(tables_data=tables)
        logger.info(f"字段映射构建完成，共 {len(field_mapping)} 个字段映射")
        results = field_processor.generate_all_ddl(tables_data, field_mapping, db_type, root_match_priority, standards_content)
        field_level_errors = [
            f"{table_name}: {ddl_content.replace('-- 生成失败:', '').strip()}"
            for table_name, ddl_content in results.items()
            if isinstance(ddl_content, str) and ddl_content.startswith("-- 生成失败")
        ]
        if field_level_errors:
            errors.extend(field_level_errors)
            raise RuntimeError("字段级主流程生成DDL失败: " + "; ".join(field_level_errors))

        existing_chinese = {r.get('chinese_name', '') for r in load_word_roots()}
        for chinese_root, english_root in root_translations.items():
            if chinese_root not in existing_chinese:
                all_new_roots.append({
                    'chinese_name': chinese_root,
                    'full_root': english_root,
                    'abbr_root': english_root[:abbr_max_len] if len(english_root) > abbr_max_len else english_root,
                    'category': 'root',
                    'recommended_type': 'VARCHAR(255)',
                    'business_domain': '基础通用'
                })

        completed_field_progress = None
        if field_stats:
            completed_field_progress = {
                "phase": "field_generation_completed",
                "phase_label": "字段级处理完成",
                "unique_root_count": field_stats.get("total_fields", 0),
                "target_item_label": "字段映射",
                "total_items": len(field_mapping),
                "completed_items": len(field_mapping),
                "batch_count": 0,
                "completed_batches": 0,
                "thread_count": max_workers
            }
        update_progress(
            task_id,
            6,
            8,
            stage="✅ 字段级处理完成",
            matched_count=field_stats.get("matched_count") if field_stats else None,
            unmatched_count=field_stats.get("unmatched_count") if field_stats else None,
            total_fields=field_stats.get("total_fields") if field_stats else None,
            field_progress=completed_field_progress
        )
    except Exception as e:
        elapsed_time = (datetime.now() - start_time).total_seconds()
        logger.error(f"字段级处理失败，终止任务: {e}")
        task_results[task_id] = {
            "code": 1,
            "message": f"字段级主流程失败: {str(e)}",
            "elapsed_time": elapsed_time,
            "errors": errors.copy()
        }
        task_validation_flags.pop(task_id, None)
        return

    deduplicated_new_roots = []
    seen_chinese_names = set()
    for root in all_new_roots:
        if root['chinese_name'] not in seen_chinese_names:
            deduplicated_new_roots.append(root)
            seen_chinese_names.add(root['chinese_name'])

    with cache_lock:
        batch_cache[task_id] = {
            'table_names': table_names,
            'results': results.copy(),
            'errors': errors.copy(),
            'next_index': len(table_names),
            'db_type': db_type,
            'root_match_priority': root_match_priority,
            'new_roots': deduplicated_new_roots.copy()
        }

    all_ddl = "\n\n".join(results.values())
    validation_info = None
    valid_results = {
        table_name: ddl_content
        for table_name, ddl_content in results.items()
        if isinstance(ddl_content, str)
        and ddl_content.strip()
        and not ddl_content.startswith("-- 生成失败")
    }

    if enable_validation and valid_results:
        update_progress(task_id, 7, 8, stage="🔍 统一校验中...")
        logger.info(f"开始统一校验，共 {len(valid_results)} 张表")

        error_count = 0
        warning_count = 0
        all_violations = []
        corrected = False

        try:
            from app.validators.unified_validator import UnifiedValidator

            unified_validator = UnifiedValidator(
                word_roots=load_word_roots(),
                standards=load_standards(),
                root_match_priority=root_match_priority,
                abbr_max_len=abbr_max_len,
            )
            validation_result = unified_validator.validate_batch(valid_results)
            all_violations = validation_result['single_violations'] + validation_result['cross_violations']
            error_count = validation_result['error_count']
            warning_count = validation_result['warning_count']

            if error_count > 0:
                update_progress(task_id, 7, 8, stage="🔧 统一修正中...")
                try:
                    parsed_tables = unified_validator.parse_all_ddl(valid_results)
                    error_fields = unified_validator.extract_error_fields(
                        validation_result['single_violations'],
                        parsed_tables
                    )
                    if error_fields:
                        fix_prompt = unified_validator.generate_field_fix_prompt(error_fields)
                        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
                        fix_data = {
                            "model": model,
                            "messages": [
                                {"role": "system", "content": get_system_prompt()},
                                {"role": "user", "content": fix_prompt}
                            ],
                            "max_tokens": 2048,
                            "temperature": temperature
                        }
                        fix_data = add_deepseek_thinking_body(api_url, fix_data)
                        fix_response = requests_session.post(
                            build_llm_url(api_url),
                            headers=headers,
                            json=fix_data,
                            timeout=300
                        )
                        if fix_response.status_code == 200:
                            fix_response.encoding = 'utf-8'
                            fix_result = fix_response.json()
                            llm_output = extract_llm_content(fix_result)
                            fix_results = unified_validator.parse_field_fix_result(llm_output)
                            if fix_results:
                                corrected_results = unified_validator.reassemble_ddl(
                                    valid_results,
                                    fix_results,
                                    parsed_tables
                                )
                                results.update(corrected_results)
                                corrected = True
                        else:
                            logger.error(f"修正请求失败: {fix_response.status_code}")
                except Exception as e:
                    logger.error(f"统一修正异常: {e}")
                    update_progress(task_id, 7, 8, stage=f"⚠️ 修正失败: {str(e)[:50]}")

                all_ddl = "\n\n".join(results.values())
        except Exception as e:
            logger.error(f"统一校验或修正失败: {e}")
            update_progress(task_id, 7, 8, stage=f"❌ 校验失败: {str(e)[:50]}")

        validation_info = {
            "error_count": error_count,
            "warning_count": warning_count,
            "violations": all_violations[:50],
            "total_violations": len(all_violations),
            "corrected": corrected
        }

    elapsed_time = (datetime.now() - start_time).total_seconds()
    elapsed_time_str = f"{int(elapsed_time // 60)}分{int(elapsed_time % 60)}秒" if elapsed_time >= 60 else f"{elapsed_time:.2f}秒"

    with cache_lock:
        if task_id in batch_cache:
            del batch_cache[task_id]
            logger.info(f"任务 {task_id} 完成，已清除缓存")

    save_ddl_history({
        "type": history_type,
        "tables_count": len(tables),
        "success_count": len(tables) - len(errors),
        "ddl": all_ddl,
        "description": history_description,
        "db_type": db_type,
        "root_match_priority": root_match_priority,
        "new_roots": deduplicated_new_roots
    })

    result_data = {
        "tables": results,
        "total_tables": len(tables),
        "success_count": len(tables) - len(errors),
        "error_count": len(errors),
        "errors": errors,
        "full_ddl": all_ddl,
        "new_roots": deduplicated_new_roots,
        "new_roots_count": len(deduplicated_new_roots),
        "elapsed_time": elapsed_time,
        "elapsed_time_str": elapsed_time_str,
        "validation_info": validation_info
    }
    if field_stats:
        result_data["field_stats"] = field_stats

    task_results[task_id] = {
        "code": 0,
        "data": result_data
    }
    update_progress(task_id, 8, 8, stage="✅ 完成建表")
    task_validation_flags.pop(task_id, None)
    if task_id in progress_store:
        del progress_store[task_id]
        logger.info(f"任务 {task_id} 进度数据已清除")


def process_batch_task(task_id: str, content: bytes, api_key: str, api_url: str, model: str, db_type: str,
                       root_match_priority: str, custom_prompt: str = "", enable_validation: bool = True,
                       max_workers: int = 5, temperature: float = 0.3, cut_mode: str = "accurate",
                       abbr_max_len: int = DEFAULT_ABBR_MAX_LEN, industry_context: str = "", use_field_level: bool = True):
    start_time = datetime.now()
    try:
        tables = parse_batch_table_excel(content)
        if not tables:
            raise ValueError("未解析到有效的表数据")
    except Exception as e:
        elapsed_time = (datetime.now() - start_time).total_seconds()
        logger.error(f"Excel解析失败: {e}，耗时: {elapsed_time:.2f}秒")
        task_results[task_id] = {
            "code": 1,
            "message": f"Excel解析失败: {str(e)}",
            "elapsed_time": elapsed_time
        }
        task_validation_flags.pop(task_id, None)
        return

    run_structured_ddl_task(
        task_id, tables, api_key, api_url, model, db_type, root_match_priority,
        custom_prompt, enable_validation, max_workers, temperature, cut_mode,
        abbr_max_len, industry_context, history_type="batch",
        history_description=f"批量建表: {', '.join(list(tables.keys()))[:50]}...",
        input_stage_label="解析Excel文件..."
    )


def process_text_generate_task(req: TextGenerateTaskRequest):
    task_id = req.task_id or str(uuid.uuid4())
    start_time = datetime.now()
    update_progress(task_id, 1, 8, stage="输入解析中...", source_label="输入解析")

    try:
        tables = parse_single_text_to_table_schema(req.description, req.llm_config, req.db_type)
        if len(tables) != 1:
            raise ValueError("输入解析返回了多张表")
        update_progress(task_id, 1, 8, stage="输入解析完成", source_label="输入解析")
    except Exception as e:
        elapsed_time = (datetime.now() - start_time).total_seconds()
        logger.error(f"输入解析失败: {e}")
        task_results[task_id] = {
            "code": 1,
            "message": f"输入解析失败: {str(e)}",
            "elapsed_time": elapsed_time
        }
        task_validation_flags.pop(task_id, None)
        return

    run_structured_ddl_task(
        task_id, tables,
        req.llm_config.api_key, req.llm_config.api_url, req.llm_config.model,
        req.db_type, req.root_match_priority, req.custom_prompt, req.enable_validation,
        1, req.llm_config.temperature or 0.3, req.cut_mode, req.llm_config.abbr_max_len,
        req.llm_config.industry_context or "", history_type="single",
        history_description=req.description[:80],
        input_stage_label="输入解析结果整理中..."
    )


@app.post("/api/batch-generate-ddl")
async def batch_generate_ddl(
    file: UploadFile = File(...),
    api_key: str = Form(...),
    api_url: str = Form(...),
    model: str = Form(...),
    industry_context: str = Form(""),
    db_type: str = Form("mysql"),
    root_match_priority: str = Form("abbr"),
    task_id: str = Form(None),
    custom_prompt: str = Form(""),
    enable_validation: bool = Form(True),
    max_workers: int = Form(5),
    temperature: float = Form(0.3),
    cut_mode: str = Form("accurate"),
    abbr_max_len: int = Form(DEFAULT_ABBR_MAX_LEN),
):
    abbr_max_len = resolve_abbr_max_len(abbr_max_len)
    logger.info(f"批量生成DDL请求 - 文件: {file.filename}, 数据库类型: {db_type}, 任务ID: {task_id}, 自定义提示词: {'是' if custom_prompt else '否'}, 行业背景: {'是' if normalize_industry_context(industry_context) else '否'}, 三层校验: {'启用' if enable_validation else '禁用'}, 并发线程数: {max_workers}, 温度: {temperature}, 分词模式: {cut_mode}, 缩写上限: {abbr_max_len}")
    
    if not task_id:
        return {"code": 1, "message": "缺少任务ID"}
    
    if task_id in batch_cache:
        logger.info(f"新上传文件，清除旧缓存 {task_id}，从头开始")
        del batch_cache[task_id]
    task_validation_flags[task_id] = enable_validation
    
    content = await file.read()
    
    asyncio.get_event_loop().run_in_executor(executor, process_batch_task, task_id, content, api_key, api_url, model, db_type, root_match_priority, custom_prompt, enable_validation, max_workers, temperature, cut_mode, abbr_max_len, industry_context)

    return {"code": 0, "message": "任务已启动", "data": {"task_id": task_id, "enable_validation": enable_validation, "max_workers": max_workers, "abbr_max_len": abbr_max_len}}


@app.post("/api/text-generate-ddl-task")
async def text_generate_ddl_task(req: TextGenerateTaskRequest):
    task_id = req.task_id or str(uuid.uuid4())
    logger.info(
        f"文本生成DDL任务启动 - 任务ID: {task_id}, 数据库类型: {req.db_type}, "
        f"词根模式: {req.root_match_priority}, 分词模式: {req.cut_mode}, 三层校验: {'启用' if req.enable_validation else '禁用'}"
    )
    task_validation_flags[task_id] = req.enable_validation
    payload = req.model_copy(update={"task_id": task_id})
    asyncio.get_event_loop().run_in_executor(executor, process_text_generate_task, payload)
    return {"code": 0, "message": "任务已启动", "data": {"task_id": task_id, "enable_validation": req.enable_validation, "max_workers": 1, "abbr_max_len": req.llm_config.abbr_max_len}}


@app.get("/api/batch-result/{task_id}")
async def get_batch_result(task_id: str):
    result = task_results.get(task_id)
    if result:
        return result
    return {"code": 1, "message": "任务处理中或不存在"}


@app.post("/api/batch-task/{task_id}/cancel")
async def cancel_batch_task(task_id: str):
    logger.info(f"终止任务请求: {task_id}")
    task_cancel_flags[task_id] = True
    return {"code": 0, "message": "任务已终止"}


@app.get("/api/db-connections")
async def get_db_connections_api():
    connections = load_db_connections(include_password=False)
    return {"code": 0, "data": connections}


@app.post("/api/db-connections")
async def create_db_connection_api(data: DbConnectionRequest):
    try:
        record = upsert_db_connection(data)
        return {"code": 0, "data": record, "message": "数据库连接保存成功"}
    except Exception as e:
        logger.error(f"保存数据库连接失败: {e}")
        return {"code": 1, "message": f"保存失败: {str(e)}"}


@app.put("/api/db-connections/{connection_id}")
async def update_db_connection_api(connection_id: str, data: DbConnectionRequest):
    try:
        record = upsert_db_connection(data, connection_id)
        return {"code": 0, "data": record, "message": "数据库连接更新成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新数据库连接失败: {e}")
        return {"code": 1, "message": f"更新失败: {str(e)}"}


@app.delete("/api/db-connections/{connection_id}")
async def delete_db_connection_api(connection_id: str):
    stored_connections = []
    if os.path.exists(DB_CONNECTIONS_FILE):
        with open(DB_CONNECTIONS_FILE, 'r', encoding='utf-8') as f:
            stored_connections = json.load(f)
    next_connections = [conn for conn in stored_connections if conn.get("id") != connection_id]
    if len(next_connections) == len(stored_connections):
        return {"code": 1, "message": "数据库连接不存在"}
    save_db_connections(next_connections)
    return {"code": 0, "message": "数据库连接删除成功"}


@app.post("/api/db-connections/{connection_id}/test")
async def test_db_connection_api(connection_id: str):
    connection = get_db_connection(connection_id, include_password=True)
    if not connection:
        return {"code": 1, "message": "数据库连接不存在"}
    try:
        result = test_native_connection(connection)
        return {"code": 0 if result.get("success") else 1, "message": result.get("message", "")}
    except Exception as e:
        logger.error(f"测试数据库连接失败: {e}")
        return {"code": 1, "message": f"测试连接失败: {str(e)}"}


@app.get("/api/ddl-history")
async def get_ddl_history_api(limit: int = 20):
    logger.info(f"获取建表历史记录，限制: {limit}")
    history = get_ddl_history(limit)
    return {"code": 0, "data": history}


@app.delete("/api/ddl-history/{record_id}")
async def delete_ddl_history_api(record_id: str):
    logger.info(f"删除建表历史记录: {record_id}")
    success = delete_ddl_history(record_id)
    if success:
        return {"code": 0, "message": "删除成功"}
    return {"code": 1, "message": "删除失败"}


@app.post("/api/ddl-history/{record_id}/execute")
async def execute_history_ddl_api(record_id: str, req: ExecuteDDLRequest):
    logger.info(f"执行历史建表SQL: record={record_id}, connection={req.connection_id}")
    history = get_ddl_history(100)
    record = next((item for item in history if item.get("id") == record_id), None)
    if not record:
        return {"code": 1, "message": "历史记录不存在"}

    connection = get_db_connection(req.connection_id, include_password=True)
    if not connection:
        return {"code": 1, "message": "数据库连接不存在"}

    if record.get("db_type") != connection.get("db_type"):
        return {"code": 1, "message": "SQL脚本数据库类型与连接类型不匹配"}

    try:
        result = execute_native_ddl(connection, record.get("ddl", ""))
        execution_info = {
            "status": "success" if result.get("success") else "failed",
            "time": datetime.now().isoformat(),
            "connection_name": connection.get("name", ""),
            "message": result.get("message", ""),
            "executed_count": result.get("executed_count", 0),
        }
        update_ddl_history_execution(record_id, execution_info)
        return {
            "code": 0 if result.get("success") else 1,
            "data": result,
            "message": result.get("message", ""),
        }
    except Exception as e:
        logger.error(f"执行历史建表SQL失败: {e}")
        update_ddl_history_execution(record_id, {
            "status": "failed",
            "time": datetime.now().isoformat(),
            "connection_name": connection.get("name", ""),
            "message": str(e),
            "executed_count": 0,
        })
        return {"code": 1, "message": f"执行失败: {str(e)}"}


@app.post("/api/save-new-roots")
async def save_new_roots(new_roots: List[WordRootItem]):
    logger.info(f"保存新词根请求，共 {len(new_roots)} 条")
    try:
        roots_data = [root.dict() for root in new_roots]
        with roots_lock:
            merge_and_save_roots(roots_data)
        return {"code": 0, "message": f"成功保存 {len(new_roots)} 条新词根"}
    except Exception as e:
        logger.error(f"保存新词根失败: {e}")
        return {"code": 1, "message": f"保存失败: {str(e)}"}


@app.get("/api/ddl-history/{record_id}/download")
async def download_history_ddl(record_id: str):
    logger.info(f"下载历史记录DDL: {record_id}")
    history = get_ddl_history(100)
    record = next((h for h in history if h.get('id') == record_id), None)
    
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    ddl_content = record.get('ddl', '')
    filename = f"ddl_{record_id}_{record['timestamp'][:10]}.sql"
    
    return StreamingResponse(
        iter([ddl_content]),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/progress/{task_id}")
async def get_progress(task_id: str):
    progress = progress_store.get(task_id)
    if progress:
        cached = batch_cache.get(task_id)
        current_ddl = ""
        new_roots = []
        if cached:
            results = cached.get('results', {})
            if results:
                current_ddl = "\n\n".join(results.values())
            new_roots = cached.get('new_roots', [])
        return {"code": 0, "data": {**progress, "current_ddl": current_ddl, "new_roots": new_roots}}
    return {"code": 1, "message": "任务不存在或已完成"}


@app.get("/api/batch-cache/{task_id}")
async def get_batch_cache(task_id: str):
    cached = batch_cache.get(task_id)
    if cached:
        return {
            "code": 0,
            "data": {
                "has_cache": True,
                "table_names": cached.get('table_names', []),
                "completed": cached.get('next_index', 0),
                "total": len(cached.get('table_names', [])),
                "errors": cached.get('errors', [])
            }
        }
    return {"code": 0, "data": {"has_cache": False}}


@app.delete("/api/batch-cache/{task_id}")
async def clear_batch_cache(task_id: str):
    if task_id in batch_cache:
        del batch_cache[task_id]
        return {"code": 0, "message": "缓存已清除"}
    return {"code": 1, "message": "缓存不存在"}


@app.get("/api/custom-prompt")
async def get_custom_prompt():
    logger.info("获取自定义提示词请求")
    if os.path.exists(CUSTOM_PROMPT_FILE):
        try:
            with open(CUSTOM_PROMPT_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
                if content.strip():
                    logger.info(f"成功读取自定义提示词，长度: {len(content)}")
                    return {"code": 0, "data": content}
        except Exception as e:
            logger.error(f"读取自定义提示词失败: {e}")
    logger.info("自定义提示词文件不存在或为空，尝试从内置默认文件初始化")
    builtin_content = seed_from_builtin(BUILTIN_PROMPT_FILE, CUSTOM_PROMPT_FILE, "text")
    if builtin_content:
        logger.info(f"从内置默认文件初始化成功，长度: {len(builtin_content)}")
        return {"code": 0, "data": builtin_content}
    logger.info("内置默认提示词文件不可用，返回空内容")
    return {"code": 0, "data": ""}


@app.post("/api/custom-prompt")
async def save_custom_prompt(data: dict = Body(...)):
    content = data.get('content', '')
    logger.info(f"保存自定义提示词请求，长度: {len(content)}")
    try:
        with open(CUSTOM_PROMPT_FILE, 'w', encoding='utf-8') as f:
            f.write(content)
        logger.info("自定义提示词保存成功")
        return {"code": 0, "message": "保存成功"}
    except Exception as e:
        logger.error(f"保存自定义提示词失败: {e}")
        return {"code": 1, "message": "保存失败"}


@app.post("/api/custom-prompt/reset")
async def reset_custom_prompt():
    logger.info("重置自定义提示词为内置默认")
    builtin_content = seed_from_builtin(BUILTIN_PROMPT_FILE, CUSTOM_PROMPT_FILE, "text")
    if builtin_content:
        return {"code": 0, "data": builtin_content, "message": "已重置为内置默认提示词"}
    return {"code": 1, "message": "内置默认提示词文件不可用"}


@app.get("/api/progress/sse/{task_id}")
async def progress_sse(task_id: str):
    async def event_stream():
        last_progress = None
        while True:
            progress = progress_store.get(task_id)
            if progress != last_progress:
                last_progress = progress
                yield f"data: {json.dumps(progress)}\n\n"
            await asyncio.sleep(0.5)
    
    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive"}
    )


if getattr(sys, 'frozen', False):
    exe_dir = os.path.dirname(sys.executable)
    frontend_candidates = [
        os.path.join(exe_dir, 'frontend'),
        os.path.join(exe_dir, 'frontend', 'dist'),
    ]
else:
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    frontend_candidates = [
        os.path.join(project_root, 'frontend', 'dist'),
        os.path.join(project_root, 'frontend'),
    ]

frontend_dir = next((path for path in frontend_candidates if os.path.exists(path)), None)
if frontend_dir:
    logger.info(f'前端文件已加载，目录: {frontend_dir}')
    app.mount('/', StaticFiles(directory=frontend_dir, html=True), name='frontend')
else:
    logger.warning('=' * 60)
    logger.warning('WARNING: 前端文件未找到！')
    for candidate in frontend_candidates:
        logger.warning(f'  已检查路径: {candidate}')
    logger.warning('  开发环境请先构建前端，打包环境请确保 frontend 或 frontend/dist 随程序分发')
    logger.warning('=' * 60)





