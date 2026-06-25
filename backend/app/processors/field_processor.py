import io
import json
import re
import hashlib
import logging
import requests
from typing import List, Dict, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import jieba
from app.field_type_resolver import get_excel_field_type, resolve_field_type
from app.name_normalizer import NameNormalizer, VALIDATION_ROOT_TOKENS
from app.root_policy import (
    DEFAULT_ABBR_MAX_LEN,
    build_root_sets,
    ensure_theme_prefix,
    filter_translation_by_mode,
    get_root_constraints,
    get_root_reuse_principle,
    get_theme_prefix_map,
    infer_theme_prefix,
    normalize_table_business_domain,
    normalize_priority,
    render_theme_prefix_guide,
    resolve_abbr_max_len,
    validate_identifier_mode,
)

logger = logging.getLogger(__name__)

TABLE_NAME_PROMPT = """
【任务】根据以下信息生成符合规范的英文表名

【中文表名】{chinese_table_name}

【开发规范】
{standards_content}

【数据库类型】{db_type}

【表名格式要求】{table_format}

【可用词根列表】
{available_roots}

{root_constraints}
{root_reuse_principle}

【输出格式】
请只输出英文表名，不要包含任何解释或额外内容。

【输出示例】
{table_name_example}
"""

THEME_PREFIX_PROMPT_BLOCK = """
銆愪富棰樺煙缂╁啓銆?{theme_prefix}

銆愪富棰樺煙缂╁啓瑙勫垯銆?
{theme_prefix_guide}
"""

MAX_FIELDS_PER_GROUP = 50
MAX_CONCURRENT_WORKERS = 5


def create_batches_by_ascii_order(roots, batch_size=50):
    """
    按中文ASCII排序创建批次
    中文按Unicode编码排序，相近语义的词根会相邻排列
    :param roots: 词根列表
    :param batch_size: 每批最大数量，默认50
    :return: 分批后的词根列表
    """
    # 中文ASCII排序（按Unicode编码）
    sorted_roots = sorted(roots)
    
    # 分批
    batches = []
    for i in range(0, len(sorted_roots), batch_size):
        batch = sorted_roots[i:i + batch_size]
        batches.append(batch)
    
    logger.info(f"【分批】词根按ASCII排序后分成 {len(batches)} 批")
    return batches

class FieldInfo:
    def __init__(self, table_name: str, chinese_name: str, suggested_type: str,
                 table_layer: str = "", field_index: int = 0):
        self.table_name = table_name
        self.table_layer = table_layer
        self.field_index = field_index
        self.chinese_name = chinese_name
        self.suggested_type = suggested_type
    
    def to_dict(self):
        return {
            'table_name': self.table_name,
            'table_layer': self.table_layer,
            'field_index': self.field_index,
            'chinese_name': self.chinese_name,
            'suggested_type': self.suggested_type
        }

class FieldProcessor:
    def __init__(self, api_key: str, api_url: str, model: str, temperature: float = 0.3, 
                 task_id: str = None, progress_callback=None, max_workers: int = 5,
                 custom_prompt: str = "", word_roots: list = None, root_match_priority: str = 'full',
                 cut_mode: str = "accurate", standards: str = "", abbr_max_len: int = DEFAULT_ABBR_MAX_LEN):
        self.api_key = api_key
        self.api_url = api_url
        self.model = model
        self.temperature = temperature
        self.field_mapping = {}
        self.field_tokenization = {}
        self.root_translations = {}
        self.semantic_root_map = {}
        self.fallback_root_cache = {}
        self.task_id = task_id
        self.progress_callback = progress_callback
        self.max_workers = max_workers
        self.custom_prompt = custom_prompt
        self.word_roots = word_roots or []
        self.root_match_priority = root_match_priority
        self.field_stats = None  # 保存字段统计信息
        self.cut_mode = cut_mode  # normalize and tokenize naming inputs
        self.standards = standards
        self.abbr_max_len = resolve_abbr_max_len(abbr_max_len)
        self.name_normalizer = NameNormalizer()
        for term in self.name_normalizer.iter_custom_terms():
            try:
                jieba.add_word(term)
            except Exception:
                pass
        self.full_root_set, self.abbr_root_set, self.abbr_to_full_roots = build_root_sets(self.word_roots)
        for token in VALIDATION_ROOT_TOKENS:
            self.full_root_set.add(token)
            if len(token) <= self.abbr_max_len:
                self.abbr_root_set.add(token)
        
        # 构建历史词根映射表（中文 -> 英文）
        self.existing_root_map = {}
        loaded_count = 0
        logger.info(f"【字段级处理】开始加载历史词根，共接收 {len(self.word_roots)} 条词根数据")
        
        for idx, root in enumerate(self.word_roots):
            chinese_name = root.get('chinese_name', '').strip()  # 去除前后空格
            full_root = root.get('full_root', '')
            abbr_root = root.get('abbr_root', '')
            
            if chinese_name:
                self.existing_root_map[chinese_name] = {
                    'full': full_root,
                    'abbr': abbr_root,
                    'type': root.get('recommended_type', 'VARCHAR(255)')
                }
                loaded_count += 1
                
                # 打印前10条词根用于调试
                if idx < 10:
                    logger.debug(f"【字段级处理】加载词根 {idx}: '{chinese_name}' -> full='{full_root}', abbr='{abbr_root}'")
        
        logger.info(f"【字段级处理】已加载 {loaded_count} 条历史词根到映射表")
        logger.info(f"【字段级处理】映射表中包含的词根示例（前10条）: {list(self.existing_root_map.keys())[:10]}")

    def _root_style_guide(self) -> str:
        return get_root_constraints(self.root_match_priority, self.abbr_max_len)

    def _choose_root_by_priority(self, root_info: dict) -> str:
        if normalize_priority(self.root_match_priority) == 'abbr':
            return root_info.get('abbr') or ''
        return root_info.get('full') or ''

    def _validate_identifier_by_priority(self, identifier: str, context: str) -> bool:
        identifier_to_check = self.name_normalizer.normalize_english_identifier(identifier)
        parts = [part for part in identifier_to_check.split("_") if part]
        if parts:
            strip_prefixes = set(get_theme_prefix_map(self.standards).values()) | {'dim', 'dwd', 'dws', 'ods', 'ads', 'input'}
            while len(parts) > 1 and parts[0] in strip_prefixes:
                parts = parts[1:]
            identifier_to_check = "_".join(parts) or identifier_to_check

        errors = validate_identifier_mode(
            identifier_to_check,
            self.root_match_priority,
            self.full_root_set,
            self.abbr_root_set,
            self.abbr_to_full_roots,
            self.abbr_max_len,
        )
        for error in errors:
            logger.warning(f"【字段级处理】{context}不符合词根优先级: {error}")
        return not errors
    def _filter_translation_by_priority(self, chinese_root: str, english_root: str) -> Optional[str]:
        normalized_root = self.name_normalizer.normalize_english_identifier(english_root)
        if not normalized_root or self._contains_chinese(normalized_root) or self._has_hash_root_token(normalized_root):
            logger.warning(f"Discard unsafe root translation: {chinese_root} -> {english_root}")
            return None
        is_valid, result = filter_translation_by_mode(
            chinese_root,
            normalized_root,
            self.root_match_priority,
            self.full_root_set,
            self.abbr_root_set,
            self.abbr_to_full_roots,
            self.abbr_max_len,
        )
        if not is_valid:
            logger.warning(f"??????????????????????????????: {result}")
            return None
        result = self.name_normalizer.normalize_english_identifier(result)
        if not result or self._has_hash_root_token(result) or self._contains_chinese(result):
            return None
        return result
    def _available_roots_for_priority(self) -> List[str]:
        if normalize_priority(self.root_match_priority) == 'abbr':
            return sorted(root for root in self.abbr_root_set if len(root) <= self.abbr_max_len)
        return sorted(self.full_root_set)

    def _mode_safe_table_fallback(self) -> str:
        return 'tbl' if normalize_priority(self.root_match_priority) == 'abbr' else 'table'

    def _extract_llm_content(self, result: dict) -> str:
        choices = result.get("choices") or []
        if not choices:
            return ""
        message = choices[0].get("message") or {}
        content = message.get("content")
        if isinstance(content, list):
            content = "\n".join(
                item.get("text", "") if isinstance(item, dict) else str(item)
                for item in content
            )
        if content and str(content).strip():
            return str(content).strip()
        fallback_text = str(choices[0].get("text") or "").strip()
        if fallback_text:
            return fallback_text
        reasoning = str(message.get("reasoning_content") or "").strip()
        if reasoning:
            logger.warning("LLM returned reasoning_content but no final content; treating response as empty")
        return ""

    def _short_hash_token(self, text: str, prefix: str = "x", max_len: int = None) -> str:
        max_len = self.abbr_max_len if max_len is None else max_len
        digest = hashlib.md5((text or "").encode("utf-8")).hexdigest()
        return f"{prefix}{digest}"[:max_len]

    def _abbr_from_english(self, english_root: str) -> str:
        normalized = self.name_normalizer.normalize_english_identifier(english_root)
        parts = [part for part in normalized.split("_") if part]
        if not parts:
            return ""
        if len(parts) == 1:
            part = parts[0]
            if len(part) <= self.abbr_max_len:
                return part
            consonants = "".join(ch for ch in part if ch not in "aeiou")
            return consonants[:self.abbr_max_len] or part[:self.abbr_max_len]
        return "".join(part[0] for part in parts if part)[:self.abbr_max_len]

    def _has_hash_root_token(self, identifier: str) -> bool:
        normalized = self.name_normalizer.normalize_english_identifier(identifier)
        return any(re.fullmatch(r"x[0-9a-f]{2,}", part) for part in normalized.split("_") if part)

    def _contains_chinese(self, value: str) -> bool:
        return any('\u4e00' <= ch <= '\u9fff' for ch in str(value or ""))

    def _is_safe_generated_identifier(self, identifier: str, context: str = "") -> bool:
        normalized = self.name_normalizer.normalize_english_identifier(identifier)
        if not normalized or self._contains_chinese(normalized) or self._has_hash_root_token(normalized):
            return False
        if not re.fullmatch(r"[a-z][a-z0-9_]*", normalized):
            return False
        return self._validate_identifier_by_priority(normalized, context or "generated identifier")

    def _is_safe_table_body(self, table_name: str, theme_prefix: str = "") -> bool:
        normalized = self.name_normalizer.normalize_english_identifier(table_name)
        if not normalized or self._contains_chinese(normalized) or self._has_hash_root_token(normalized):
            return False
        if not re.fullmatch(r"[a-z][a-z0-9_]*", normalized):
            return False

        parts = [part for part in normalized.split("_") if part]
        if not parts:
            return False

        invalid_singletons = {"tbl", "table"} | set(get_theme_prefix_map(self.standards).values())
        selected_theme = (theme_prefix or "").lower().strip("_")
        if selected_theme:
            invalid_singletons.add(selected_theme)

        if len(parts) == 1 and parts[0] in invalid_singletons:
            return False
        if selected_theme and parts[0] == selected_theme and len(parts) < 2:
            return False
        return True

    def _is_deepseek_native_api(self) -> bool:
        normalized = (self.api_url or '').strip().rstrip('/').lower()
        if normalized.endswith('/chat/completions'):
            normalized = normalized[:-len('/chat/completions')].rstrip('/')
        return normalized == 'https://api.deepseek.com'

    def _build_llm_url(self) -> str:
        api_url = self.api_url.rstrip('/')
        if api_url.endswith('/chat/completions'):
            return api_url
        return f"{api_url}/chat/completions"

    def _prepare_llm_payload(self, payload: dict) -> dict:
        request_payload = dict(payload)
        if self._is_deepseek_native_api():
            request_payload['thinking'] = {'type': 'disabled'}
        return request_payload

    def _post_llm(self, payload: dict, timeout: int):
        return requests.post(
            self._build_llm_url(),
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json"
            },
            json=self._prepare_llm_payload(payload),
            timeout=timeout
        )

    def _fallback_root_for_mode(self, chinese_root: str, preferred_english: str = "") -> str:
        cleaned_root = self.name_normalizer.clean_chinese_text(chinese_root) or str(chinese_root or "").strip()
        cache_key = (normalize_priority(self.root_match_priority), cleaned_root)
        if cache_key in self.fallback_root_cache:
            return self.fallback_root_cache[cache_key]

        normalized_preferred = self.name_normalizer.normalize_english_identifier(preferred_english)
        token = self._abbr_from_english(normalized_preferred) if normalize_priority(self.root_match_priority) == "abbr" else normalized_preferred
        token = self.name_normalizer.normalize_english_identifier(token)

        if token and self._is_safe_generated_identifier(token, f"fallback root {cleaned_root}"):
            self.fallback_root_cache[cache_key] = token
            return token

        self.fallback_root_cache[cache_key] = ""
        logger.warning(f"Field root '{cleaned_root}' has no safe translation; hash fallback is disabled")
        return ""

    def _ensure_identifier_by_priority(self, identifier: str, chinese_context: str) -> str:
        normalized = self.name_normalizer.normalize_english_identifier(identifier)
        if normalized and self._is_safe_generated_identifier(normalized, chinese_context):
            return normalized

        parts = [part for part in normalized.split("_") if part]
        safe_parts = [
            self._fallback_root_for_mode(f"{chinese_context}_{index}", part)
            for index, part in enumerate(parts)
        ]
        safe_identifier = "_".join(part for part in safe_parts if part)
        if safe_identifier and self._is_safe_generated_identifier(safe_identifier, f"{chinese_context} fallback"):
            return safe_identifier
        return ""

    def _strip_table_structure_suffixes(self, chinese_table_name: str) -> str:
        cleaned_name = self.name_normalizer.clean_chinese_text(chinese_table_name)
        if not cleaned_name:
            return ''

        suffixes = ('维度表', '明细表', '表')
        changed = True
        while changed:
            changed = False
            for suffix in suffixes:
                if cleaned_name.endswith(suffix) and len(cleaned_name) > len(suffix):
                    cleaned_name = cleaned_name[:-len(suffix)]
                    changed = True
                    break
        return cleaned_name

    def _match_table_name_from_root_library(self, chinese_table_name: str) -> List[str]:
        cleaned_name = self._strip_table_structure_suffixes(chinese_table_name)
        if not cleaned_name or not self.existing_root_map:
            return []

        candidates = []
        for chinese_root, root_info in self.existing_root_map.items():
            cleaned_root = self.name_normalizer.clean_chinese_text(chinese_root)
            selected_root = self._choose_root_by_priority(root_info)
            if cleaned_root and selected_root and not cleaned_root.isdigit():
                candidates.append((cleaned_root, selected_root))

        candidates.sort(key=lambda item: len(item[0]), reverse=True)
        roots = []
        index = 0
        while index < len(cleaned_name):
            match = next(
                (
                    (chinese_root, selected_root)
                    for chinese_root, selected_root in candidates
                    if cleaned_name.startswith(chinese_root, index)
                ),
                None,
            )
            if match:
                chinese_root, selected_root = match
                roots.append(selected_root)
                index += len(chinese_root)
                continue

            char = cleaned_name[index]
            if char.isascii() and char.isalnum():
                index += 1
                continue
            return []

        return roots

    def _resolve_root_translation(self, root: str, llm_translations: dict = None, allow_fallback: bool = False) -> Optional[str]:
        if not root:
            return None

        cleaned_root = self.name_normalizer.clean_chinese_text(root)
        if not cleaned_root:
            return None

        matched = self.match_existing_root(cleaned_root)
        if matched:
            return matched[0]

        canonical_root = self.name_normalizer.canonical_root_for_token(cleaned_root)
        if canonical_root:
            return canonical_root

        normalized_root = self.name_normalizer.normalize_english_identifier(cleaned_root)
        if normalized_root and re.fullmatch(r'[a-z][a-z0-9_]*', normalized_root):
            return normalized_root

        translation_sources = []
        if llm_translations:
            translation_sources.append(llm_translations)
        if self.root_translations:
            translation_sources.append(self.root_translations)

        for translations in translation_sources:
            if cleaned_root in translations:
                translated = self.name_normalizer.normalize_english_identifier(translations[cleaned_root])
                if self._validate_identifier_by_priority(translated, f"词根 {cleaned_root}"):
                    return translated
                return self._fallback_root_for_mode(cleaned_root, translated) if allow_fallback else None

        return self._fallback_root_for_mode(cleaned_root) if allow_fallback else None

    def _compose_field_name_from_roots(self, chinese_name: str) -> Optional[str]:
        roots = self.field_tokenization.get(chinese_name, [])
        if not roots:
            logger.warning(f"????????? '{chinese_name}' ??????????????")
            return None

        english_parts = []
        for root in roots:
            english_root = self._resolve_root_translation(root, allow_fallback=True)
            if not english_root:
                english_root = self._fallback_root_for_mode(root)
            english_parts.append(english_root)

        english_name = '_'.join(english_parts).lower()
        normalized_name = self.name_normalizer.normalize_english_identifier(english_name)
        if not normalized_name:
            return None
        return normalized_name
    
    def match_existing_root(self, chinese_name: str) -> Optional[tuple]:
        """
        Match a Chinese root against history or canonical aliases.
        :param chinese_name: Chinese root token
        :return: (english_root, recommended_type) or None
        """
        cleaned_name = self.name_normalizer.clean_chinese_text(chinese_name)
        if not cleaned_name:
            return None

        logger.debug(f"【字段级处理】尝试匹配 '{cleaned_name}'")

        if cleaned_name in self.existing_root_map:
            root_info = self.existing_root_map[cleaned_name]
            selected_root = self._choose_root_by_priority(root_info)
            if selected_root:
                logger.debug(f"【字段级处理】完全匹配成功: '{cleaned_name}' -> '{selected_root}'")
                return (selected_root, root_info.get('type', 'VARCHAR(255)'))
            fallback_source = root_info.get('full') or root_info.get('abbr') or ''
            fallback_root = self._fallback_root_for_mode(cleaned_name, fallback_source)
            logger.debug(f"【字段级处理】词根 '{cleaned_name}' 命中历史库但当前模式无可用词根，兜底为 '{fallback_root}'")
            return (fallback_root, root_info.get('type', 'VARCHAR(255)'))

        canonical = self.name_normalizer.canonical_root_for_token(cleaned_name)
        if canonical:
            return (canonical, 'VARCHAR(255)')

        logger.debug(f"【字段级处理】未命中历史库，映射示例: {list(self.existing_root_map.keys())[:20]}")
        return None
    def _try_split_and_match(self, chinese_name):
        """
        Attempt to split a Chinese name into multiple roots and match each part.
        """
        if not chinese_name:
            return None

        parts = self.name_normalizer.tokenize_chinese_name(chinese_name, jieba.lcut)
        if not parts:
            return None

        logger.info(f"【字段级处理】jieba 分词结果: '{chinese_name}' -> {parts}")

        english_parts = []
        for phrase in parts:
            translated = self._resolve_root_translation(phrase)
            if translated:
                english_parts.append(translated)
            else:
                logger.debug(f"【字段级处理】词组 '{phrase}' 未命中历史词根")
                return None

        if english_parts:
            logger.info(f"【字段级处理】拆分匹配成功: '{chinese_name}' -> {english_parts}")
            return english_parts

        logger.info(f"【字段级处理】部分词组未匹配到历史词根，交给 LLM 处理")
        return None
    def _update_progress(self, current, total, stage=None, matched_count=None, unmatched_count=None,
                         total_fields=None, field_progress=None):
        if self.progress_callback:
            # 如果没有传递新的统计信息，但有保存的统计信息，就使用保存的
            if matched_count is None and self.field_stats:
                matched_count = self.field_stats.get('matched_count')
            if unmatched_count is None and self.field_stats:
                unmatched_count = self.field_stats.get('unmatched_count')
            if total_fields is None and self.field_stats:
                total_fields = self.field_stats.get('total_fields')
            
            self.progress_callback(self.task_id, current, total, stage=stage, matched_count=matched_count, 
                                 unmatched_count=unmatched_count, total_fields=total_fields,
                                 field_progress=field_progress)
        
    def extract_fields_from_excel(self, file_content: bytes):
        logger.info("【字段级处理】解析批量建表Excel文件")
        self._update_progress(1, 8, "解析Excel文件...")
        
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        tables = {}
        subject_domain_values = {}
        all_fields = []
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = None
            
            for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_num == 1:
                    headers = [str(cell).strip() if cell else '' for cell in row]
                    continue
                
                if not row or all(not cell for cell in row):
                    continue
                
                row_data = {headers[i]: str(row[i]).strip() if row[i] else '' for i in range(len(headers))}
                
                table_name = row_data.get('表名', '')
                table_layer = row_data.get('表分层', '').lower()
                subject_domain = row_data.get('主题域', '').strip().lower()
                field_name = row_data.get('字段名', '')
                field_type = get_excel_field_type(row_data) or resolve_field_type(field_name)
                
                if not table_name or not field_name:
                    continue
                
                if table_name not in tables:
                    tables[table_name] = {
                        'layer': table_layer,
                        'fields': [],
                        'user_specified_subject_domain': ''
                    }
                    subject_domain_values[table_name] = set()
                
                if subject_domain:
                    subject_domain_values.setdefault(table_name, set()).add(subject_domain)
                
                field_index = len(tables[table_name]['fields']) + 1
                tables[table_name]['fields'].append({
                    'name': field_name,
                    'type': field_type,
                    'table_name': table_name,
                    'table_layer': table_layer,
                    'field_index': field_index
                })
                
                all_fields.append(FieldInfo(
                    table_name=table_name,
                    table_layer=table_layer,
                    field_index=field_index,
                    chinese_name=field_name,
                    suggested_type=field_type
                ))
        
        for table_name, domains in subject_domain_values.items():
            if len(domains) > 1:
                conflict_values = ", ".join(sorted(domains))
                raise ValueError(f"表[{table_name}]存在多个冲突的主题域值: {conflict_values}，请统一后重试")
            if len(domains) == 1:
                tables[table_name]['user_specified_subject_domain'] = next(iter(domains))
        
        logger.info(f"【字段级处理】Excel解析完成，共 {len(tables)} 张表，{len(all_fields)} 个字段")
        return tables, all_fields
    
    def group_fields_by_chinese(self, fields):
        logger.info("【字段级处理】按中文词义分组字段")
        self._update_progress(2, 8, "按中文词义分组...")
        
        groups = {}
        
        for field in fields:
            # 去除前后空格，确保分组不受空格影响
            key = field.chinese_name.strip()
            if key not in groups:
                groups[key] = []
            groups[key].append(field)
        
        logger.info(f"【字段级处理】字段分组完成，共 {len(groups)} 个不同的中文词义")
        return groups
    
    def split_into_batches(self, groups):
        logger.info("【字段级处理】将字段分组分批处理（含历史词根匹配）")
        logger.info(f"【字段级处理】历史词根库大小: {len(self.existing_root_map)} 条")
        logger.info(f"【字段级处理】待匹配的字段分组数量: {len(groups)} 个")
        logger.info(f"【字段级处理】待匹配的字段分组示例: {list(groups.keys())[:10]}")
        
        # 先进行历史词根匹配
        matched_fields = {}  # 已匹配的历史词根
        unmatched_groups = {}  # 需要提交给LLM的字段组
        
        for chinese_name, field_list in groups.items():
            # 调试：打印正在匹配的中文字段名
            logger.info(f"【字段级处理】正在匹配字段: '{chinese_name}'")
            
            matched_result = self.match_existing_root(chinese_name)
            if matched_result:
                english_name, field_type = matched_result
                matched_fields[chinese_name] = (english_name, field_type)
                logger.info(f"【字段级处理】✅ 历史词根匹配成功: '{chinese_name}' -> '{english_name}'")
            else:
                # 完全匹配失败，尝试拆分匹配
                split_parts = self._try_split_and_match(chinese_name)
                if split_parts:
                    # 拆分匹配成功，使用组合的英文名
                    english_name = '_'.join(split_parts)
                    matched_fields[chinese_name] = (english_name, field_list[0].suggested_type)
                    logger.info(f"【字段级处理】✅ 拆分匹配成功: '{chinese_name}' -> '{english_name}'")
                else:
                    unmatched_groups[chinese_name] = field_list
                    logger.info(f"【字段级处理】❌ 历史词根匹配失败: '{chinese_name}'")
        
        logger.info(f"【字段级处理】历史词根匹配完成，已匹配 {len(matched_fields)} 个，未匹配 {len(unmatched_groups)} 个")
        
        # 打印匹配成功的字段
        if matched_fields:
            matched_list = list(matched_fields.keys())[:10]
            logger.info(f"【字段级处理】已匹配成功的字段示例: {matched_list}")
        
        # 如果有未匹配的字段，列出前10个以便调试
        if unmatched_groups:
            unmatched_list = list(unmatched_groups.keys())[:10]
            logger.info(f"【字段级处理】未匹配的字段示例: {unmatched_list}")
        
        # 仅对未匹配的字段进行分批
        batches = []
        current_batch = []
        current_count = 0
        
        for chinese_name, field_list in unmatched_groups.items():
            field_count = len(field_list)
            
            if current_count + field_count > MAX_FIELDS_PER_GROUP and current_batch:
                batches.append(current_batch)
                current_batch = []
                current_count = 0
            
            current_batch.append((chinese_name, field_list))
            current_count += field_count
        
        if current_batch:
            batches.append(current_batch)
        
        logger.info(f"【字段级处理】未匹配字段分成 {len(batches)} 批处理，每批最多 {MAX_FIELDS_PER_GROUP} 个字段")
        
        # 计算统计信息
        stats = {
            'matched_count': len(matched_fields),
            'unmatched_count': len(unmatched_groups),
            'total_fields': len(groups)
        }
        
        # 保存统计信息，供后续进度更新使用
        self.field_stats = stats
        
        logger.info(f"【字段级处理】即将发送字段统计信息: matched={stats['matched_count']}, unmatched={stats['unmatched_count']}, total={stats['total_fields']}")
        
        # 更新进度时传递统计信息
        self._update_progress(3, 8, "jieba分词中...", matched_count=stats['matched_count'], 
                            unmatched_count=stats['unmatched_count'], total_fields=stats['total_fields'])
        
        logger.info(f"【字段级处理】字段统计信息已发送")
        
        return batches, matched_fields, stats
    
    def build_prompt_for_batch(self, batch):
        field_lines = []
        
        for chinese_name, field_list in batch:
            first_field = field_list[0]
            field_lines.append(f"{chinese_name}:{first_field.suggested_type}")
        
        fields_text = "\n".join(field_lines)
        
        # 根据用户设置的优先级选择提示词风格
        style_guide = self._root_style_guide()
        
        if self.custom_prompt:
            prompt = f"""请为以下中文业务词根生成英文翻译，严禁使用无意义的命名翻译
{self._render_custom_prompt_root_mode()}

【字段列表】
{fields_text}

【词根规则】
1. 字段名必须使用下划线分隔，全部小写
2. {style_guide}
3. 同一中文含义必须使用相同的英文词根
4. 如果有多个中文词组合，使用下划线连接各个词根

请按以下格式输出，每行一个：
中文名称:英文词根:推荐类型
示例：
订单:order:INT
客户:customer:VARCHAR(64)

不要输出任何解释，只输出字段映射。"""
        else:
            prompt = f"""请为以下中文业务字段生成英文词根：

【字段列表】
{fields_text}

【词根规则】
1. 字段名必须使用下划线分隔，全部小写
2. {style_guide}
3. 同一中文含义必须使用相同的英文词根
4. 如果有多个中文词组合，使用下划线连接各个词根

请按以下格式输出，每行一个：
中文名称:英文词根:推荐类型
示例：
订单:order:INT
客户:customer:VARCHAR(64)
创建时间:create_time:DATETIME
用户ID:user_id:VARCHAR(64)

不要输出任何解释，只输出字段映射。"""
        
        return prompt
    
    def call_llm_for_batch(self, batch):
        import requests
        
        prompt = self.build_prompt_for_batch(batch)
        
        data = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": "你是一位数据仓库专家，请根据中文业务字段生成符合规范的英文词根。"},
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 2048,
            "temperature": self.temperature
        }

        response = self._post_llm(data, timeout=300)
        
        if response.status_code != 200:
            logger.error(f"LLM调用失败: {response.status_code}")
            return {}
        
        response.encoding = 'utf-8'
        result = response.json()
        content = self._extract_llm_content(result)
        
        return self.parse_llm_response(content)
    
    def parse_llm_response(self, content):
        mapping = {}
        
        for line in content.split('\n'):
            line = line.strip()
            if not line or line.startswith('--') or line.startswith('#'):
                continue
            
            parts = line.split(':')
            if len(parts) >= 2:
                chinese_name = parts[0].strip()
                english_name = parts[1].strip() if len(parts) > 1 else ''
                field_type = parts[2].strip() if len(parts) > 2 else 'VARCHAR(255)'
                
                if chinese_name and english_name:
                    mapping[chinese_name] = (english_name, field_type)
                    logger.debug(f"解析映射: {chinese_name} -> {english_name}")
        
        logger.info(f"从LLM响应中解析到 {len(mapping)} 个字段映射")
        return mapping
    
    def generate_table_name(self, chinese_table_name, db_type, standards_content='', layer=''):
        """
        调用LLM生成英文表名
        :param chinese_table_name: 中文表名
        :param db_type: 数据库类型
        :param standards_content: 开发规范内容
        :param layer: 表分层（如 dim, dwd, dws）
        :return: 英文表名
        """
        logger.info(f"【字段级处理】调用LLM生成表名: '{chinese_table_name}', layer: '{layer}'")
        
        theme_prefix = infer_theme_prefix(chinese_table_name, layer, standards_content=standards_content)
        example_domain = theme_prefix or "biz"

        # 根据数据库类型设置输出示例
        table_name_example = {
            'mysql': f'dwd_{example_domain}_order',
            'postgresql': f'dwd.{example_domain}_order',
            'oracle': f'DWD.{example_domain.upper()}_ORDER'
        }.get(db_type.lower(), f'dwd_{example_domain}_order')
        
        # 获取数据库表名格式要求
        table_format = {
            'mysql': f'使用下划线前缀表示分层与业务域（如 dwd_{example_domain}_order）',
            'postgresql': f'使用 schema.table 格式，表名包含业务域前缀（如 dwd.{example_domain}_order）',
            'oracle': f'使用大写 SCHEMA.TABLE 格式，表名包含业务域前缀（如 DWD.{example_domain.upper()}_ORDER）'
        }.get(db_type.lower(), '使用下划线前缀表示分层与业务域')
        
        # 构建提示词
        prompt = TABLE_NAME_PROMPT.format(
            chinese_table_name=chinese_table_name,
            standards_content=standards_content,
            db_type=db_type,
            table_format=table_format,
            table_name_example=table_name_example,
            available_roots=", ".join(self._available_roots_for_priority()) or "无",
            root_constraints=self._root_style_guide(),
            root_reuse_principle=get_root_reuse_principle(self.root_match_priority, self.abbr_max_len)
        )
        prompt += THEME_PREFIX_PROMPT_BLOCK.format(
            theme_prefix=theme_prefix,
            theme_prefix_guide=render_theme_prefix_guide(standards_content)
        )
        
        # 调用LLM
        data = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": "你是一位数据仓库专家，请根据中文表名生成符合规范的英文表名。"},
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 64,
            "temperature": 0.1
        }

        try:
            response = self._post_llm(data, timeout=60)
            
            if response.status_code == 200:
                response.encoding = 'utf-8'
                result = response.json()
                table_name = self._extract_llm_content(result)
                if not table_name:
                    logger.warning(f"【字段级处理】表名生成响应为空，改用动态词根fallback: '{chinese_table_name}'")
                    return self.translate_table_name(chinese_table_name, layer)
                logger.info(f"【字段级处理】表名生成成功: '{chinese_table_name}' -> '{table_name}'")
                
                # 处理LLM返回的表名，提取纯表名部分
                # 情况1: 表名可能已经是 schema.table 格式（如 dim.product_info）
                if '.' in table_name:
                    parts = table_name.split('.')
                    # 如果最后一部分不是空的，使用最后一部分作为表名
                    if parts[-1]:
                        table_name = parts[-1]
                        logger.info(f"【字段级处理】从schema.table格式中提取表名: '{table_name}'")
                
                # 情况2: 检查是否包含layer作为独立部分（下划线分隔）
                if layer:
                    layer_prefixes = ['dim', 'dwd', 'dws', 'ods', 'ads', 'input']
                    layer_lower = layer.lower()
                    if layer_lower in layer_prefixes:
                        parts = table_name.split('_')
                        if layer_lower in parts:
                            new_parts = [p for p in parts if p != layer_lower]
                            if new_parts:
                                table_name = '_'.join(new_parts)
                                logger.info(f"【字段级处理】移除重复的layer '{layer_lower}'，表名变为: '{table_name}'")

                table_name = table_name.strip().strip('`"').lower()
                table_name = normalize_table_business_domain(
                    table_name,
                    theme_prefix,
                    standards_content,
                    layer
                )
                if self._is_safe_table_body(table_name, theme_prefix):
                    return table_name
                logger.warning(f"【字段级处理】LLM表名不符合表名规则，改用动态词根fallback: {table_name}")
                return self.translate_table_name(chinese_table_name, layer)
            else:
                logger.error(f"【字段级处理】表名生成失败: HTTP {response.status_code}")
                return self.translate_table_name(chinese_table_name, layer)
        except Exception as e:
            logger.error(f"【字段级处理】表名生成异常: {str(e)}")
            return self.translate_table_name(chinese_table_name, layer)
    
    def process_batches_parallel(self, batches):
        logger.info("【字段级处理】并行调用LLM生成字段映射")
        self._update_progress(5, 8, "批量生成字段英文名...")
        
        all_mappings = {}
        completed_batches = 0
        total_batches = len(batches)
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {}
            
            for batch_idx, batch in enumerate(batches):
                future = executor.submit(self.call_llm_for_batch, batch)
                futures[future] = batch_idx
            
            for future in as_completed(futures):
                batch_idx = futures[future]
                completed_batches += 1
                try:
                    mapping = future.result()
                    all_mappings.update(mapping)
                    logger.info(f"【字段级处理】批次 {batch_idx + 1}/{total_batches} 处理完成")
                    self._update_progress(5, 8, f"生成字段英文名 [{completed_batches}/{total_batches}]")
                except Exception as e:
                    logger.error(f"【字段级处理】批次 {batch_idx + 1} 处理失败: {e}")
        
        self.field_mapping = all_mappings
        logger.info(f"【字段级处理】所有批次处理完成，共 {len(all_mappings)} 个字段映射")
        return all_mappings
    
    def tokenize_all_fields_root_level(self, groups):
        """
        对所有字段进行 jieba 分词，收集所有词根并去重
        :param groups: 按中文词义分组的字段字典 {chinese_name: [FieldInfo, ...]}
        :return: (field_tokenization, unique_roots)
            field_tokenization: {chinese_name: [root1, root2, ...]}  每个字段的分词结果
            unique_roots: [root1, root2, ...]  所有去重后的词根列表
        """
        logger.info(f"【字段级处理】开始对所有字段进行 jieba 词根级分词，模式: {self.cut_mode}")
        
        # 更新进度：开始jieba分词
        self._update_progress(3, 8, "jieba分词中...")
        
        field_tokenization = {}  # {chinese_name: [roots]}
        all_roots = []  # 收集所有词根（不去重）
        total_fields = len(groups)
        processed_fields = 0
        
        for chinese_name in groups.keys():
            cleaned_name = self.name_normalizer.clean_chinese_text(chinese_name)
            # 根据分词模式选择不同的分词方法
            if not cleaned_name:
                roots = []
            elif self.cut_mode == "full":
                # 全模式：把句子中所有可以成词的词语都扫描出来
                roots = jieba.lcut(cleaned_name, cut_all=True)
            elif self.cut_mode == "search":
                # 搜索引擎模式：先精确分词，然后对长词再次切分
                roots = jieba.lcut_for_search(cleaned_name)
            else:
                # 精准模式（默认）：试图将句子最精确地切开
                roots = jieba.lcut(cleaned_name, cut_all=False)
            
            # 过滤空字符串
            roots = [r.strip() for r in roots if r.strip()]
            
            field_tokenization[chinese_name] = roots
            all_roots.extend(roots)
            processed_fields += 1
            
            # 每处理100个字段更新一次进度
            if processed_fields % 100 == 0:
                self._update_progress(3, 8, f"jieba分词 [{processed_fields}/{total_fields}]")
            
            logger.debug(f"【字段级处理】分词: '{chinese_name}' -> '{cleaned_name}' -> {roots}")
        
        # 去重
        unique_roots = list(dict.fromkeys(all_roots))
        
        # 更新进度：jieba分词完成
        self._update_progress(
            3,
            8,
            f"分词完成：原始词根{len(all_roots)}个，去重后{len(unique_roots)}个",
            field_progress={
                "phase": "tokenized",
                "phase_label": "jieba分词完成",
                "raw_root_count": len(all_roots),
                "unique_root_count": len(unique_roots),
                "completed_items": len(unique_roots),
                "total_items": len(unique_roots),
                "thread_count": 0,
                "batch_count": 0,
                "completed_batches": 0
            }
        )
        
        logger.info(f"【字段级处理】词根分词完成，共 {len(field_tokenization)} 个字段，{len(unique_roots)} 个唯一词根")
        logger.debug(f"【字段级处理】唯一词根列表: {unique_roots[:20]}...")
        
        return field_tokenization, unique_roots
    
    def match_roots_against_history(self, unique_roots):
        """
        对所有词根进行历史词根匹配
        :param unique_roots: 唯一词根列表
        :return: (matched_roots, unmatched_roots)
            matched_roots: {chinese_root: english_root}
            unmatched_roots: [chinese_root1, chinese_root2, ...]
        """
        logger.info(f"【字段级处理】开始匹配历史词根，共 {len(unique_roots)} 个词根")
        
        # 更新进度：开始历史词根匹配
        self._update_progress(4, 8, "历史词根匹配中...")
        
        matched_roots = {}
        unmatched_roots = []
        total_roots = len(unique_roots)
        processed_roots = 0
        
        for root in unique_roots:
            matched_result = self.match_existing_root(root)
            if matched_result:
                english_root, _ = matched_result
                matched_roots[root] = english_root
                logger.debug(f"【字段级处理】词根匹配成功: '{root}' -> '{english_root}'")
            else:
                unmatched_roots.append(root)
                logger.debug(f"【字段级处理】词根未匹配: '{root}'")
            
            processed_roots += 1
            # 每处理50个词根更新一次进度
            if processed_roots % 50 == 0:
                self._update_progress(4, 8, f"历史词根匹配 [{processed_roots}/{total_roots}]")
        
        # 更新进度：匹配完成
        self._update_progress(
            4,
            8,
            f"历史词根匹配完成：匹配{len(matched_roots)}个，LLM需生成{len(unmatched_roots)}个，去重词根{len(unique_roots)}个",
            matched_count=len(matched_roots),
            unmatched_count=len(unmatched_roots),
            total_fields=len(unique_roots),
            field_progress={
                "phase": "history_matched",
                "phase_label": "历史词根匹配完成",
                "matched_count": len(matched_roots),
                "unmatched_count": len(unmatched_roots),
                "unique_root_count": len(unique_roots),
                "completed_items": len(unique_roots),
                "total_items": len(unique_roots),
                "thread_count": 0,
                "batch_count": 0,
                "completed_batches": 0
            }
        )
        
        logger.info(f"【字段级处理】历史词根匹配完成，已匹配 {len(matched_roots)} 个，未匹配 {len(unmatched_roots)} 个")
        
        return matched_roots, unmatched_roots

    def normalize_unmatched_roots_semantically(self, unmatched_roots):
        """
        Ask LLM to normalize synonymous unmatched roots before batch translation.
        :param unmatched_roots: unmatched Chinese root list
        :return: {chinese_root: english_root}
        """
        if not unmatched_roots:
            return {}

        roots = [root for root in unmatched_roots if root]
        if not roots:
            return {}
        if not self.api_url or not self.model:
            logger.info("【字段级处理】未配置LLM接口，跳过词根语义归一")
            return {}

        history_rows = []
        for cn_root, root_info in self.existing_root_map.items():
            selected_root = self._choose_root_by_priority(root_info)
            if selected_root:
                history_rows.append(f"{cn_root}:{selected_root}")

        roots_text = "\n".join(roots)
        history_text = "\n".join(history_rows) if history_rows else "无"
        style_guide = self._root_style_guide()

        prompt = f"""请对以下中文业务词根做语义归一和英文翻译。

【待归一词根列表】
{roots_text}

【历史词根映射（优先沿用）】
{history_text}

【当前词根模式约束】
{style_guide}

【要求】
1. 判断待归一词根中是否存在同义词、近义词或同一业务概念的不同说法。
2. 同义/近义词必须输出同一个英文词根，例如“预计”和“预估”如果语义相同，应使用同一个英文词根。
3. 如果历史词根映射中已有相同语义，请优先沿用历史英文词根。
4. 必须为每个待归一词根输出一行。

【输出格式】
中文词根:英文词根

不要输出任何解释，只输出词根映射。"""

        data = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": "你是一位数据仓库词根语义归一专家，请识别同义/近义业务词根并统一英文命名。"},
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 4096,
            "temperature": 0.1
        }

        try:
            response = self._post_llm(data, timeout=300)

            if response.status_code != 200:
                logger.error(f"词根语义归一LLM调用失败: {response.status_code}")
                return {}

            response.encoding = 'utf-8'
            result = response.json()
            content = self._extract_llm_content(result)

            normalized = {}
            expected_roots = set(roots)
            for line in content.split('\n'):
                line = line.strip()
                if not line or line.startswith('--') or line.startswith('#') or ':' not in line:
                    continue

                chinese_root, english_root = line.split(':', 1)
                chinese_root = chinese_root.strip()
                english_root = english_root.strip().rstrip('_')
                if chinese_root not in expected_roots or not english_root:
                    continue

                filtered_root = self._filter_translation_by_priority(chinese_root, english_root)
                if filtered_root:
                    normalized[chinese_root] = filtered_root

            logger.info(f"【字段级处理】词根语义归一完成，获得 {len(normalized)} 个统一词根")
            return normalized
        except Exception as e:
            logger.error(f"【字段级处理】词根语义归一异常: {str(e)}")
            return {}
    
    def translate_unmatched_roots_via_llm(self, unmatched_roots, semantic_root_map=None):
        """
        将未匹配的词根交给 LLM 翻译（多线程并行）
        :param unmatched_roots: 未匹配的词根列表
        :param semantic_root_map: 全量语义归一得到的词根映射
        :return: {chinese_root: english_root} 新词根翻译结果
        """
        if not unmatched_roots:
            return {}
        semantic_root_map = semantic_root_map or {}
        self.semantic_root_map = dict(semantic_root_map)
        
        logger.info(f"【字段级处理】开始 LLM 翻译 {len(unmatched_roots)} 个未匹配词根")

        all_translations = {}
        remaining_roots = []
        for root in unmatched_roots:
            if root in semantic_root_map:
                filtered_root = self._filter_translation_by_priority(root, semantic_root_map[root])
                if filtered_root:
                    all_translations[root] = filtered_root
                    continue
            remaining_roots.append(root)

        if not remaining_roots:
            logger.info(f"【字段级处理】全部未匹配词根已由语义归一覆盖，获得 {len(all_translations)} 个词根翻译")
            return all_translations
        
        # 按中文ASCII排序分批
        batches = create_batches_by_ascii_order(remaining_roots, MAX_FIELDS_PER_GROUP)
        
        logger.info(f"【字段级处理】词根分成 {len(batches)} 批，每批最多 {MAX_FIELDS_PER_GROUP} 个")
        
        # 多线程并行调用 LLM
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(self._translate_roots_batch, batch): batch for batch in batches}
            
            for future in as_completed(futures):
                try:
                    batch_translations = future.result()
                    all_translations.update(batch_translations)
                except Exception as e:
                    logger.error(f"【字段级处理】词根批次翻译失败: {e}")
        
        logger.info(f"【字段级处理】LLM翻译完成，获得 {len(all_translations)} 个新词根翻译")
        return all_translations
    
    def _translate_roots_batch(self, roots_batch):
        """
        翻译一批词根（供多线程调用）
        :param roots_batch: 词根列表
        :return: {chinese_root: english_root}
        """
        if not roots_batch:
            return {}
        
        roots_text = "\n".join(roots_batch)
        semantic_rows = [
            f"{cn_root}:{en_root}"
            for cn_root, en_root in self.semantic_root_map.items()
            if cn_root and en_root
        ]
        semantic_text = "\n".join(semantic_rows) if semantic_rows else "无"
        
        style_guide = self._root_style_guide()
        
        if self.custom_prompt:
            prompt = f"""请为以下中文业务词根生成英文翻译，严禁使用无意义的命名翻译
{self._render_custom_prompt_root_mode()}

【待翻译词根列表】
{roots_text}

【词根规则】
1. 词根必须全部小写
2. {style_guide}
3. 每个词根单独一行
4. 同义/近义中文词根必须翻译为同一个英文词根

【已确定同义词归一表】
{semantic_text}

如果待翻译词根与归一表中的词根同义或近义，必须使用归一表中的相同英文词根，不得改写。

请按以下格式输出，每行一个：
中文词根:英文词根
示例：
订单:order
客户:customer

不要输出任何解释，只输出词根映射。"""
        else:
            prompt = f"""请为以下中文业务词根生成英文翻译，严禁使用无意义的命名翻译：

【待翻译词根列表】
{roots_text}

【词根规则】
1. 词根必须全部小写
2. {style_guide}
3. 每个词根单独一行
4. 同义/近义中文词根必须翻译为同一个英文词根

【已确定同义词归一表】
{semantic_text}

如果待翻译词根与归一表中的词根同义或近义，必须使用归一表中的相同英文词根，不得改写。

请按以下格式输出，每行一个：
中文词根:英文词根
示例：
订单:order
客户:customer
创建时间:create_time
用户ID:user_id

不要输出任何解释，只输出词根映射。"""
        
        try:
            # 构建system prompt：开发规范 + 角色定义
            if self.standards:
                system_content = f"""你是一位数据仓库专家，请根据中文业务词根生成符合规范的英文翻译。

【开发规范】
{self.standards}

请严格按照开发规范和已有词根进行翻译，并注意同一汉语意思要翻译为同一词根

【重要约束】{style_guide}
"""
            else:
                system_content = f"""你是一位数据仓库专家，请根据中文业务词根生成符合规范的英文翻译。

【重要约束】{style_guide}
"""
            
            data = {
                "model": self.model,
                "messages": [
                    {"role": "system", "content": system_content},
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 2048,
                "temperature": self.temperature
            }

            response = self._post_llm(data, timeout=300)
            
            if response.status_code != 200:
                logger.error(f"LLM调用失败: {response.status_code}")
                return {}
            
            response.encoding = 'utf-8'
            result = response.json()
            content = self._extract_llm_content(result)
            if not content:
                logger.warning("【字段级处理】LLM词根翻译响应为空，跳过本批次并等待上层兜底")
                return {}
            
            # 解析 LLM 响应
            batch_translations = {}
            for line in content.split('\n'):
                line = line.strip()
                if not line or line.startswith('--') or line.startswith('#'):
                    continue
                
                parts = line.split(':', 1)
                if len(parts) >= 2:
                    chinese_root = parts[0].strip()
                    english_root = parts[1].strip()
                    
                    if chinese_root and english_root:
                        # 清理词根：移除首尾空格和末尾下划线
                        english_root_cleaned = english_root.strip().rstrip('_')
                        filtered_root = self._filter_translation_by_priority(chinese_root, english_root_cleaned)
                        if filtered_root:
                            batch_translations[chinese_root] = filtered_root
                            logger.debug(f"【字段级处理】LLM翻译: '{chinese_root}' -> '{filtered_root}'")
            
            logger.info(f"【字段级处理】批次翻译完成，获得 {len(batch_translations)} 个词根翻译")
            return batch_translations
            
        except Exception as e:
            logger.error(f"【字段级处理】批次翻译异常: {str(e)}")
            return {}
    
    def process_layer1_fields(self, layer1_fields, existing_root_map):
        """
        Handle fully matched fields.
        """
        results = {}
        for chinese_name, roots in layer1_fields.items():
            english_parts = []
            valid_parts = True
            for root in roots:
                selected_root = self._resolve_root_translation(root)
                if selected_root:
                    english_parts.append(selected_root)
                else:
                    logger.warning(f"【字段级处理】词根 '{root}' 无法映射为当前模式下的英文根词")
                    valid_parts = False
                    break
            if english_parts and valid_parts:
                english_name = '_'.join(english_parts)
                if self._validate_identifier_by_priority(english_name, f"Layer1字段 '{chinese_name}'"):
                    results[chinese_name] = english_name
                    logger.debug(f"Layer1匹配: {chinese_name} -> {results[chinese_name]}")
        return results
    def process_layer2_fields(self, layer2_fields, existing_root_map, llm_translations):
        """
        Handle partially matched fields.
        """
        results = {}
        for chinese_name, roots in layer2_fields.items():
            english_parts = []
            valid_parts = True
            for root in roots:
                selected_root = self._resolve_root_translation(root, llm_translations)
                if selected_root:
                    english_parts.append(selected_root)
                else:
                    logger.warning(f"【字段级处理】词根 '{root}' 没有符合当前模式的翻译结果")
                    valid_parts = False
                    break
            if english_parts and valid_parts:
                english_name = '_'.join(english_parts)
                if self._validate_identifier_by_priority(english_name, f"Layer2字段 '{chinese_name}'"):
                    results[chinese_name] = english_name
                    logger.debug(f"Layer2匹配: {chinese_name} -> {results[chinese_name]}")
        return results
    def process_layer3_fields(self, layer3_fields, llm_designs):
        """
        Handle fully unmatched fields after LLM design.
        """
        results = {}
        new_roots = {}
        for chinese_name, roots in layer3_fields.items():
            if chinese_name in llm_designs:
                design = llm_designs[chinese_name]
                for cn_root, en_root in design['roots'].items():
                    filtered_root = self._filter_translation_by_priority(cn_root, en_root)
                    if filtered_root and cn_root not in new_roots:
                        new_roots[cn_root] = filtered_root
                english_parts = []
                valid_parts = True
                for root in roots:
                    translated = new_roots.get(root) or self._resolve_root_translation(root, new_roots)
                    if not translated:
                        logger.warning(f"Layer3字段 '{chinese_name}' 的词根 '{root}' 没有符合当前模式的翻译")
                        valid_parts = False
                        break
                    english_parts.append(translated)
                if valid_parts and english_parts:
                    field_name = '_'.join(english_parts)
                    if self._validate_identifier_by_priority(field_name, f"Layer3字段 '{chinese_name}'"):
                        results[chinese_name] = field_name
                        logger.debug(f"Layer3匹配: {chinese_name} -> {field_name}")
        return results, new_roots
    def parse_layer3_output(self, output_lines):
        """
        解析Layer 3的LLM输出
        """
        results = {}
        for line in output_lines:
            line = line.strip()
            if not line or ':' not in line:
                continue
            try:
                parts = line.split('|')
                field_part = parts[0]
                root_part = parts[1] if len(parts) > 1 else ""
                chinese_field, english_field = field_part.split(':', 1)
                roots = {}
                if root_part:
                    for root_pair in root_part.split(','):
                        if ':' in root_pair:
                            cn_root, en_root = root_pair.split(':', 1)
                            roots[cn_root.strip()] = en_root.strip()
                results[chinese_field.strip()] = {
                    "field_name": english_field.strip(),
                    "roots": roots
                }
            except Exception as e:
                logger.warning(f"解析Layer3输出失败: {line}, 错误: {e}")
        return results
    
    def build_layer3_prompt(self, unmatched_fields):
        """构建Layer 3的LLM提示词"""
        fields_text = "\n".join([
            f"{i+1}. \"{field['chinese_name']}\" - 分词: {field['roots']}"
            for i, field in enumerate(unmatched_fields)
        ])
        
        style_guide = self._root_style_guide()
        
        return f"""请为以下中文字段的每个分词词根生成英文词根翻译。字段名将由系统按分词顺序拼接，请不要自行改变分词顺序：

{fields_text}

【开发规范】
{self.standards}

【设计要求】
1. 只翻译给出的中文词根，不要创造整字段名绕过分词结果
2. 词根必须全部小写，系统会使用下划线拼接
3. {style_guide}

【基础词根速查】
- "是否"类 → "is"
- "状态"类 → "status"
- "类型"类 → "type"
- "名称"类 → "name"
- "数量"类 → "num"
- "金额"类 → "amt"
- "日期"类 → "date"
- "时间"类 → "time"
- "ID"类 → "id"

【输出格式】
中文字段名:英文字段名|中文词根1:英文词根1,中文词根2:英文词根2

【输出示例】
特殊标识:special_id|特殊:special,标识:id"""
    
    def _design_fields_batch(self, fields_batch):
        """设计一批字段（Layer 3专用）"""
        if not fields_batch:
            return {}
        
        prompt = self.build_layer3_prompt(fields_batch)
        
        try:
            data = {
                "model": self.model,
                "messages": [
                    {"role": "system", "content": "你是一位数据仓库专家。"},
                    {"role": "user", "content": prompt}
                ],
                "max_tokens": 4096,
                "temperature": self.temperature
            }

            response = self._post_llm(data, timeout=300)
            
            if response.status_code != 200:
                logger.error(f"Layer3 LLM调用失败: {response.status_code}")
                return {}
            
            result = response.json()
            content = self._extract_llm_content(result)
            
            return self.parse_layer3_output(content.split('\n'))
            
        except Exception as e:
            logger.error(f"【字段级处理】Layer3字段设计异常: {str(e)}")
            return {}
    
    def reassemble_fields_from_roots(self, field_tokenization, root_translation):
        """
        使用翻译好的词根重新拼接成完整的英文字段名
        :param field_tokenization: {chinese_name: [roots]}  每个字段的分词结果
        :param root_translation: {chinese_root: english_root}  词根翻译映射
        :return: {chinese_name: (english_name, field_type)}
        """
        logger.info("【字段级处理】开始将翻译好的词根拼接回完整字段名")
        
        field_mapping = {}
        
        for chinese_name, roots in field_tokenization.items():
            english_parts = []
            valid_parts = True
            for root in roots:
                if root in root_translation:
                    english_parts.append(root_translation[root])
                else:
                    logger.warning(f"【字段级处理】词根 '{root}' 没有符合当前模式的翻译结果，跳过字段 '{chinese_name}'")
                    valid_parts = False
                    break

            if not valid_parts:
                continue
            
            english_name = '_'.join(english_parts)
            if self._validate_identifier_by_priority(english_name, f"字段拼接 '{chinese_name}'"):
                field_mapping[chinese_name] = english_name
                logger.debug(f"【字段级处理】字段拼接: '{chinese_name}' -> '{english_name}'")
        
        logger.info(f"【字段级处理】字段拼接完成，共 {len(field_mapping)} 个字段")
        return field_mapping
    
    def process_fields_root_level(self, groups):
        """
        Hybrid root-level field processing pipeline.
        1. jieba tokenization
        2. history root matching
        3. global unmatched-root translation
        4. field reassembly from the same root translation map
        """
        logger.info("【字段级处理】开始混合模式词根级处理流程")

        field_tokenization, unique_roots = self.tokenize_all_fields_root_level(groups)
        self.field_tokenization = field_tokenization

        matched_roots, unmatched_roots = self.match_roots_against_history(unique_roots)
        logger.info(f"【字段级处理】词根匹配完成: 匹配 {len(matched_roots)} 个, 未匹配 {len(unmatched_roots)} 个")

        layer1_fields = {}
        layer2_fields = {}
        layer3_fields = {}
        direct_overrides = {}

        for chinese_name, roots in field_tokenization.items():
            override = self.name_normalizer.get_field_override(chinese_name)
            if override:
                direct_overrides[chinese_name] = (override, groups[chinese_name][0].suggested_type if groups.get(chinese_name) else 'VARCHAR(255)')
                continue

            matched_in_field = [r for r in roots if r in matched_roots]
            if len(matched_in_field) == len(roots):
                layer1_fields[chinese_name] = roots
            elif len(matched_in_field) > 0:
                layer2_fields[chinese_name] = roots
            else:
                layer3_fields[chinese_name] = roots

        logger.info(
            f"【字段级处理】字段分层完成: Layer1={len(layer1_fields)}, Layer2={len(layer2_fields)}, Layer3={len(layer3_fields)}"
        )

        planned_batches = create_batches_by_ascii_order(unmatched_roots, MAX_FIELDS_PER_GROUP) if unmatched_roots else []
        planned_batch_count = len(planned_batches)
        planned_thread_count = min(self.max_workers, planned_batch_count) if planned_batch_count else 0

        self._update_progress(
            5,
            8,
            f"生成字段名准备: 去重词根{len(unique_roots)}个, 字段{len(field_tokenization)}个, Layer1={len(layer1_fields)}, Layer2={len(layer2_fields)}, Layer3={len(layer3_fields)}, 预计{planned_batch_count}批, 线程{planned_thread_count}个",
            matched_count=len(matched_roots),
            unmatched_count=len(unmatched_roots),
            total_fields=len(unique_roots),
            field_progress={
                "phase": "field_generation_prepared",
                "phase_label": "字段分层完成，准备生成字段名",
                "unique_root_count": len(unique_roots),
                "target_item_label": "字段",
                "total_items": len(field_tokenization),
                "completed_items": 0,
                "batch_count": planned_batch_count,
                "completed_batches": 0,
                "thread_count": planned_thread_count,
                "layer1_count": len(layer1_fields),
                "layer2_count": len(layer2_fields),
                "layer3_count": len(layer3_fields)
            }
        )

        self._update_progress(
            5,
            8,
            f"Layer2词根生成中: 预计{planned_batch_count}批, 线程{planned_thread_count}个",
            field_progress={
                "phase": "layer2_generating",
                "phase_label": "Layer2词根生成中",
                "unique_root_count": len(unique_roots),
                "target_item_label": "词根",
                "total_items": len(unmatched_roots),
                "completed_items": 0,
                "batch_count": planned_batch_count,
                "completed_batches": 0,
                "thread_count": planned_thread_count,
                "layer1_count": len(layer1_fields),
                "layer2_count": len(layer2_fields),
                "layer3_count": len(layer3_fields)
            }
        )

        semantic_root_map = self.normalize_unmatched_roots_semantically(unmatched_roots)
        llm_translations = self.translate_unmatched_roots_via_llm(unmatched_roots, semantic_root_map)

        all_root_translations = {}
        for cn_root, trans in self.existing_root_map.items():
            selected_root = self._choose_root_by_priority(trans)
            if selected_root:
                all_root_translations[cn_root] = selected_root
        all_root_translations.update(llm_translations)
        self.root_translations = all_root_translations

        new_roots = {
            cn_root: en_root
            for cn_root, en_root in llm_translations.items()
            if cn_root not in self.existing_root_map
        }

        self._update_progress(
            5,
            8,
            f"生成字段名: 全局词根翻译完成，历史匹配{len(matched_roots)}个，新生成{len(llm_translations)}个",
            field_progress={
                "phase": "layer2_generating",
                "phase_label": "全局词根翻译完成",
                "unique_root_count": len(unique_roots),
                "target_item_label": "词根",
                "total_items": len(unique_roots),
                "completed_items": len(matched_roots) + len(llm_translations),
                "batch_count": planned_batch_count,
                "completed_batches": planned_batch_count,
                "thread_count": planned_thread_count,
                "layer1_count": len(layer1_fields),
                "layer2_count": len(layer2_fields),
                "layer3_count": len(layer3_fields)
            }
        )

        field_mapping = {}
        missing_fields = []
        for chinese_name, roots in field_tokenization.items():
            field_list = groups[chinese_name]
            field_type = field_list[0].suggested_type if field_list else 'VARCHAR(255)'

            if chinese_name in direct_overrides:
                override_name, override_type = direct_overrides[chinese_name]
                field_mapping[chinese_name] = (override_name, override_type)
                continue

            english_parts = []
            valid_parts = True
            for root in roots:
                selected_root = all_root_translations.get(root) or self._resolve_root_translation(root, llm_translations)
                if selected_root:
                    english_parts.append(selected_root)
                else:
                    valid_parts = False
                    logger.warning(f"【字段级处理】词根 '{root}' 没有符合当前模式的翻译结果")
                    break

            if valid_parts and english_parts:
                english_name = '_'.join(english_parts)
                if self._validate_identifier_by_priority(english_name, f"字段 '{chinese_name}'"):
                    field_mapping[chinese_name] = (english_name, field_type)
                else:
                    missing_fields.append(chinese_name)
            else:
                missing_fields.append(chinese_name)

        stats = {
            'layer1_count': len(layer1_fields),
            'layer2_count': len(layer2_fields),
            'layer3_count': len(layer3_fields),
            'matched_count': len(matched_roots),
            'unmatched_count': len(unmatched_roots),
            'total_fields': len(unique_roots),
            'new_roots_count': len(new_roots)
        }

        if missing_fields:
            logger.warning(f"【字段级处理】字段映射缺失 {len(missing_fields)} 个: {missing_fields[:20]}")

        logger.info(
            f"【字段级处理】混合模式处理完成: L1={stats['layer1_count']}, L2={stats['layer2_count']}, L3={stats['layer3_count']}, 新词根={stats['new_roots_count']}"
        )

        self._update_progress(
            5,
            8,
            f"字段名生成完成: 共生成{len(field_mapping)}/{len(field_tokenization)}个字段映射",
            matched_count=stats['matched_count'],
            unmatched_count=stats['unmatched_count'],
            total_fields=stats['total_fields'],
            field_progress={
                "phase": "field_names_completed",
                "phase_label": "字段名生成完成",
                "unique_root_count": len(unique_roots),
                "target_item_label": "字段映射",
                "total_items": len(field_tokenization),
                "completed_items": len(field_mapping),
                "batch_count": planned_batch_count,
                "completed_batches": planned_batch_count,
                "thread_count": planned_thread_count,
                "layer1_count": len(layer1_fields),
                "layer2_count": len(layer2_fields),
                "layer3_count": len(layer3_fields)
            }
        )

        return field_mapping, stats, new_roots
    def _convert_db_type(self, original_type, db_type):
        """将字段类型转换为目标数据库的类型"""
        type_mappings = {
            'mysql': {
                'VARCHAR': 'VARCHAR',
                'INT': 'INT',
                'BIGINT': 'BIGINT',
                'DECIMAL': 'DECIMAL',
                'DATETIME': 'DATETIME',
                'DATE': 'DATE',
                'TIMESTAMP': 'DATETIME',
                'TINYINT': 'TINYINT',
                'TEXT': 'TEXT',
                'LONGTEXT': 'LONGTEXT'
            },
            'postgresql': {
                'VARCHAR': 'VARCHAR',
                'INT': 'INTEGER',
                'BIGINT': 'BIGINT',
                'DECIMAL': 'DECIMAL',
                'DATETIME': 'TIMESTAMP',
                'DATE': 'DATE',
                'TIMESTAMP': 'TIMESTAMP',
                'TINYINT': 'SMALLINT',
                'TEXT': 'TEXT',
                'LONGTEXT': 'TEXT'
            },
            'oracle': {
                'VARCHAR': 'VARCHAR2',
                'INT': 'NUMBER(10)',
                'BIGINT': 'NUMBER(19)',
                'DECIMAL': 'NUMBER',
                'DATETIME': 'DATE',
                'DATE': 'DATE',
                'TIMESTAMP': 'TIMESTAMP',
                'TINYINT': 'NUMBER(3)',
                'TEXT': 'CLOB',
                'LONGTEXT': 'CLOB'
            }
        }
        
        db_type_lower = db_type.lower()
        mapping = type_mappings.get(db_type_lower, type_mappings['mysql'])
        
        # 提取类型名称（不包含括号内的内容）
        base_type = original_type.upper().split('(')[0]
        
        if base_type in mapping:
            # 保留括号内的参数
            if '(' in original_type:
                params = original_type[original_type.index('('):]
                target_type = mapping[base_type]
                if '(' in target_type:
                    return target_type
                return target_type + params
            return mapping[base_type]
        
        return original_type
    
    def _repair_field_name(self, chinese_name: str, candidate: str = "", used_names: set = None) -> str:
        used_names = used_names or set()
        candidate = self.name_normalizer.normalize_english_identifier(candidate)
        if candidate and self._is_safe_generated_identifier(candidate, f"field {chinese_name}") and candidate not in used_names:
            return candidate

        roots = self.field_tokenization.get(chinese_name, [])
        english_parts = []
        for root in roots:
            translated = self._resolve_root_translation(root, self.root_translations, allow_fallback=False)
            if translated and not self._has_hash_root_token(translated):
                english_parts.append(translated)
        rebuilt = self.name_normalizer.normalize_english_identifier("_".join(english_parts))
        if rebuilt and self._is_safe_generated_identifier(rebuilt, f"rebuilt field {chinese_name}") and rebuilt not in used_names:
            return rebuilt

        base = re.sub(r"(^|_)x[0-9a-f]{2,}(?=_|$)", "_", candidate).strip("_")
        base = self.name_normalizer.normalize_english_identifier(base)
        if normalize_priority(self.root_match_priority) == "abbr":
            base = self._abbr_from_english(base) or "fld"
        else:
            base = base or "field"
        if not re.match(r"^[a-z]", base):
            base = f"f_{base}" if normalize_priority(self.root_match_priority) == "abbr" else f"field_{base}"
        suffix = 2
        repaired = base
        while repaired in used_names or not self._is_safe_generated_identifier(repaired, f"dedup field {chinese_name}"):
            if suffix > 200:
                raise ValueError(f"字段 '{chinese_name}' 缺少符合当前词根模式的词根翻译，无法生成DDL")
            repaired = f"{base}_{suffix}"
            suffix += 1
        logger.warning(f"Field name repaired for '{chinese_name}': '{candidate}' -> '{repaired}'")
        return repaired

    def _deduplicate_table_field_names(self, field_rows: list) -> list:
        used_names = set()
        repaired_rows = []
        for chinese_name, english_name, final_type in field_rows:
            repaired_name = self._repair_field_name(chinese_name, english_name, used_names)
            used_names.add(repaired_name)
            repaired_rows.append((chinese_name, repaired_name, final_type))
        return repaired_rows

    def _repair_table_name(self, chinese_table_name: str, candidate: str = "", layer: str = "", theme_prefix: str = "") -> str:
        candidate = self.name_normalizer.normalize_english_identifier(candidate)
        if candidate:
            parts = [part for part in candidate.split("_") if part and part != layer]
            candidate = "_".join(parts)
        if candidate and self._is_safe_table_body(candidate, theme_prefix):
            return candidate

        fallback = self.translate_table_name(chinese_table_name, layer, theme_prefix)
        if fallback and fallback != candidate and self._is_safe_table_body(fallback, theme_prefix):
            logger.warning(f"Table name repaired for '{chinese_table_name}': '{candidate}' -> '{fallback}'")
            return fallback

        logger.warning(f"Table name repair failed for '{chinese_table_name}': '{candidate}'")
        raise ValueError(f"表 '{chinese_table_name}' 缺少符合规范的业务表名，无法生成DDL")

    def generate_ddl_for_table(self, table_name, table_info, field_mapping, 
                               db_type='mysql', root_match_priority='full',
                               standards_content=''):
        layer = table_info.get('layer', '')
        fields = table_info.get('fields', [])
        user_subject_domain = (table_info.get('user_specified_subject_domain') or '').strip().lower()
        theme_prefix = user_subject_domain or infer_theme_prefix(table_name, layer, standards_content=standards_content)
        
        # 调用LLM生成英文表名（如果有开发规范），否则使用fallback
        if standards_content.strip():
            english_table_name = self.generate_table_name(table_name, db_type, standards_content, layer)
        else:
            english_table_name = self.translate_table_name(table_name, layer, theme_prefix)
        english_table_name = normalize_table_business_domain(
            english_table_name,
            theme_prefix,
            standards_content,
            layer
        )
        english_table_name = self._repair_table_name(table_name, english_table_name, layer, theme_prefix)
        
        field_definitions = []
        field_comments = []
        field_rows = []
        
        db_type_lower = db_type.lower()
        
        for field_index, field in enumerate(fields, start=1):
            chinese_name = field['name']
            suggested_type = field['type']
            
            if chinese_name in field_mapping:
                english_name, llm_type = field_mapping[chinese_name]
                english_name = self._ensure_identifier_by_priority(english_name, f"field {chinese_name}")
                final_type = llm_type if llm_type else suggested_type
            else:
                english_name = self._compose_field_name_from_roots(chinese_name)
                final_type = suggested_type

            if not self._is_safe_generated_identifier(english_name, f"field {chinese_name}"):
                english_name = self._repair_field_name(chinese_name, english_name)
            
            final_type = self._convert_db_type(final_type, db_type)
            field_rows.append((chinese_name, english_name, final_type))

        field_rows = self._deduplicate_table_field_names(field_rows)
        for chinese_name, english_name, final_type in field_rows:
            if db_type_lower == 'mysql':
                field_definitions.append(f"    `{english_name}` {final_type} COMMENT '{chinese_name}'")
            else:
                field_definitions.append(f'    "{english_name}" {final_type}')
                field_comments.append((english_name, chinese_name))
        
        table_name_with_layer = f"{layer}_{english_table_name}" if layer else english_table_name
        
        if db_type_lower == 'mysql':
            ddl = f"""CREATE TABLE `{table_name_with_layer}` (
{',\n'.join(field_definitions)}
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='{table_name}';"""
        elif db_type_lower == 'postgresql':
            # PostgreSQL 使用 schema.table 格式
            schema_name = layer if layer else 'public'
            full_table_name = f'"{schema_name}"."{english_table_name}"'
            
            ddl = f"""CREATE TABLE {full_table_name} (
{',\n'.join(field_definitions)}
);

COMMENT ON TABLE {full_table_name} IS '{table_name}';"""
            
            # 添加字段注释
            for eng_name, chn_name in field_comments:
                ddl += f"\nCOMMENT ON COLUMN {full_table_name}.\"{eng_name}\" IS '{chn_name}';"
        elif db_type_lower == 'oracle':
            # Oracle 使用大写 SCHEMA.TABLE 格式
            schema_name = layer.upper() if layer else 'PUBLIC'
            full_table_name = f'"{schema_name}"."{english_table_name.upper()}"'
            field_definitions_upper = [fd.replace('"', '"').upper() for fd in field_definitions]
            
            ddl = f"""CREATE TABLE {full_table_name} (
{',\n'.join(field_definitions_upper)}
);

COMMENT ON TABLE {full_table_name} IS '{table_name}';"""
            
            # 添加字段注释
            for eng_name, chn_name in field_comments:
                ddl += f"\nCOMMENT ON COLUMN {full_table_name}.\"{eng_name.upper()}\" IS '{chn_name}';"
        else:
            ddl = f"""CREATE TABLE `{table_name_with_layer}` (
{',\n'.join(field_definitions)}
);"""
        
        return ddl

    def _semantic_table_name_from_keywords(self, chinese_table_name, layer=''):
        roots = self._match_table_name_from_root_library(chinese_table_name)
        return '_'.join(root for root in roots if root)
    
    def translate_table_name(self, chinese_table_name, layer='', theme_prefix=''):
        """
        Translate a Chinese table name into an English table name.
        """
        logger.info(f"【字段级处理】尝试转换表名: '{chinese_table_name}', layer: '{layer}'")

        layer_prefixes = ['dim', 'dwd', 'dws', 'ods', 'ads', 'input']
        theme_prefix = theme_prefix or infer_theme_prefix(chinese_table_name, layer, standards_content=self.standards)
        cleaned_name = self._strip_table_structure_suffixes(chinese_table_name)
        parts = self.name_normalizer.tokenize_chinese_name(cleaned_name, jieba.lcut)

        english_parts = []
        semantic_name = self._semantic_table_name_from_keywords(cleaned_name, layer)
        matched_from_root_library = bool(semantic_name)
        partial_table_match = False
        if semantic_name:
            english_parts = semantic_name.split('_')
            if len(english_parts) < 2:
                logger.warning(f"【字段级处理】表名只匹配到部分历史词根: '{chinese_table_name}' -> '{semantic_name}'")
                english_parts = []
                matched_from_root_library = False
                partial_table_match = True
        else:
            translated_parts = []
            all_parts_translated = bool(parts)
            for part in parts:
                translated = self._resolve_root_translation(part)
                if translated:
                    translated_parts.append(translated)
                elif not (part.isascii() and part.isalnum()):
                    all_parts_translated = False
                    break
            if all_parts_translated:
                english_parts = translated_parts

        if not english_parts and not partial_table_match:
            split_parts = self._try_split_and_match(cleaned_name)
            if split_parts:
                english_parts.extend(split_parts)

        if not english_parts and not partial_table_match:
            fallback_part = self._fallback_root_for_mode(cleaned_name or chinese_table_name)
            if fallback_part:
                english_parts = [fallback_part]

        english_name = '_'.join(english_parts)

        if '.' in english_name:
            english_name = english_name.split('.')[-1]

        if layer and layer.lower() in layer_prefixes:
            layer_lower = layer.lower()
            parts = english_name.split('_')
            if layer_lower in parts:
                new_parts = [p for p in parts if p != layer_lower]
                if new_parts:
                    english_name = '_'.join(new_parts)
                    if '_' not in english_name and len(english_name) <= self.abbr_max_len:
                        english_name += '_info'

        if theme_prefix and not matched_from_root_library:
            english_name = ensure_theme_prefix(english_name, theme_prefix)
        english_name = normalize_table_business_domain(
            english_name,
            theme_prefix,
            self.standards,
            layer
        )
        english_name = self.name_normalizer.normalize_english_identifier(english_name)
        if self._is_safe_table_body(english_name, theme_prefix):
            logger.info(f"【字段级处理】表名转换结果: '{chinese_table_name}' -> '{english_name}'")
            return english_name

        english_name = self._ensure_identifier_by_priority(english_name, f"fallback表名 {chinese_table_name}")
        if not self._is_safe_table_body(english_name, theme_prefix):
            logger.warning(f"【字段级处理】表名兜底转换失败: '{chinese_table_name}' -> '{english_name}'")
            raise ValueError(f"表 '{chinese_table_name}' 缺少符合规范的业务表名，无法生成DDL")
        logger.info(f"【字段级处理】表名兜底转换结果: '{chinese_table_name}' -> '{english_name}'")
        return english_name
    def build_field_mapping(self, file_content):
        tables, all_fields = self.extract_fields_from_excel(file_content)
        groups = self.group_fields_by_chinese(all_fields)
        
        # 使用新的词根级处理流程
        field_mapping, stats, root_translations = self.process_fields_root_level(groups)
        
        self.field_mapping = field_mapping
        self.root_translations = root_translations
        logger.info(f"【字段级处理】字段映射构建完成，共 {len(field_mapping)} 个字段")
        return tables, field_mapping, stats, root_translations
    
    def _generate_single_ddl(self, table_name, table_info, field_mapping, db_type, root_match_priority, standards_content):
        """生成单张表的 DDL（供多线程调用）"""
        return table_name, self.generate_ddl_for_table(
            table_name, table_info, field_mapping, db_type, root_match_priority, standards_content
        )

    def generate_all_ddl(self, tables, field_mapping, db_type='mysql', root_match_priority='full', standards_content=''):
        logger.info("【字段级处理】组装DDL（多线程并行）")
        self._update_progress(6, 8, "组装DDL...")
        
        results = {}
        total_tables = len(tables)
        
        # 多线程并行生成各表的 DDL
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {}
            for table_name, table_info in tables.items():
                future = executor.submit(
                    self._generate_single_ddl,
                    table_name, table_info, field_mapping, db_type, root_match_priority, standards_content
                )
                futures[future] = table_name
            
            completed_tables = 0
            for future in as_completed(futures):
                try:
                    table_name, ddl = future.result()
                    results[table_name] = ddl
                    completed_tables += 1
                    logger.info(f"【字段级处理】表 {table_name} DDL生成完成 [{completed_tables}/{total_tables}]")
                    self._update_progress(6, 8, f"组装DDL [{completed_tables}/{total_tables}]")
                except Exception as e:
                    failed_table = futures[future]
                    logger.error(f"【字段级处理】表 {failed_table} DDL生成失败: {e}")
                    results[failed_table] = f"-- 生成失败: {e}"
        
        self._update_progress(6, 8, "DDL组装完成")
        success_count = sum(1 for ddl in results.values() if not ddl.startswith("-- 生成失败"))
        logger.info(f"【字段级处理】所有表DDL生成完成，成功 {success_count} 张，失败 {len(results) - success_count} 张")
        return results




