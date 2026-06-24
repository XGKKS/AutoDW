#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从Excel文件读取字段并进行词根匹配测试
"""
import json
import os
import sys
import io

# 添加 app 目录到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'app'))

from app.processors.field_processor import FieldProcessor
import openpyxl


def load_word_roots():
    """加载词根文件"""
    word_roots_path = os.path.join(os.path.dirname(__file__), 'word_roots.json')
    if os.path.exists(word_roots_path):
        try:
            with open(word_roots_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载词根文件失败: {e}")
            return []
    else:
        print(f"词根文件不存在: {word_roots_path}")
        return []


def extract_fields_from_excel(excel_path):
    """从Excel文件提取字段"""
    print(f"\n从Excel文件提取字段: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path)
    all_fields = []
    tables = {}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = None
        
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num == 1:
                headers = [str(cell).strip() if cell else '' for cell in row]
                print(f"  表头: {headers}")
                continue
            
            if not row or all(not cell for cell in row):
                continue
            
            row_data = {headers[i]: str(row[i]).strip() if row[i] else '' for i in range(len(headers))}
            
            table_name = row_data.get('表名', '')
            table_layer = row_data.get('表分层', '').lower()
            field_name = row_data.get('字段名', '')
            field_type = row_data.get('推荐字段类型', 'VARCHAR(255)')
            
            if not table_name or not field_name:
                continue
            
            all_fields.append({
                'table_name': table_name,
                'chinese_name': field_name,
                'suggested_type': field_type
            })
            
            if table_name not in tables:
                tables[table_name] = {
                    'layer': table_layer,
                    'fields': []
                }
            tables[table_name]['fields'].append({
                'name': field_name,
                'type': field_type
            })
    
    print(f"  共提取 {len(all_fields)} 个字段，来自 {len(tables)} 张表")
    return all_fields, tables


def group_fields_by_chinese(fields):
    """按中文名称分组"""
    groups = {}
    for field in fields:
        key = field['chinese_name'].strip()
        if key not in groups:
            groups[key] = []
        groups[key].append(field)
    return groups


def test_excel_root_matching():
    """测试Excel字段的词根匹配"""
    print("=" * 80)
    print("Excel文件字段词根匹配测试")
    print("=" * 80)
    
    # 1. 加载词根
    print("\n1. 加载词根文件...")
    word_roots = load_word_roots()
    print(f"   共加载 {len(word_roots)} 条词根")
    
    # 2. 创建 FieldProcessor 实例
    print("\n2. 创建 FieldProcessor 实例...")
    processor = FieldProcessor(
        api_key="test_key",
        api_url="http://localhost",
        model="test_model",
        word_roots=word_roots,
        root_match_priority="full"
    )
    
    # 3. 从Excel提取字段
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'DDL建表测试用例4表.xlsx')
    if not os.path.exists(excel_path):
        excel_path = os.path.join(os.path.dirname(__file__), 'DDL建表测试用例4表.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"\nExcel文件不存在: {excel_path}")
        print("请确保 'DDL建表测试用例4表.xlsx' 文件在正确位置")
        return
    
    all_fields, tables = extract_fields_from_excel(excel_path)
    
    # 4. 按中文名称分组
    print("\n3. 按中文名称分组...")
    groups = group_fields_by_chinese(all_fields)
    print(f"   共 {len(groups)} 个不同的中文字段名")
    
    # 打印所有不同的字段名
    print(f"\n   所有不同的字段名 ({len(groups)} 个):")
    for i, (name, field_list) in enumerate(groups.items(), 1):
        print(f"   {i:2d}. {name} (出现 {len(field_list)} 次)")
    
    # 5. 词根匹配测试
    print("\n4. 词根匹配测试...")
    print("-" * 80)
    
    matched_fields = {}
    unmatched_fields = {}
    
    for chinese_name, field_list in groups.items():
        result = processor.match_existing_root(chinese_name)
        if result:
            english, field_type = result
            matched_fields[chinese_name] = (english, field_type)
        else:
            unmatched_fields[chinese_name] = field_list
    
    print(f"\n匹配结果:")
    print(f"  成功匹配: {len(matched_fields)} 个字段")
    print(f"  未匹配:   {len(unmatched_fields)} 个字段")
    print(f"  匹配率:   {len(matched_fields)/len(groups)*100:.1f}%")
    
    # 6. 显示匹配成功的字段
    if matched_fields:
        print(f"\n5. 匹配成功的字段 ({len(matched_fields)} 个):")
        print("-" * 80)
        for i, (name, (english, field_type)) in enumerate(matched_fields.items(), 1):
            print(f"   {i:2d}. {name:20s} -> {english:20s} ({field_type})")
    
    # 7. 显示未匹配的字段
    if unmatched_fields:
        print(f"\n6. 未匹配的字段 ({len(unmatched_fields)} 个):")
        print("-" * 80)
        for i, (name, field_list) in enumerate(unmatched_fields.items(), 1):
            print(f"   {i:2d}. {name:20s}")
    
    # 8. 尝试模糊匹配
    if unmatched_fields:
        print(f"\n7. 尝试模糊匹配未匹配的字段...")
        print("-" * 80)
        
        all_roots = list(processor.existing_root_map.keys())
        
        for chinese_name, field_list in unmatched_fields.items():
            print(f"\n   '{chinese_name}' 的分析:")
            
            # 检查是否包含已知词根
            found_parts = []
            for char in chinese_name:
                if char in all_roots:
                    info = processor.existing_root_map[char]
                    print(f"     - 字符 '{char}' -> '{info['full']}'")
                    found_parts.append((char, info['full']))
            
            if found_parts:
                print(f"     建议: 可以拆分为 {found_parts}")
            else:
                print(f"     建议: 需要添加新词根或使用LLM生成")
    
    print("\n" + "=" * 80)
    print("测试完成!")
    print("=" * 80)


if __name__ == "__main__":
    test_excel_root_matching()
